"""
API endpoints for admin panel features: Challenges, Referrals, Loyalty, Notifications, Gallery
"""
from fastapi import APIRouter, Request, Cookie, HTTPException, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional, List, Set
from db.connection import get_db_connection
from utils.utils import require_auth
from utils.logger import log_info, log_error
from datetime import datetime, timedelta
import os
import uuid
import json

router = APIRouter(tags=["Admin Features"])

# ============================================================================
# DASHBOARD STATS
# ============================================================================

@router.get("/admin/stats")
async def get_admin_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику для админ дашборда"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total clients (not users)
        c.execute("SELECT COUNT(*) FROM clients")
        total_users = c.fetchone()[0]

        # Active challenges
        c.execute("SELECT COUNT(*) FROM active_challenges WHERE is_active = TRUE")
        active_challenges = c.fetchone()[0]

        # Total loyalty points issued
        c.execute("SELECT COALESCE(SUM(points), 0) FROM loyalty_transactions WHERE points > 0")
        total_loyalty_points = c.fetchone()[0]

        # Total referrals
        c.execute("SELECT COUNT(*) FROM client_referrals")
        total_referrals = c.fetchone()[0]

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_users": total_users,
                "active_challenges": active_challenges,
                "total_loyalty_points": int(total_loyalty_points),
                "total_referrals": total_referrals
            }
        }
    except Exception as e:
        log_error(f"Error getting admin stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

# ============================================================================
# CHALLENGES API
# ============================================================================

def _get_table_columns(cursor, table_name: str) -> Set[str]:
    cursor.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = %s
    """, (table_name,))
    return {row[0] for row in cursor.fetchall()}


def _quote_sql_identifier(identifier: str) -> str:
    escaped_identifier = identifier.replace('"', '""')
    return f'"{escaped_identifier}"'


def _get_template_value_expression(template_columns: Set[str], base_column: str) -> str:
    if base_column in template_columns:
        return _quote_sql_identifier(base_column)

    prefixed_columns = sorted([
        column_name
        for column_name in template_columns
        if column_name.startswith(f"{base_column}_")
    ])
    if len(prefixed_columns) == 0:
        return "''::text"

    prefixed_sql = ", ".join([
        f"NULLIF({_quote_sql_identifier(column_name)}, '')"
        for column_name in prefixed_columns
    ])
    return f"COALESCE({prefixed_sql}, '')"


def _get_template_storage_columns(template_columns: Set[str], base_column: str) -> List[str]:
    if base_column in template_columns:
        return [base_column]
    return sorted([
        column_name
        for column_name in template_columns
        if column_name.startswith(f"{base_column}_")
    ])


def _normalize_challenge_type(raw_type: Optional[str]) -> str:
    if raw_type == "spend":
        return "spending"
    if isinstance(raw_type, str) and len(raw_type) > 0:
        return raw_type
    return "visits"

@router.get("/admin/challenges")
async def get_admin_challenges(session_token: Optional[str] = Cookie(None)):
    """Получить все челленджи для админки"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        challenge_columns = _get_table_columns(c, "active_challenges")

        challenge_type_expr = "'visits'"
        if "challenge_type" in challenge_columns:
            challenge_type_expr = "c.challenge_type"
        elif "reward_type" in challenge_columns:
            challenge_type_expr = "c.reward_type"

        target_value_expr = "0"
        if "target_value" in challenge_columns:
            target_value_expr = "COALESCE(c.target_value, 0)"

        reward_points_expr = "0"
        if "bonus_points" in challenge_columns:
            reward_points_expr = "COALESCE(c.bonus_points, 0)"
        elif "points_reward" in challenge_columns:
            reward_points_expr = "COALESCE(c.points_reward, 0)"

        start_date_expr = "c.created_at"
        if "start_date" in challenge_columns:
            start_date_expr = "COALESCE(c.start_date::timestamp, c.created_at)"

        end_date_expr = "(c.created_at + INTERVAL '30 days')"
        if "end_date" in challenge_columns:
            end_date_expr = "COALESCE(c.end_date::timestamp, c.created_at + INTERVAL '30 days')"

        c.execute(f"""
            SELECT
                c.id,
                c.title,
                c.description,
                {challenge_type_expr} as type,
                {target_value_expr} as target_value,
                {reward_points_expr} as reward_points,
                {start_date_expr} as start_date,
                {end_date_expr} as end_date,
                c.is_active,
                COUNT(DISTINCT cp.client_id) as participants,
                COUNT(DISTINCT CASE WHEN cp.is_completed THEN cp.client_id END) as completions
            FROM active_challenges c
            LEFT JOIN challenge_progress cp ON c.id = cp.challenge_id
            GROUP BY c.id
            ORDER BY c.created_at DESC
        """)

        challenges = []
        for row in c.fetchall():
            is_active = bool(row[8])
            status = "active" if is_active else "completed"

            start_date = row[6]
            end_date = row[7]
            participants = int(row[9] or 0)
            completions = int(row[10] or 0)
            completion_rate = int((completions * 100 / participants) if participants > 0 else 0)

            if start_date is None:
                start_date = datetime.now()
            if end_date is None:
                end_date = datetime.now() + timedelta(days=30)

            if hasattr(start_date, "date"):
                now_date = datetime.now().date()
                start_date_only = start_date.date()
                end_date_only = end_date.date() if hasattr(end_date, "date") else now_date
                if is_active and now_date < start_date_only:
                    status = "upcoming"
                elif is_active and now_date > end_date_only:
                    status = "completed"

            challenges.append({
                "id": str(row[0]),
                "title": row[1] or "Challenge",
                "description": row[2] or "",
                "type": _normalize_challenge_type(row[3]),
                "target_value": row[4] or 0,
                "reward_points": row[5] or 0,
                "start_date": start_date.isoformat() if hasattr(start_date, "isoformat") else datetime.now().isoformat(),
                "end_date": end_date.isoformat() if hasattr(end_date, "isoformat") else (datetime.now() + timedelta(days=30)).isoformat(),
                "status": status,
                "participants": participants,
                "participants_count": participants,
                "completions": completions,
                "completion_rate": completion_rate,
                "is_active": is_active
            })

        conn.close()
        return {"success": True, "challenges": challenges}
    except Exception as e:
        log_error(f"Error getting admin challenges: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/challenges/stats")
async def get_admin_challenges_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику челленджей"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("SELECT COUNT(*) FROM active_challenges WHERE is_active = TRUE")
        active_challenges = int(c.fetchone()[0] or 0)

        c.execute("SELECT COUNT(DISTINCT client_id) FROM challenge_progress")
        total_participants = int(c.fetchone()[0] or 0)

        c.execute("""
            SELECT COUNT(*)
            FROM challenge_progress
            WHERE is_completed = TRUE
              AND completed_at IS NOT NULL
              AND DATE(completed_at) = CURRENT_DATE
        """)
        completed_today = int(c.fetchone()[0] or 0)

        conn.close()
        return {
            "success": True,
            "stats": {
                "active_challenges": active_challenges,
                "total_participants": total_participants,
                "completed_today": completed_today
            }
        }
    except Exception as e:
        log_error(f"Error getting challenge stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/challenges")
async def create_admin_challenge(request: Request, session_token: Optional[str] = Cookie(None)):
    """Создать новый челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        challenge_columns = _get_table_columns(c, "active_challenges")

        fields: List[str] = ["title", "description", "is_active"]
        values: List[object] = [
            data.get("title", ""),
            data.get("description", ""),
            True
        ]

        if "challenge_id" in challenge_columns:
            fields.append("challenge_id")
            values.append(f"challenge_{uuid.uuid4().hex[:12]}")

        normalized_type = _normalize_challenge_type(data.get("type", "visits"))
        if normalized_type == "services":
            raise HTTPException(status_code=400, detail="Services type challenges are not supported yet")

        if "challenge_type" in challenge_columns:
            fields.append("challenge_type")
            values.append(normalized_type)
        elif "reward_type" in challenge_columns:
            fields.append("reward_type")
            values.append(normalized_type)

        if "target_value" in challenge_columns:
            fields.append("target_value")
            values.append(data.get("target_value", 0))

        if "bonus_points" in challenge_columns:
            fields.append("bonus_points")
            values.append(data.get("reward_points", 0))
        elif "points_reward" in challenge_columns:
            fields.append("points_reward")
            values.append(data.get("reward_points", 0))

        if "start_date" in challenge_columns:
            fields.append("start_date")
            values.append(data.get("start_date"))
        if "end_date" in challenge_columns:
            fields.append("end_date")
            values.append(data.get("end_date"))

        placeholders = ", ".join(["%s"] * len(values))
        sql = f"""
            INSERT INTO active_challenges ({", ".join(fields)})
            VALUES ({placeholders})
            RETURNING id
        """

        c.execute(sql, tuple(values))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {"success": True, "id": new_id}
    except Exception as e:
        log_error(f"Error creating challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/challenges/{challenge_id}")
async def update_admin_challenge(challenge_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        challenge_columns = _get_table_columns(c, "active_challenges")

        updates: List[str] = [
            "title = %s",
            "description = %s"
        ]
        values: List[object] = [
            data.get("title", ""),
            data.get("description", "")
        ]

        normalized_type = _normalize_challenge_type(data.get("type", "visits"))
        if normalized_type == "services":
            raise HTTPException(status_code=400, detail="Services type challenges are not supported yet")

        if "challenge_type" in challenge_columns:
            updates.append("challenge_type = %s")
            values.append(normalized_type)
        elif "reward_type" in challenge_columns:
            updates.append("reward_type = %s")
            values.append(normalized_type)

        if "target_value" in challenge_columns:
            updates.append("target_value = %s")
            values.append(data.get("target_value", 0))

        if "bonus_points" in challenge_columns:
            updates.append("bonus_points = %s")
            values.append(data.get("reward_points", 0))
        elif "points_reward" in challenge_columns:
            updates.append("points_reward = %s")
            values.append(data.get("reward_points", 0))

        if "start_date" in challenge_columns:
            updates.append("start_date = %s")
            values.append(data.get("start_date"))
        if "end_date" in challenge_columns:
            updates.append("end_date = %s")
            values.append(data.get("end_date"))
        if "updated_at" in challenge_columns:
            updates.append("updated_at = CURRENT_TIMESTAMP")

        values.append(challenge_id)
        sql = f"UPDATE active_challenges SET {', '.join(updates)} WHERE id = %s"
        c.execute(sql, tuple(values))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/challenges/{challenge_id}")
async def delete_admin_challenge(challenge_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Удаляем прогресс челленджа
        c.execute("DELETE FROM challenge_progress WHERE challenge_id = %s", (challenge_id,))
        # Удаляем сам челлендж
        c.execute("DELETE FROM active_challenges WHERE id = %s", (challenge_id,))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/challenges/{challenge_id}/check-progress")
async def check_challenge_progress(
    challenge_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Проверить и обновить прогресс челленджа для всех участников"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        challenge_columns = _get_table_columns(c, "active_challenges")

        challenge_type_col = None
        if "challenge_type" in challenge_columns:
            challenge_type_col = "challenge_type"
        elif "reward_type" in challenge_columns:
            challenge_type_col = "reward_type"

        target_value_col = "target_value" if "target_value" in challenge_columns else None
        reward_points_col = None
        if "bonus_points" in challenge_columns:
            reward_points_col = "bonus_points"
        elif "points_reward" in challenge_columns:
            reward_points_col = "points_reward"
        start_date_col = "start_date" if "start_date" in challenge_columns else None
        end_date_col = "end_date" if "end_date" in challenge_columns else None

        select_parts = [
            f"{challenge_type_col} AS challenge_type" if challenge_type_col else "'visits' AS challenge_type",
            f"{target_value_col} AS target_value" if target_value_col else "0 AS target_value",
            f"{reward_points_col} AS reward_points" if reward_points_col else "0 AS reward_points",
            f"{start_date_col} AS start_date" if start_date_col else "NULL::timestamp AS start_date",
            f"{end_date_col} AS end_date" if end_date_col else "NULL::timestamp AS end_date"
        ]

        c.execute(f"""
            SELECT {", ".join(select_parts)}
            FROM active_challenges
            WHERE id = %s AND is_active = TRUE
        """, (challenge_id,))

        challenge = c.fetchone()
        if not challenge:
            raise HTTPException(status_code=404, detail="Challenge not found or inactive")

        challenge_type, target_value, bonus_points, start_date, end_date = challenge
        challenge_type = _normalize_challenge_type(challenge_type)
        if target_value is None:
            target_value = 0

        # Проверяем прогресс в зависимости от типа челленджа
        if challenge_type == 'visits':
            # Считаем количество визитов (завершенных броней) для каждого клиента
            c.execute("""
                SELECT b.client_id, COUNT(*) as visit_count
                FROM bookings b
                WHERE b.status = 'completed'
                  AND (%s IS NULL OR b.booking_date >= %s)
                  AND (%s IS NULL OR b.booking_date <= %s)
                GROUP BY b.client_id
                HAVING COUNT(*) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        elif challenge_type in ('spending', 'spend'):
            # Считаем общую сумму трат для каждого клиента
            c.execute("""
                SELECT b.client_id, COALESCE(SUM(b.total_price), 0) as total_spent
                FROM bookings b
                WHERE b.status = 'completed'
                  AND (%s IS NULL OR b.booking_date >= %s)
                  AND (%s IS NULL OR b.booking_date <= %s)
                GROUP BY b.client_id
                HAVING COALESCE(SUM(b.total_price), 0) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        elif challenge_type == 'referrals':
            # Считаем количество успешных рефералов для каждого клиента
            c.execute("""
                SELECT r.referrer_id as client_id, COUNT(*) as referral_count
                FROM client_referrals r
                WHERE r.status = 'completed'
                  AND (%s IS NULL OR r.created_at >= %s)
                  AND (%s IS NULL OR r.created_at <= %s)
                GROUP BY r.referrer_id
                HAVING COUNT(*) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        else:
            # Для типа 'services' нужна дополнительная логика
            conn.close()
            return {"success": False, "message": "Services type challenges not yet implemented"}

        completed_clients = c.fetchall()
        updated_count = 0

        for client_row in completed_clients:
            client_id = client_row[0]
            current_value = client_row[1]

            # Проверяем, есть ли уже запись прогресса
            c.execute("""
                SELECT id, is_completed
                FROM challenge_progress
                WHERE challenge_id = %s AND client_id = %s
            """, (challenge_id, client_id))

            progress = c.fetchone()

            if progress:
                progress_id, is_completed = progress
                if not is_completed:
                    # Обновляем прогресс и помечаем как завершенный
                    c.execute("""
                        UPDATE challenge_progress
                        SET current_value = %s,
                            is_completed = TRUE,
                            completed_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (current_value, progress_id))

                    # Начисляем бонусные баллы клиенту
                    c.execute("""
                        UPDATE clients
                        SET loyalty_points = COALESCE(loyalty_points, 0) + %s
                        WHERE instagram_id = %s
                    """, (bonus_points, client_id))

                    # Создаем транзакцию лояльности
                    c.execute("""
                        INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
                        VALUES (%s, %s, 'earn', %s)
                    """, (client_id, bonus_points, f"Challenge completed: {challenge_type}"))

                    updated_count += 1
            else:
                # Создаем новую запись прогресса как завершенную
                c.execute("""
                    INSERT INTO challenge_progress (
                        challenge_id, client_id, current_value,
                        is_completed, completed_at
                    ) VALUES (%s, %s, %s, TRUE, CURRENT_TIMESTAMP)
                """, (challenge_id, client_id, current_value))

                # Начисляем бонусные баллы
                c.execute("""
                    UPDATE clients
                    SET loyalty_points = COALESCE(loyalty_points, 0) + %s
                    WHERE instagram_id = %s
                """, (bonus_points, client_id))

                # Создаем транзакцию
                c.execute("""
                    INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
                    VALUES (%s, %s, 'earn', %s)
                """, (client_id, bonus_points, f"Challenge completed: {challenge_type}"))

                updated_count += 1

        conn.commit()
        conn.close()

        return {
            "success": True,
            "updated_count": updated_count,
            "message": f"Updated {updated_count} clients who completed the challenge"
        }

    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Error checking challenge progress: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# REFERRAL PROGRAM API
# ============================================================================

@router.get("/admin/referrals")
async def get_admin_referrals(
    session_token: Optional[str] = Cookie(None),
    status: Optional[str] = None,
    sort_by: Optional[str] = "date",
    order: Optional[str] = "desc"
):
    """Получить все рефералы с фильтрацией и сортировкой"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Построение WHERE клаузы
        where_clause = ""
        params = []
        if status:
            where_clause = "WHERE r.status = %s"
            params.append(status)

        # Определение сортировки
        sort_column = {
            "date": "r.created_at",
            "points": "r.points_awarded",
            "status": "r.status"
        }.get(sort_by, "r.created_at")

        sort_order = "DESC" if order == "desc" else "ASC"

        query = f"""
            SELECT
                r.id,
                c1.name as referrer_name,
                c1.email as referrer_email,
                c2.name as referred_name,
                c2.email as referred_email,
                r.status,
                r.points_awarded,
                r.created_at
            FROM client_referrals r
            LEFT JOIN clients c1 ON r.referrer_id = c1.instagram_id
            LEFT JOIN clients c2 ON r.referred_id = c2.instagram_id
            {where_clause}
            ORDER BY {sort_column} {sort_order}
        """

        c.execute(query, params)

        referrals = []
        for row in c.fetchall():
            referrals.append({
                "id": str(row[0]),
                "referrer_name": row[1] or "Unknown",
                "referrer_email": row[2] or "",
                "referred_name": row[3] or "Unknown",
                "referred_email": row[4] or "",
                "status": row[5] or "pending",
                "points_awarded": row[6] or 0,
                "created_at": row[7].isoformat() if row[7] else datetime.now().isoformat()
            })

        conn.close()
        return {"success": True, "referrals": referrals}
    except Exception as e:
        log_error(f"Error getting referrals: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/referrals/settings")
async def get_referral_settings(session_token: Optional[str] = Cookie(None)):
    """Получить настройки реферальной программы"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT setting_value FROM settings WHERE setting_key = 'referral_settings'
        """)

        result = c.fetchone()
        if result:
            import json
            settings = json.loads(result[0])
        else:
            settings = {
                "referrer_bonus": 500,
                "referred_bonus": 200,
                "min_purchase_amount": 0
            }

        conn.close()
        return {"success": True, "settings": settings}
    except Exception as e:
        log_error(f"Error getting referral settings: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/referrals/settings")
async def update_referral_settings(request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить настройки реферальной программы"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        import json
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO settings (setting_key, setting_value)
            VALUES ('referral_settings', %s)
            ON CONFLICT (setting_key) DO UPDATE SET setting_value = EXCLUDED.setting_value
        """, (json.dumps(data),))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating referral settings: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/referrals/stats")
async def get_referral_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику рефералов"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total referrals
        c.execute("SELECT COUNT(*) FROM client_referrals")
        total_referrals = c.fetchone()[0] or 0

        # Completed referrals
        c.execute("SELECT COUNT(*) FROM client_referrals WHERE status = 'completed'")
        completed_referrals = c.fetchone()[0] or 0

        # Points distributed
        c.execute("SELECT COALESCE(SUM(points_awarded), 0) FROM client_referrals WHERE status = 'completed'")
        points_distributed = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_referrals": total_referrals,
                "completed_referrals": completed_referrals,
                "points_distributed": points_distributed
            }
        }
    except Exception as e:
        log_error(f"Error getting referral stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# LOYALTY MANAGEMENT API
# ============================================================================

@router.get("/admin/loyalty/tiers")
async def get_loyalty_tiers(session_token: Optional[str] = Cookie(None)):
    """Получить уровни лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT id, name, min_points, discount, COALESCE(is_active, TRUE), color
            FROM loyalty_tiers
            ORDER BY min_points ASC
        """)

        tiers = []
        for row in c.fetchall():
            tiers.append({
                "id": str(row[0]),
                "name": row[1],
                "min_points": row[2] or 0,
                "discount": row[3] or 0,
                "is_active": bool(row[4]),
                "color": row[5] or "#CD7F32"
            })

        # Если тиров нет, создаем дефолтные
        if not tiers:
            default_tiers = [
                ('Bronze', 0, 0, True, '#CD7F32'),
                ('Silver', 1000, 0, True, '#C0C0C0'),
                ('Gold', 5000, 0, True, '#FFD700'),
                ('Platinum', 10000, 0, True, '#E5E4E2')
            ]

            for tier in default_tiers:
                c.execute("""
                    INSERT INTO loyalty_tiers (name, min_points, discount, is_active, color)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id, name, min_points, discount, is_active, color
                """, tier)
                row = c.fetchone()
                tiers.append({
                    "id": str(row[0]),
                    "name": row[1],
                    "min_points": row[2],
                    "discount": row[3],
                    "is_active": bool(row[4]),
                    "color": row[5]
                })

            conn.commit()

        conn.close()
        return {"success": True, "tiers": tiers}
    except Exception as e:
        log_error(f"Error getting loyalty tiers: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/loyalty/tiers")
async def create_loyalty_tier(request: Request, session_token: Optional[str] = Cookie(None)):
    """Создать уровень лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        tier_name = str(data.get("name") or "").strip()
        min_points = data.get("min_points")
        discount = data.get("discount")
        color = data.get("color")
        is_active_raw = data.get("is_active", True)

        if len(tier_name) == 0:
            return JSONResponse({"error": "Tier name is required"}, status_code=400)

        min_points_value = int(min_points) if min_points is not None else 0
        discount_value = float(discount) if discount is not None else 0
        is_active_value = bool(is_active_raw) if isinstance(is_active_raw, bool) else str(is_active_raw).strip().lower() in ["1", "true", "yes", "on"]
        color_value = str(color or '#CD7F32').strip()

        if min_points_value < 0:
            min_points_value = 0

        if discount_value < 0:
            discount_value = 0

        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO loyalty_tiers (name, min_points, discount, is_active, color)
            VALUES (%s, %s, %s, %s, %s)
            RETURNING id, name, min_points, discount, is_active, color
        """, (
            tier_name,
            min_points_value,
            discount_value,
            is_active_value,
            color_value
        ))

        row = c.fetchone()
        conn.commit()
        conn.close()

        created_tier = {
            "id": str(row[0]),
            "name": row[1],
            "min_points": row[2] or 0,
            "discount": row[3] or 0,
            "is_active": bool(row[4]),
            "color": row[5] or "#CD7F32"
        }
        return {"success": True, "tier": created_tier}
    except Exception as e:
        log_error(f"Error creating loyalty tier: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/loyalty/tiers/{tier_id}")
async def update_loyalty_tier(tier_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить уровень лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        tier_name = str(data.get("name") or "").strip()
        min_points = data.get("min_points")
        discount = data.get("discount")
        color = data.get("color")
        is_active_raw = data.get("is_active", True)

        if len(tier_name) == 0:
            return JSONResponse({"error": "Tier name is required"}, status_code=400)

        min_points_value = int(min_points) if min_points is not None else 0
        discount_value = float(discount) if discount is not None else 0
        is_active_value = bool(is_active_raw) if isinstance(is_active_raw, bool) else str(is_active_raw).strip().lower() in ["1", "true", "yes", "on"]
        color_value = str(color or '#CD7F32').strip()

        if min_points_value < 0:
            min_points_value = 0

        if discount_value < 0:
            discount_value = 0

        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            UPDATE loyalty_tiers SET
                name = %s,
                min_points = %s,
                discount = %s,
                is_active = %s,
                color = %s
            WHERE id = %s
        """, (
            tier_name,
            min_points_value,
            discount_value,
            is_active_value,
            color_value,
            tier_id
        ))

        if c.rowcount == 0:
            conn.close()
            return JSONResponse({"error": "Tier not found"}, status_code=404)

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating loyalty tier: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/loyalty/tiers/{tier_id}")
async def delete_loyalty_tier(tier_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить уровень лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("SELECT COUNT(*) FROM loyalty_tiers")
        tiers_count = c.fetchone()[0] or 0
        if tiers_count <= 1:
            conn.close()
            return JSONResponse({"error": "At least one tier must remain"}, status_code=400)

        c.execute("DELETE FROM loyalty_tiers WHERE id = %s", (tier_id,))
        if c.rowcount == 0:
            conn.close()
            return JSONResponse({"error": "Tier not found"}, status_code=404)

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting loyalty tier: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/loyalty/transactions")
async def get_loyalty_transactions(session_token: Optional[str] = Cookie(None)):
    """Получить историю транзакций лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT
                lt.id,
                c.name as client_name,
                c.email as client_email,
                lt.points,
                lt.transaction_type as type,
                lt.reason,
                lt.created_at
            FROM loyalty_transactions lt
            LEFT JOIN clients c ON lt.client_id = c.instagram_id
            ORDER BY lt.created_at DESC
            LIMIT 100
        """)

        transactions = []
        for row in c.fetchall():
            transactions.append({
                "id": str(row[0]),
                "client_name": row[1] or "Unknown",
                "client_email": row[2] or "",
                "points": row[3] or 0,
                "type": row[4] or "adjust",
                "reason": row[5] or "",
                "created_at": row[6].isoformat() if row[6] else datetime.now().isoformat()
            })

        conn.close()
        return {"success": True, "transactions": transactions}
    except Exception as e:
        log_error(f"Error getting loyalty transactions: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/loyalty/transactions/{transaction_id}")
async def delete_loyalty_transaction(transaction_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить транзакцию лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT client_id, points
            FROM loyalty_transactions
            WHERE id = %s
        """, (transaction_id,))
        row = c.fetchone()
        if row is None:
            conn.close()
            return JSONResponse({"error": "Transaction not found"}, status_code=404)

        client_id = row[0]
        points = row[1] if row[1] is not None else 0

        c.execute("DELETE FROM loyalty_transactions WHERE id = %s", (transaction_id,))

        if client_id:
            c.execute("""
                UPDATE clients
                SET loyalty_points = GREATEST(0, COALESCE(loyalty_points, 0) - %s)
                WHERE instagram_id = %s
            """, (points, client_id))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting loyalty transaction: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/loyalty/adjust-points")
async def adjust_loyalty_points(request: Request, session_token: Optional[str] = Cookie(None)):
    """Скорректировать баллы лояльности клиента"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        # Найти клиента по email
        c.execute("SELECT instagram_id, loyalty_points FROM clients WHERE email = %s", (data.get("client_email"),))
        result = c.fetchone()

        if not result:
            raise HTTPException(status_code=404, detail="Client not found")

        client_id = result[0]
        current_points = result[1] or 0
        points = data.get("points", 0)
        new_points = current_points + points

        # Обновить баллы клиента
        c.execute("UPDATE clients SET loyalty_points = %s WHERE instagram_id = %s", (new_points, client_id))

        # Создать транзакцию
        c.execute("""
            INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
            VALUES (%s, %s, %s, %s)
        """, (client_id, points, 'adjust', data.get("reason", "Manual adjustment")))

        conn.commit()
        conn.close()

        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Error adjusting loyalty points: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/loyalty/stats")
async def get_loyalty_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику программы лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total points issued
        c.execute("SELECT COALESCE(SUM(points), 0) FROM loyalty_transactions WHERE points > 0")
        total_points_issued = c.fetchone()[0] or 0

        # Points redeemed
        c.execute("SELECT COALESCE(SUM(ABS(points)), 0) FROM loyalty_transactions WHERE points < 0")
        points_redeemed = c.fetchone()[0] or 0

        # Active members (clients with points > 0)
        c.execute("SELECT COUNT(*) FROM clients WHERE loyalty_points > 0")
        active_members = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_points_issued": total_points_issued,
                "points_redeemed": points_redeemed,
                "active_members": active_members
            }
        }
    except Exception as e:
        log_error(f"Error getting loyalty stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# NOTIFICATIONS DASHBOARD API
# ============================================================================

@router.get("/admin/notifications")
async def get_admin_notifications(session_token: Optional[str] = Cookie(None)):
    """Получить историю уведомлений"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        notification_columns = _get_table_columns(c, "notification_history")
        failed_count_expr = "COALESCE(failed_count, 0) AS failed_count" if "failed_count" in notification_columns else "0 AS failed_count"

        c.execute("""
            SELECT
                id, title, message, notification_type,
                recipients_count, sent_count, {failed_count_expr},
                status, created_at, sent_at
            FROM notification_history
            ORDER BY created_at DESC
            LIMIT 100
        """.format(failed_count_expr=failed_count_expr))

        notifications = []
        for row in c.fetchall():
            notifications.append({
                "id": str(row[0]),
                "title": row[1] or "",
                "message": row[2] or "",
                "type": row[3] or "push",
                "recipients": row[4] or 0,
                "sent_count": row[5] or 0,
                "failed_count": row[6] or 0,
                "status": row[7] or "pending",
                "created_at": row[8].isoformat() if row[8] else datetime.now().isoformat(),
                "sent_at": row[9].isoformat() if row[9] else None
            })

        conn.close()
        return {"success": True, "notifications": notifications}
    except Exception as e:
        log_error(f"Error getting notifications: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/notifications/templates")
async def get_notification_templates(session_token: Optional[str] = Cookie(None)):
    """Получить шаблоны уведомлений"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        template_columns = _get_table_columns(c, "notification_templates")
        has_legacy_columns = {"title", "message", "notification_type"}.issubset(template_columns)

        if has_legacy_columns:
            c.execute("""
                SELECT id, name, title, message, notification_type
                FROM notification_templates
                ORDER BY created_at DESC
            """)
        else:
            order_by_expression = "COALESCE(updated_at, created_at)" if "updated_at" in template_columns else "created_at"
            subject_sql = _get_template_value_expression(template_columns, "subject")
            body_sql = _get_template_value_expression(template_columns, "body")
            c.execute("""
                SELECT id, name, category, {subject_sql} AS subject, {body_sql} AS body
                FROM notification_templates
                ORDER BY {order_by_expression} DESC
            """.format(
                subject_sql=subject_sql,
                body_sql=body_sql,
                order_by_expression=order_by_expression
            ))

        templates = []
        rows = c.fetchall()

        if has_legacy_columns:
            for row in rows:
                templates.append({
                    "id": str(row[0]),
                    "name": row[1] or "",
                    "title": row[2] or "",
                    "message": row[3] or "",
                    "type": row[4] or "push"
                })
        else:
            for row in rows:
                category_value = row[2] or ""
                mapped_type = category_value if category_value in ["email", "push", "sms"] else "push"
                templates.append({
                    "id": str(row[0]),
                    "name": row[1] or "",
                    "title": row[3] or row[1] or "",
                    "message": row[4] or "",
                    "type": mapped_type
                })

        conn.close()
        return {"success": True, "templates": templates}
    except Exception as e:
        log_error(f"Error getting notification templates: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/notifications/templates")
async def create_notification_template(request: Request, session_token: Optional[str] = Cookie(None)):
    """Создать шаблон уведомления"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        template_columns = _get_table_columns(c, "notification_templates")
        has_legacy_columns = {"title", "message", "notification_type"}.issubset(template_columns)
        template_name = data.get("name") or f"template_{uuid.uuid4().hex[:8]}"
        template_type = data.get("type", "push")

        if has_legacy_columns:
            c.execute("""
                INSERT INTO notification_templates (name, title, message, notification_type)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (
                template_name,
                data.get("title"),
                data.get("message"),
                template_type
            ))
        else:
            template_category = template_type if template_type in ["email", "push", "sms"] else "transactional"
            template_title = data.get("title")
            template_message = data.get("message")
            insert_columns = ["name", "category"]
            insert_values = [
                template_name,
                template_category
            ]
            update_clauses = [
                "category = EXCLUDED.category"
            ]

            for subject_column in _get_template_storage_columns(template_columns, "subject"):
                insert_columns.append(subject_column)
                insert_values.append(template_title)
                quoted_column = _quote_sql_identifier(subject_column)
                update_clauses.append(f"{quoted_column} = EXCLUDED.{quoted_column}")

            for body_column in _get_template_storage_columns(template_columns, "body"):
                insert_columns.append(body_column)
                insert_values.append(template_message)
                quoted_column = _quote_sql_identifier(body_column)
                update_clauses.append(f"{quoted_column} = EXCLUDED.{quoted_column}")

            if "variables" in template_columns:
                insert_columns.append("variables")
                insert_values.append(json.dumps([]))
                update_clauses.append("variables = EXCLUDED.variables")

            if "updated_at" in template_columns:
                update_clauses.append("updated_at = CURRENT_TIMESTAMP")

            c.execute("""
                INSERT INTO notification_templates ({columns})
                VALUES ({placeholders})
                ON CONFLICT (name) DO UPDATE SET
                    {updates}
                RETURNING id
            """.format(
                columns=", ".join(insert_columns),
                placeholders=", ".join(["%s"] * len(insert_values)),
                updates=", ".join(update_clauses)
            ), tuple(insert_values))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {"success": True, "id": new_id}
    except Exception as e:
        log_error(f"Error creating notification template: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/notifications/send")
async def send_notification(request: Request, session_token: Optional[str] = Cookie(None)):
    """Отправить уведомление"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        notification_columns = _get_table_columns(c, "notification_history")

        # Определить целевую аудиторию
        target_segment = data.get("target_segment", "all")
        tier_filter = data.get("tier_filter", "")

        # Новые параметры фильтрации
        appointment_filter = data.get("appointment_filter", "")
        appointment_date = data.get("appointment_date", "")
        appointment_start_date = data.get("appointment_start_date", "")
        appointment_end_date = data.get("appointment_end_date", "")
        service_filter = data.get("service_filter", "")

        # Получить список клиентов
        # Выбираем instagram_id, telegram_id и preferences для фильтрации по настройкам клиента
        columns = "instagram_id, telegram_id, preferences"
        
        if target_segment == "all":
            c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")
        elif target_segment == "active":
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND is_active = TRUE")
        elif target_segment == "inactive":
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND is_active = FALSE")
        elif target_segment == "tier" and tier_filter:
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND loyalty_tier = %s", (tier_filter,))
        elif target_segment == "appointment-based":
            # Фильтр по датам записей
            if appointment_filter == "tomorrow":
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime::date = CURRENT_DATE + INTERVAL '1 day'
                      AND b.status IN ('pending', 'confirmed')
                """)
            elif appointment_filter == "specific_date" and appointment_date:
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime::date = %s::date
                      AND b.status IN ('pending', 'confirmed')
                """, (appointment_date,))
            elif appointment_filter == "date_range" and appointment_start_date and appointment_end_date:
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime BETWEEN %s AND %s
                      AND b.status IN ('pending', 'confirmed')
                """, (appointment_start_date, appointment_end_date))
            else:
                c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")
        elif target_segment == "service-based" and service_filter:
            # Фильтр по процедурам/услугам
            c.execute(f"""
                SELECT DISTINCT c.instagram_id, c.telegram_id
                FROM clients c
                JOIN bookings b ON c.instagram_id = b.client_id
                WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                  AND (b.service_name ILIKE %s OR b.service_id::text = %s)
            """, (f"%{service_filter}%", service_filter))
        else:
            c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")

        recipients = c.fetchall()
        notification_type = data.get("type", "push")

        filtered_recipients = []
        for recipient in recipients:
            recipient_preferences = {}
            preferences_raw = recipient[2] if len(recipient) > 2 else None
            if preferences_raw:
                try:
                    recipient_preferences = json.loads(preferences_raw)
                except Exception:
                    recipient_preferences = {}

            notification_preferences = recipient_preferences.get("notification_prefs", {})
            if notification_type == "email" and notification_preferences.get("email") is False:
                continue
            if notification_type == "sms" and notification_preferences.get("sms") is False:
                continue
            if notification_type == "push" and notification_preferences.get("push") is False:
                continue

            filtered_recipients.append(recipient)

        recipients_count = len(filtered_recipients)

        # Параметры планирования
        scheduled = data.get("scheduled", False)
        schedule_date = data.get("schedule_date", "")
        schedule_time = data.get("schedule_time", "")
        repeat_enabled = data.get("repeat_enabled", False)
        repeat_interval = data.get("repeat_interval", "")
        repeat_end_date = data.get("repeat_end_date", "")

        # Подготовить datetime для планирования
        schedule_datetime = None
        if scheduled and schedule_date and schedule_time:
            schedule_datetime = f"{schedule_date} {schedule_time}:00"

        # Сохранить параметры фильтрации в JSON
        filter_params = {
            "target_segment": target_segment,
            "tier_filter": tier_filter,
            "appointment_filter": appointment_filter,
            "appointment_date": appointment_date,
            "appointment_start_date": appointment_start_date,
            "appointment_end_date": appointment_end_date,
            "service_filter": service_filter
        }

        # Определить статус уведомления
        status = "pending" if scheduled else "sent"

        # Создать запись в истории
        c.execute("""
            INSERT INTO notification_history (
                title, message, notification_type,
                recipients_count, status,
                scheduled, schedule_datetime,
                repeat_enabled, repeat_interval, repeat_end_date,
                target_segment, filter_params
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data.get("title"),
            data.get("message"),
            notification_type,
            recipients_count,
            status,
            scheduled,
            schedule_datetime,
            repeat_enabled,
            repeat_interval if repeat_enabled else None,
            repeat_end_date if repeat_enabled else None,
            target_segment,
            json.dumps(filter_params)
        ))

        notification_id = c.fetchone()[0]

        # Если не запланировано - отправить сразу
        # Если не запланировано - отправить сразу
        if not scheduled:
            # Реальная отправка уведомлений через Telegram
            from integrations.telegram_bot import telegram_bot
            
            sent_count = 0
            failed_count = 0
            message_text = data.get("message")
            
            for row in filtered_recipients:
                inst_id = row[0]
                tg_id = row[1] if len(row) > 1 else None
                
                chat_id = None
                if tg_id:
                     chat_id = tg_id
                elif inst_id and str(inst_id).startswith('telegram_'):
                     chat_id = str(inst_id).replace('telegram_', '')
                
                if chat_id:
                    try:
                        res = telegram_bot.send_message(int(chat_id), message_text)
                        if res.get("ok"):
                            sent_count += 1
                        else:
                            failed_count += 1
                            log_error(f"Telegram send failed for {chat_id}: {res}", "api")
                    except Exception as e:
                        failed_count += 1
                        log_error(f"Error sending to {chat_id}: {e}", "api")

            if "failed_count" in notification_columns:
                c.execute("""
                    UPDATE notification_history
                    SET sent_count = %s, failed_count = %s, sent_at = CURRENT_TIMESTAMP, status = 'sent'
                    WHERE id = %s
                """, (sent_count, failed_count, notification_id))
            else:
                c.execute("""
                    UPDATE notification_history
                    SET sent_count = %s, sent_at = CURRENT_TIMESTAMP, status = 'sent'
                    WHERE id = %s
                """, (sent_count, notification_id))

        conn.commit()
        conn.close()

        return {"success": True, "id": notification_id, "scheduled": scheduled}
    except Exception as e:
        log_error(f"Error sending notification: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# PHOTO GALLERY API
# ============================================================================

def _remove_gallery_file(path_url: Optional[str]):
    if not path_url:
        return

    file_path = ""
    if path_url.startswith("/uploads/"):
        file_path = path_url[1:]
    elif path_url.startswith("/static/uploads/"):
        file_path = path_url.replace("/static/", "/", 1)[1:]

    if len(file_path) == 0:
        return

    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            log_info(f"Deleted file: {file_path}", "api")
        except Exception as e:
            log_error(f"Error deleting file {file_path}: {e}", "api")

@router.get("/admin/gallery/photos")
async def get_gallery_photos(session_token: Optional[str] = Cookie(None)):
    """Получить все фото галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")

        if len(gallery_columns) > 0:
            url_sql = "url" if "url" in gallery_columns else "''::text AS url"
            title_sql = "title" if "title" in gallery_columns else "''::text AS title"
            description_sql = "description" if "description" in gallery_columns else "''::text AS description"
            category_sql = "category" if "category" in gallery_columns else "'other'::text AS category"
            uploaded_by_sql = "uploaded_by" if "uploaded_by" in gallery_columns else "'Admin'::text AS uploaded_by"
            created_at_sql = "created_at" if "created_at" in gallery_columns else "CURRENT_TIMESTAMP AS created_at"
            is_featured_sql = "is_featured" if "is_featured" in gallery_columns else "FALSE AS is_featured"
            views_sql = "views" if "views" in gallery_columns else "0 AS views"
            before_sql = "before_photo_url" if "before_photo_url" in gallery_columns else "''::text AS before_photo_url"
            after_sql = "after_photo_url" if "after_photo_url" in gallery_columns else "''::text AS after_photo_url"
            client_sql = "client_id" if "client_id" in gallery_columns else "''::text AS client_id"
            visible_sql = "is_visible" if "is_visible" in gallery_columns else "TRUE AS is_visible"
            order_by_expression = "created_at DESC" if "created_at" in gallery_columns else "id DESC"

            c.execute("""
                SELECT
                    id,
                    {url_sql},
                    {title_sql},
                    {description_sql},
                    {category_sql},
                    {uploaded_by_sql},
                    {created_at_sql},
                    {is_featured_sql},
                    {views_sql},
                    {before_sql},
                    {after_sql},
                    {client_sql},
                    {visible_sql}
                FROM gallery_photos
                ORDER BY {order_by_expression}
            """.format(
                url_sql=url_sql,
                title_sql=title_sql,
                description_sql=description_sql,
                category_sql=category_sql,
                uploaded_by_sql=uploaded_by_sql,
                created_at_sql=created_at_sql,
                is_featured_sql=is_featured_sql,
                views_sql=views_sql,
                before_sql=before_sql,
                after_sql=after_sql,
                client_sql=client_sql,
                visible_sql=visible_sql,
                order_by_expression=order_by_expression
            ))
        else:
            c.execute("""
                SELECT
                    cg.id,
                    COALESCE(NULLIF(cg.after_photo, ''), NULLIF(cg.before_photo, ''), '') AS url,
                    COALESCE(cl.name, cg.client_id, '') AS title,
                    COALESCE(cg.notes, '') AS description,
                    COALESCE(cg.category, 'other') AS category,
                    COALESCE(u.full_name, 'Admin') AS uploaded_by,
                    cg.created_at,
                    FALSE AS is_featured,
                    0 AS views,
                    COALESCE(cg.before_photo, '') AS before_photo_url,
                    COALESCE(cg.after_photo, '') AS after_photo_url,
                    COALESCE(cg.client_id, '') AS client_id,
                    TRUE AS is_visible
                FROM client_gallery cg
                LEFT JOIN clients cl ON cl.instagram_id = cg.client_id
                LEFT JOIN users u ON u.id = cg.master_id
                ORDER BY cg.created_at DESC
            """)

        photos = []
        for row in c.fetchall():
            photos.append({
                "id": str(row[0]),
                "url": row[1] or "",
                "title": row[2] or "",
                "description": row[3] or "",
                "category": row[4] or "other",
                "uploaded_by": row[5] or "Admin",
                "created_at": row[6].isoformat() if row[6] else datetime.now().isoformat(),
                "is_featured": row[7] or False,
                "views": row[8] or 0,
                "before_photo_url": row[9] or "",
                "after_photo_url": row[10] or "",
                "client_id": row[11] or "",
                "is_visible": True if row[12] is None else row[12]
            })

        conn.close()
        return {"success": True, "photos": photos}
    except Exception as e:
        log_error(f"Error getting gallery photos: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/gallery/categories")
async def get_gallery_categories(session_token: Optional[str] = Cookie(None)):
    """Получить список категорий из базы данных"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Получаем уникальные категории из services таблицы
        c.execute("""
            SELECT DISTINCT category
            FROM services
            WHERE category IS NOT NULL AND category != ''
            ORDER BY category
        """)

        categories = []
        for row in c.fetchall():
            if row[0]:
                categories.append({
                    "value": row[0].lower().replace(' ', '_'),
                    "label": row[0]
                })

        # Добавляем дополнительные категории если нужно
        default_categories = [
            {"value": "portfolio", "label": "Портфолио"},
            {"value": "interior", "label": "Интерьер"},
            {"value": "other", "label": "Другое"}
        ]

        # Объединяем, избегая дубликатов
        existing_values = {cat["value"] for cat in categories}
        for cat in default_categories:
            if cat["value"] not in existing_values:
                categories.append(cat)

        conn.close()
        return {"success": True, "categories": categories}
    except Exception as e:
        log_error(f"Error getting gallery categories: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/gallery/stats")
async def get_gallery_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")

        if len(gallery_columns) > 0:
            c.execute("SELECT COUNT(*) FROM gallery_photos")
            total_photos = c.fetchone()[0] or 0

            if "views" in gallery_columns:
                c.execute("SELECT COALESCE(SUM(views), 0) FROM gallery_photos")
                total_views = c.fetchone()[0] or 0
            else:
                total_views = 0

            if "is_featured" in gallery_columns:
                c.execute("SELECT COUNT(*) FROM gallery_photos WHERE is_featured = TRUE")
                featured_count = c.fetchone()[0] or 0
            else:
                featured_count = 0

            if "created_at" in gallery_columns:
                c.execute("SELECT COUNT(*) FROM gallery_photos WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'")
                recent_uploads = c.fetchone()[0] or 0
            else:
                recent_uploads = 0
        else:
            c.execute("SELECT COUNT(*) FROM client_gallery")
            total_photos = c.fetchone()[0] or 0
            total_views = 0
            featured_count = 0
            c.execute("SELECT COUNT(*) FROM client_gallery WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'")
            recent_uploads = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_photos": total_photos,
                "total_views": total_views,
                "featured_count": featured_count,
                "recent_uploads": recent_uploads
            }
        }
    except Exception as e:
        log_error(f"Error getting gallery stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/gallery/photos")
async def upload_gallery_photo(
    file: UploadFile = File(None),
    before_photo: UploadFile = File(None),
    after_photo: UploadFile = File(None),
    title: str = Form(...),
    description: str = Form(""),
    category: str = Form("other"),
    is_featured: str = Form("false"),
    client_id: Optional[str] = Form(None),
    session_token: Optional[str] = Cookie(None)
):
    """Загрузить фото в галерею (обычное или до/после)"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        upload_dir = "uploads/gallery"
        os.makedirs(upload_dir, exist_ok=True)

        photo_url = ""
        before_photo_url = ""
        after_photo_url = ""

        # Обработка обычного фото
        if file:
            file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
            unique_filename = f"{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)

            photo_url = f"/uploads/gallery/{unique_filename}"

        # Обработка фото "до"
        if before_photo:
            file_extension = before_photo.filename.split('.')[-1] if '.' in before_photo.filename else 'jpg'
            unique_filename = f"before_{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await before_photo.read()
                f.write(content)

            before_photo_url = f"/uploads/gallery/{unique_filename}"

        # Обработка фото "после"
        if after_photo:
            file_extension = after_photo.filename.split('.')[-1] if '.' in after_photo.filename else 'jpg'
            unique_filename = f"after_{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await after_photo.read()
                f.write(content)

            after_photo_url = f"/uploads/gallery/{unique_filename}"

        # Если нет обычного фото, но есть фото "после", используем его как основное
        if not photo_url and after_photo_url:
            photo_url = after_photo_url

        # Сохранить в БД
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")

        if len(gallery_columns) > 0:
            insert_columns = ["url", "title", "description", "category"]
            insert_values = [photo_url, title, description, category]

            if "uploaded_by" in gallery_columns:
                insert_columns.append("uploaded_by")
                insert_values.append(user.get("email", "admin"))
            if "is_featured" in gallery_columns:
                insert_columns.append("is_featured")
                insert_values.append(is_featured.lower() == "true")
            if "client_id" in gallery_columns:
                insert_columns.append("client_id")
                insert_values.append(client_id if client_id else None)
            if "before_photo_url" in gallery_columns:
                insert_columns.append("before_photo_url")
                insert_values.append(before_photo_url if before_photo_url else None)
            if "after_photo_url" in gallery_columns:
                insert_columns.append("after_photo_url")
                insert_values.append(after_photo_url if after_photo_url else None)

            c.execute("""
                INSERT INTO gallery_photos ({columns})
                VALUES ({placeholders})
                RETURNING id
            """.format(
                columns=", ".join(insert_columns),
                placeholders=", ".join(["%s"] * len(insert_values))
            ), tuple(insert_values))
        else:
            normalized_client_id = (client_id or "").strip()
            if len(normalized_client_id) == 0:
                conn.close()
                raise HTTPException(status_code=400, detail="client_id is required")

            notes_value = (description or "").strip()
            if len(notes_value) == 0:
                notes_value = (title or "").strip()

            c.execute("""
                INSERT INTO client_gallery (
                    client_id, before_photo, after_photo, category, notes, master_id
                ) VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                normalized_client_id,
                before_photo_url if len(before_photo_url) > 0 else None,
                after_photo_url if len(after_photo_url) > 0 else (photo_url if len(photo_url) > 0 else None),
                category,
                notes_value,
                user.get("id")
            ))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {
            "success": True,
            "id": new_id,
            "url": photo_url,
            "before_photo_url": before_photo_url,
            "after_photo_url": after_photo_url
        }
    except Exception as e:
        log_error(f"Error uploading gallery photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/gallery/photos/{photo_id}")
async def delete_gallery_photo(photo_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить фото из галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")

        if len(gallery_columns) > 0:
            url_sql = "url" if "url" in gallery_columns else "NULL::text AS url"
            before_sql = "before_photo_url" if "before_photo_url" in gallery_columns else "NULL::text AS before_photo_url"
            after_sql = "after_photo_url" if "after_photo_url" in gallery_columns else "NULL::text AS after_photo_url"
            c.execute("""
                SELECT {url_sql}, {before_sql}, {after_sql}
                FROM gallery_photos
                WHERE id = %s
            """.format(
                url_sql=url_sql,
                before_sql=before_sql,
                after_sql=after_sql
            ), (photo_id,))
            result = c.fetchone()

            if result:
                photo_url = result[0]
                before_url = result[1]
                after_url = result[2]
                _remove_gallery_file(photo_url)
                if before_url and before_url != photo_url:
                    _remove_gallery_file(before_url)
                if after_url and after_url != photo_url:
                    _remove_gallery_file(after_url)

            c.execute("DELETE FROM gallery_photos WHERE id = %s", (photo_id,))
        else:
            c.execute("""
                SELECT before_photo, after_photo
                FROM client_gallery
                WHERE id = %s
            """, (photo_id,))
            result = c.fetchone()
            if result:
                _remove_gallery_file(result[0])
                if result[1] != result[0]:
                    _remove_gallery_file(result[1])

            c.execute("DELETE FROM client_gallery WHERE id = %s", (photo_id,))

        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting gallery photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/gallery/photos/{photo_id}/featured")
async def toggle_featured_photo(photo_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Переключить статус избранного для фото"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")
        if len(gallery_columns) > 0 and "is_featured" in gallery_columns:
            c.execute("""
                UPDATE gallery_photos
                SET is_featured = %s
                WHERE id = %s
            """, (data.get("is_featured", False), photo_id))

        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error toggling featured photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/gallery/photos/{photo_id}/visibility")
async def toggle_gallery_photo_visibility(
    photo_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Переключить видимость фото"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()
        gallery_columns = _get_table_columns(c, "gallery_photos")
        if len(gallery_columns) > 0 and "is_visible" in gallery_columns:
            c.execute("""
                UPDATE gallery_photos
                SET is_visible = %s
                WHERE id = %s
            """, (data.get("is_visible", True), photo_id))

        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error toggling photo visibility: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/admin/export-report")
async def export_report(request: Request, session_token: Optional[str] = Cookie(None)):
    """Экспорт отчета в CSV/PDF/Excel"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    try:
        data = await request.json()
        export_format = data.get('format', 'csv').lower()
        
        # Получаем данные
        from db import get_stats
        stats = get_stats()
        
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        period_str = f"{data.get('start_date', 'All')} - {data.get('end_date', 'Now')}"
        
        if export_format == 'csv':
            import io
            import csv
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            writer.writerow(["Report Generated", report_date])
            writer.writerow(["Period", period_str])
            writer.writerow([])
            writer.writerow(["Metric", "Value"])
            
            if stats:
                for k, v in stats.items():
                    if isinstance(v, (int, float, str)):
                        writer.writerow([k.replace('_', ' ').title(), v])
                    elif isinstance(v, dict) and 'value' in v:
                        writer.writerow([k.replace('_', ' ').title(), v['value']])

            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.csv"
                }
            )
            
        elif export_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                import io
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"
                
                # Header
                ws['A1'] = "Report Generated"
                ws['B1'] = report_date
                ws['A1'].font = Font(bold=True)
                
                ws['A2'] = "Period"
                ws['B2'] = period_str
                ws['A2'].font = Font(bold=True)
                
                # Data headers
                ws['A4'] = "Metric"
                ws['B4'] = "Value"
                ws['A4'].font = Font(bold=True)
                ws['B4'].font = Font(bold=True)
                
                row = 5
                if stats:
                    for k, v in stats.items():
                        if isinstance(v, (int, float, str)):
                            ws[f'A{row}'] = k.replace('_', ' ').title()
                            ws[f'B{row}'] = v
                            row += 1
                        elif isinstance(v, dict) and 'value' in v:
                            ws[f'A{row}'] = k.replace('_', ' ').title()
                            ws[f'B{row}'] = v['value']
                            row += 1
                
                # Auto-size columns (openpyxl: type stubs may not expose column_dimensions)
                ws.column_dimensions['A'].width = 30  # type: ignore[union-attr]
                ws.column_dimensions['B'].width = 20  # type: ignore[union-attr]
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    }
                )
            except ImportError:
                log_error("openpyxl not installed", "api")
                raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl.")
                
        elif export_format == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                import io
                
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []
                styles = getSampleStyleSheet()
                
                # Title
                title = Paragraph(f"<b>Report Generated: {report_date}</b>", styles['Heading1'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Period
                period = Paragraph(f"<b>Period:</b> {period_str}", styles['Normal'])
                elements.append(period)
                elements.append(Spacer(1, 20))
                
                # Data table
                data = [['Metric', 'Value']]
                if stats:
                    for k, v in stats.items():
                        if isinstance(v, (int, float, str)):
                            data.append([k.replace('_', ' ').title(), str(v)])
                        elif isinstance(v, dict) and 'value' in v:
                            data.append([k.replace('_', ' ').title(), str(v['value'])])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                doc.build(elements)
                
                buffer.seek(0)
                return StreamingResponse(
                    iter([buffer.getvalue()]),
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.pdf"
                    }
                )
            except ImportError:
                log_error("reportlab not installed", "api")
                raise HTTPException(status_code=501, detail="PDF export not available. Install reportlab.")
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {export_format}")
        
    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Export error: {e}", "api")
        raise HTTPException(status_code=500, detail=str(e))
