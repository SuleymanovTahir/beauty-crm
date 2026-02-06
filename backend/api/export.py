"""
API Endpoints для экспорта данных (CSV, PDF, Excel)
"""
from fastapi import APIRouter, Query, Cookie
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional
from datetime import datetime
import csv
import io

from db import get_all_bookings, get_analytics_data
from db.settings import get_salon_settings
from core.config import DATABASE_NAME, SALON_PHONE_DEFAULT
from db.connection import get_db_connection
from utils.logger import log_error, log_warning
from utils.translation import t, register_fonts

# Попытка импортировать библиотеки для PDF и Excel
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    log_warning("ReportLab не установлен. Экспорт в PDF недоступен.", "export")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log_warning("openpyxl не установлен. Экспорт в Excel недоступен.", "export")

router = APIRouter(tags=["Export"])

# ===== ФУНКЦИИ ЭКСПОРТА КЛИЕНТОВ =====

def export_clients_csv(clients, lang='en'):
    """Экспорт клиентов в CSV"""
    import json
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Headers translated if possible, or fallback to English as standard for CSV
    headers = [
        t(lang, 'booking.formName', 'Name'),
        t(lang, 'common.phone', 'Phone'),
        'Username', 'Email', 
        t(lang, 'booking.auth.gender', 'Gender'),
        'Category', 
        t(lang, 'booking.auth.birthday', 'Date of Birth'),
        'Card Number', 'Discount', 
        t(lang, 'common.status', 'Status'),
        'Total Visits', 'Total Spend, AED', 'Paid, AED',
        'First Visit', 'Last Visit', 'Notes',
        'Additional Phone Number', 'Newsletter', 'Personal data'
    ]
    writer.writerow(headers)
    
    for c in clients:
        # Extract fields from tuple based on query order
        # Query: instagram_id, username, phone, name, first_contact, last_contact, 
        # total_messages, labels, status, lifetime_value, profile_pic, notes, is_pinned,
        # gender, card_number, discount, total_visits, additional_phone, newsletter_agreed, 
        # personal_data_agreed, total_spend, paid_amount, birthday, email
        
        # Parse phone from JSON array if needed
        phone_value = c[2] or ''
        try:
            if phone_value and phone_value.startswith('['):
                phones = json.loads(phone_value)
                phone_value = phones[0] if phones else ''
        except (json.JSONDecodeError, TypeError, IndexError):
            pass
        
        writer.writerow([
            c[3] or '',             # Name
            phone_value,            # Phone (first from array)
            c[1] or '',             # Username
            c[23] if len(c) > 23 else '', # Email
            c[13] if len(c) > 13 else '', # Gender
            c[7] or '',             # Category/Labels
            c[22] if len(c) > 22 else '', # Birthday
            c[14] if len(c) > 14 else '', # Card Number
            c[15] if len(c) > 15 else 0,  # Discount
            c[8] if len(c) > 8 else 'new', # Status
            c[16] if len(c) > 16 else 0,  # Total visits
            c[20] if len(c) > 20 else 0,  # Total spend
            c[21] if len(c) > 21 else 0,  # Paid amount
            c[4],                   # First contact
            c[5],                   # Last contact
            c[11] or '',            # Notes
            c[17] if len(c) > 17 else '', # Additional Phone
            'Yes' if len(c) > 18 and c[18] else 'No', # Newsletter
            'Yes' if len(c) > 19 and c[19] else 'No'  # Personal data
        ])
    
    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_clients_pdf(clients, lang='en'):
    """Экспорт клиентов в PDF"""
    if not PDF_AVAILABLE:
        raise Exception("PDF экспорт недоступен")
    
    fontName = register_fonts()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#ec4899'),
        spaceAfter=30,
        alignment=TA_CENTER,
        fontName=fontName
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=fontName
    )
    
    salon = get_salon_settings()
    clients_base_label = t(lang, 'admin/clients:title', 'База клиентов')
    title = Paragraph(f"{clients_base_label} - {salon['name']}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    date_label = t(lang, 'booking.formDate', 'Дата')
    date_text = Paragraph(f"{date_label}: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 
                         normal_style)
    elements.append(date_text)
    elements.append(Spacer(1, 20))
    
    data = [[
        t(lang, 'booking.formName', 'Имя'),
        t(lang, 'common.phone', 'Телефон'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.visits', 'Визитов'),
        'LTV'
    ]]
    
    for c in clients:
        data.append([
            (c[3] or c[1] or 'N/A')[:25],
            (c[2] or '-')[:15],
            (c[8] if len(c) > 8 else 'new')[:10],
            str(c[16]) if len(c) > 16 else '0',
            f"{c[9] if len(c) > 9 else 0} AED"
        ])
    
    table = Table(data, colWidths=[140, 100, 80, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), fontName),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def export_clients_excel(clients, lang='en'):
    """Экспорт клиентов в Excel"""
    import json
    if not EXCEL_AVAILABLE:
        raise Exception("Excel экспорт недоступен")
    
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, 'admin/clients:title', 'Клиенты')
    
    # Headers translated if possible, or fallback to English as standard for Excel
    headers = [
        t(lang, 'booking.formName', 'Name'),
        t(lang, 'common.phone', 'Phone'),
        'Username', 'Email', 
        t(lang, 'booking.auth.gender', 'Gender'),
        'Category', 
        t(lang, 'booking.auth.birthday', 'Date of Birth'),
        'Card Number', 'Discount', 
        t(lang, 'common.status', 'Status'),
        'Total Visits', 'Total Spend, AED', 'Paid, AED',
        'First Visit', 'Last Visit', 'Notes',
        'Additional Phone Number', 'Newsletter', 'Personal data'
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for c in clients:
        # Parse phone from JSON array if needed
        phone_value = c[2] or ''
        try:
            if phone_value and phone_value.startswith('['):
                phones = json.loads(phone_value)
                phone_value = phones[0] if phones else ''
        except (json.JSONDecodeError, TypeError, IndexError):
            pass
        
        ws.append([
            c[3] or '',             # Name
            phone_value,            # Phone (first from array)
            c[1] or '',             # Username
            c[23] if len(c) > 23 else '', # Email
            c[13] if len(c) > 13 else '', # Gender
            c[7] or '',             # Category/Labels
            c[22] if len(c) > 22 else '', # Birthday
            c[14] if len(c) > 14 else '', # Card Number
            c[15] if len(c) > 15 else 0,  # Discount
            c[8] if len(c) > 8 else 'new', # Status
            c[16] if len(c) > 16 else 0,  # Total visits
            c[20] if len(c) > 20 else 0,  # Total spend
            c[21] if len(c) > 21 else 0,  # Paid amount
            c[4],                   # First contact
            c[5],                   # Last contact
            c[11] or '',            # Notes
            c[17] if len(c) > 17 else '', # Additional Phone
            'Yes' if len(c) > 18 and c[18] else 'No', # Newsletter
            'Yes' if len(c) > 19 and c[19] else 'No'  # Personal data
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ===== ФУНКЦИИ ЭКСПОРТА ЗАПИСЕЙ =====

def export_bookings_csv(bookings, lang='en'):
    """Экспорт записей в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = [
        t(lang, 'adminPanel.dashboard.id', 'ID'),
        t(lang, 'adminPanel.dashboard.client', 'Клиент'),
        t(lang, 'common.service', 'Услуга'),
        t(lang, 'adminPanel.dashboard.datetime', 'Дата/Время'),
        t(lang, 'common.phone', 'Телефон'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.revenue', 'Доход (AED)'),
        t(lang, 'common.created_at', 'Создана')
    ]
    writer.writerow(headers)
    
    for b in bookings:
        writer.writerow([
            b[0], b[5] or '', b[2] or '', b[3], b[4] or '',
            b[6], b[8] if len(b) > 8 else 0, b[7]
        ])
    
    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_bookings_pdf(bookings, lang='en'):
    """Экспорт записей в PDF"""
    if not PDF_AVAILABLE:
        raise Exception("PDF экспорт недоступен")
    
    fontName = register_fonts()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#ec4899'),
        spaceAfter=30,
        alignment=1,
        fontName=fontName
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=fontName
    )
    
    salon = get_salon_settings()
    report_title = t(lang, 'adminPanel.dashboard.report', 'Отчет')
    title = Paragraph(f"{report_title} - {salon['name']}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    date_label = t(lang, 'booking.formDate', 'Дата')
    date_text = Paragraph(f"{date_label}: {datetime.now().strftime('%d.%m.%Y %H:%M')}", 
                         normal_style)
    elements.append(date_text)
    elements.append(Spacer(1, 20))
    
    data = [[
        t(lang, 'adminPanel.dashboard.id', 'ID'),
        t(lang, 'adminPanel.dashboard.client', 'Клиент'),
        t(lang, 'common.service', 'Услуга'),
        t(lang, 'booking.formDate', 'Дата'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.revenue', 'Доход')
    ]]
    
    for b in bookings:
        try:
            date_obj = datetime.fromisoformat(b[3])
            date_str = date_obj.strftime('%d.%m %H:%M')
        except:
            date_str = str(b[3])[:16]
        
        data.append([
            str(b[0]), (b[5] or 'N/A')[:20], (b[2] or 'N/A')[:25],
            date_str, (b[6] or '')[:10],
            f"{b[8] if len(b) > 8 else 0} AED"
        ])
    
    table = Table(data, colWidths=[30, 80, 100, 70, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), fontName),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def export_bookings_excel(bookings, lang='en'):
    """Экспорт записей в Excel"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel экспорт недоступен")
    
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, 'adminPanel.dashboard.bookings', 'Записи')
    
    headers = [
        t(lang, 'adminPanel.dashboard.id', 'ID'),
        t(lang, 'adminPanel.dashboard.client', 'Клиент'),
        t(lang, 'common.service', 'Услуга'),
        t(lang, 'adminPanel.dashboard.datetime', 'Дата/Время'),
        t(lang, 'common.phone', 'Телефон'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.revenue', 'Доход (AED)'),
        t(lang, 'common.created_at', 'Создана')
    ]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for b in bookings:
        ws.append([
            b[0], b[5] or '', b[2] or '', b[3], b[4] or '',
            b[6], b[8] if len(b) > 8 else 0, b[7]
        ])
    
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ===== ФУНКЦИИ ЭКСПОРТА ВСЕХ ДАННЫХ =====

def export_full_data_csv(lang='en'):
    """Экспорт всех данных (клиенты + сообщения + записи) в один CSV"""
    conn = get_db_connection()
    c = conn.cursor()
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Заголовок
    writer.writerow(['Type', 'Client ID', 'Client Name', 'Phone', 'Username', 
                    'Data Type', 'Content', 'Date/Time', 'Status', 'Revenue'])
    
    # Получаем клиентов
    c.execute("""SELECT instagram_id, username, phone, name, status 
                 FROM clients ORDER BY instagram_id""")
    clients_data = {row[0]: {'username': row[1], 'phone': row[2], 
                             'name': row[3], 'status': row[4]} 
                    for row in c.fetchall()}
    
    # Сообщения
    c.execute("""SELECT instagram_id, message_text, sender, created_at, message_type
                 FROM messages ORDER BY instagram_id, created_at""")
    for row in c.fetchall():
        client_id = row[0]
        client = clients_data.get(client_id, {})
        writer.writerow([
            'Сообщение',
            client_id,
            client.get('name', ''),
            client.get('phone', ''),
            client.get('username', ''),
            row[4] or 'text',
            (row[1] or '')[:100],
            row[3],
            f"От: {row[2]}",
            ''
        ])
    
    # Записи
    c.execute("""SELECT b.instagram_id, b.service_name, b.booking_datetime, 
                        b.phone, b.status, b.revenue, b.client_name
                 FROM bookings b ORDER BY b.instagram_id, b.booking_datetime""")
    for row in c.fetchall():
        client_id = row[0]
        client = clients_data.get(client_id, {})
        writer.writerow([
            'Запись',
            client_id,
            row[6] or client.get('name', ''),
            row[3] or client.get('phone', ''),
            client.get('username', ''),
            row[1] or '',
            '',
            row[2],
            row[4],
            row[5] or 0
        ])
    
    conn.close()
    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_full_data_excel(lang='en'):
    """Экспорт всех данных в Excel с отдельными листами"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel экспорт недоступен")
    
    conn = get_db_connection()
    c = conn.cursor()
    
    wb = Workbook()
    
    # ===== Лист 1: Клиенты =====
    ws_clients = wb.active
    ws_clients.title = t(lang, 'admin/clients:title', 'Клиенты')
    
    headers = ['ID', 
              t(lang, 'booking.formName', 'Name'),
              'Username', 
              t(lang, 'common.phone', 'Phone'),
              'Add. Phone', 'Email', 
              t(lang, 'booking.auth.gender', 'Gender'),
              'Category', 
              t(lang, 'booking.auth.birthday', 'Date of Birth'),
              'Card', 'Discount (%)', 
              t(lang, 'common.status', 'Status'),
              'Messages', 'Total Visits', 'Total Spend', 'Paid',
              'LTV', 'First Contact', 'Last Contact', 'Notes',
              'Newsletter', 'Personal Data']
    ws_clients.append(headers)
    
    c.execute("""SELECT instagram_id, name, username, phone, additional_phone, email, gender,
                        labels, birthday, card_number, discount, status, total_messages,
                        total_visits, total_spend, paid_amount, lifetime_value,
                        first_contact, last_contact, notes, newsletter_agreed, personal_data_agreed
                 FROM clients""")
    for row in c.fetchall():
        # Convert boolean integers to Yes/No
        row_list = list(row)
        row_list[20] = 'Да' if row_list[20] else 'Нет' # newsletter
        row_list[21] = 'Да' if row_list[21] else 'Нет' # personal data
        ws_clients.append(row_list)
    
    # Форматирование
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    for cell in ws_clients[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # ===== Лист 2: Сообщения =====
    ws_messages = wb.create_sheet(t(lang, 'manager/messages:title', 'Сообщения'))
    headers = ['Client ID', t(lang, 'booking.formName', 'Name'), t(lang, 'manager/messages:message', 'Message'), 'Sender', t(lang, 'booking.formDate', 'Date'), 'Type']
    ws_messages.append(headers)
    
    c.execute("""SELECT m.instagram_id, c.name, m.message_text, m.sender, 
                        m.created_at, m.message_type
                 FROM messages m
                 LEFT JOIN clients c ON m.instagram_id = c.instagram_id
                 ORDER BY m.created_at DESC LIMIT 1000""")
    for row in c.fetchall():
        ws_messages.append([
            row[0], row[1] or '', (row[2] or '')[:200], 
            row[3], row[4], row[5] or 'text'
        ])
    
    for cell in ws_messages[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # ===== Лист 3: Записи =====
    ws_bookings = wb.create_sheet(t(lang, 'adminPanel.dashboard.bookings', 'Записи'))
    headers = ['ID', 'Client ID', t(lang, 'adminPanel.dashboard.client', 'Client'), t(lang, 'common.service', 'Service'), t(lang, 'adminPanel.dashboard.datetime', 'Date/Time'), 
              t(lang, 'common.phone', 'Phone'), t(lang, 'common.status', 'Status'), t(lang, 'adminPanel.dashboard.revenue', 'Revenue')]
    ws_bookings.append(headers)
    
    c.execute("""SELECT b.id, b.instagram_id, b.client_name, b.service_name,
                        b.booking_datetime, b.phone, b.status, b.revenue
                 FROM bookings b ORDER BY b.booking_datetime DESC""")
    for row in c.fetchall():
        ws_bookings.append(row)
    
    for cell in ws_bookings[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Автоширина колонок для всех листов
    for ws in [ws_clients, ws_messages, ws_bookings]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    conn.close()
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ===== ENDPOINTS =====

@router.get("/export/clients")
async def export_clients(
    format: str = Query("csv"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать клиентов в CSV, PDF или Excel"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    conn = get_db_connection()
    c = conn.cursor()
    
    if date_from and date_to:
        c.execute("""SELECT instagram_id, username, phone, name, first_contact, 
                     last_contact, total_messages, labels, status, lifetime_value,
                     profile_pic, notes, is_pinned,
                     gender, card_number, discount, total_visits, additional_phone, 
                     newsletter_agreed, personal_data_agreed, total_spend, paid_amount, 
                     birthday, email
                     FROM clients 
                     WHERE first_contact >= %s AND first_contact <= %s
                     ORDER BY is_pinned DESC, last_contact DESC""",
                  (date_from, date_to))
    else:
        c.execute("""SELECT instagram_id, username, phone, name, first_contact, 
                     last_contact, total_messages, labels, status, lifetime_value,
                     profile_pic, notes, is_pinned,
                     gender, card_number, discount, total_visits, additional_phone, 
                     newsletter_agreed, personal_data_agreed, total_spend, paid_amount, 
                     birthday, email
                     FROM clients 
                     ORDER BY is_pinned DESC, last_contact DESC""")
    
    clients = c.fetchall()
    conn.close()
    
    try:
        if format == "csv":
            content = export_clients_csv(clients, lang=lang)
            media_type = "text/csv"
            filename = f"clients_{date_from or 'all'}_{date_to or datetime.now().strftime('%Y%m%d')}.csv"
        elif format == "pdf":
            content = export_clients_pdf(clients, lang=lang)
            media_type = "application/pdf"
            filename = f"clients_{date_from or 'all'}_{date_to or datetime.now().strftime('%Y%m%d')}.pdf"
        elif format == "excel":
            content = export_clients_excel(clients, lang=lang)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"clients_{date_from or 'all'}_{date_to or datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/bookings")
async def export_bookings(
    format: str = Query("csv"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать записи в CSV, PDF или Excel"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    all_bookings = get_all_bookings()
    
    if date_from and date_to:
        filtered_bookings = []
        for b in all_bookings:
            try:
                booking_date = datetime.fromisoformat(b[3])
                start_date = datetime.fromisoformat(date_from)
                end_date = datetime.fromisoformat(date_to)
                
                if start_date <= booking_date <= end_date:
                    filtered_bookings.append(b)
            except:
                continue
        bookings = filtered_bookings
    else:
        bookings = all_bookings
    
    try:
        if format == "csv":
            content = export_bookings_csv(bookings, lang=lang)
            media_type = "text/csv"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.csv"
        elif format == "pdf":
            content = export_bookings_pdf(bookings, lang=lang)
            media_type = "application/pdf"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.pdf"
        elif format == "excel":
            content = export_bookings_excel(bookings, lang=lang)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/analytics")
async def export_analytics(
    format: str = Query("csv"),
    period: int = Query(30),
    date_from: str = Query(None),
    date_to: str = Query(None),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать аналитику"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    if date_from and date_to:
        analytics = get_analytics_data(date_from=date_from, date_to=date_to)
    else:
        analytics = get_analytics_data(days=period)
    
    try:
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            
            writer.writerow([t(lang, 'booking.formDate', 'Дата'), t(lang, 'adminPanel.dashboard.bookings', 'Записи')])
            for date, count in analytics.get('bookings_by_day', []):
                writer.writerow([date, count])
            
            writer.writerow([])
            writer.writerow([t(lang, 'common.service', 'Услуга'), t(lang, 'admin/services:quantity', 'Количество'), t(lang, 'adminPanel.dashboard.revenue', 'Доход')])
            for name, count, revenue in analytics.get('services_stats', []):
                writer.writerow([name, count, revenue])
            
            output.seek(0)
            content = output.getvalue().encode('utf-8')
            media_type = "text/csv"
            filename = f"analytics_{datetime.now().strftime('%Y%m%d')}.csv"
        
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                return JSONResponse({"error": "Excel export not available"}, 
                                  status_code=500)
            
            wb = Workbook()
            ws = wb.active
            ws.title = t(lang, 'adminPanel.dashboard.analytics', 'Аналитика')
            
            ws.append([t(lang, 'booking.formDate', 'Дата'), t(lang, 'adminPanel.dashboard.bookings', 'Записи')])
            for date, count in analytics.get('bookings_by_day', []):
                ws.append([date, count])
            
            ws.append([])
            ws.append([t(lang, 'common.service', 'Услуга'), t(lang, 'admin/services:quantity', 'Количество'), t(lang, 'adminPanel.dashboard.revenue', 'Доход')])
            for name, count, revenue in analytics.get('services_stats', []):
                ws.append([name, count, revenue])
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            content = buffer.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export analytics error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/messages")
async def export_messages(
    client_id: str = Query(None),
    format: str = Query("csv"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать сообщения в CSV, PDF или Excel"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    conn = get_db_connection()
    c = conn.cursor()
    
    query = """SELECT m.id, m.instagram_id, c.username, c.name, m.message_text, 
               m.sender, m.message_type, m.created_at, m.file_url, m.instagram_file_url
               FROM messages m
               LEFT JOIN clients c ON m.instagram_id = c.instagram_id"""
    
    params = []
    conditions = []
    
    if client_id:
        conditions.append("m.instagram_id = %s")
        params.append(client_id)
    
    if date_from and date_to:
        conditions.append("m.created_at >= %s AND m.created_at <= %s")
        params.extend([date_from, date_to])
    
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    
    query += " ORDER BY m.created_at DESC"
    
    c.execute(query, params)
    messages = c.fetchall()
    conn.close()
    
    try:
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            
            headers = [
                t(lang, 'adminPanel.dashboard.id', 'ID'),
                'Instagram ID', 'Username',
                t(lang, 'booking.formName', 'Имя'),
                t(lang, 'manager/messages:message', 'Сообщение'),
                'Sender', 'Type',
                t(lang, 'booking.formDate', 'Дата'),
                t(lang, 'manager/messages:file', 'Файл')
            ]
            writer.writerow(headers)
            
            for m in messages:
                writer.writerow([
                    m[0], m[1], m[2] or '', m[3] or '', 
                    (m[4] or '')[:100], m[5], m[6], m[7], 
                    m[8] or m[9] or ''
                ])
            
            output.seek(0)
            content = output.getvalue().encode('utf-8')
            media_type = "text/csv"
            filename = f"messages_{client_id or 'all'}_{datetime.now().strftime('%Y%m%d')}.csv"
        
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                return JSONResponse({"error": "Excel export not available"}, 
                                  status_code=500)
            
            wb = Workbook()
            ws = wb.active
            ws.title = t(lang, 'manager/messages:title', 'Сообщения')
            
            headers = [
                t(lang, 'adminPanel.dashboard.id', 'ID'),
                'Instagram ID', 'Username',
                t(lang, 'booking.formName', 'Имя'),
                t(lang, 'manager/messages:message', 'Сообщение'),
                'Sender', 'Type',
                t(lang, 'booking.formDate', 'Дата'),
                t(lang, 'manager/messages:file', 'Файл')
            ]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for m in messages:
                ws.append([
                    m[0], m[1], m[2] or '', m[3] or '', 
                    (m[4] or '')[:100], m[5], m[6], m[7], 
                    m[8] or m[9] or ''
                ])
            
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            content = buffer.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"messages_{client_id or 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export messages error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/full-data")
async def export_full_data(
    format: str = Query("csv"),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать все данные (клиенты + сообщения + записи)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        if format == "csv":
            content = export_full_data_csv(lang=lang)
            media_type = "text/csv"
            filename = f"full_data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        elif format == "excel":
            content = export_full_data_excel(lang=lang)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"full_data_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Full data export error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/bookings/template")
async def download_import_template(
    format: str = Query("csv"),
    lang: str = Query("en"),
    session_token: Optional[str] = Cookie(None)
):
    """Скачать шаблон для импорта записей"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        if format == "csv":
            template = "instagram_id,name,phone,service,datetime,status,revenue\n"
            template += f"example_user_1,Анна Иванова,{SALON_PHONE_DEFAULT},Маникюр,2026-01-31 14:00,pending,150\n"
            template += "example_user_2,Мария Петрова,971507654321,Педикюр,2026-02-01 15:30,confirmed,200\n"
            
            return StreamingResponse(
                iter([template.encode('utf-8')]),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=bookings_template.csv"}
            )
        
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                return JSONResponse({"error": "Excel support not available"}, 
                                  status_code=500)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Записи"
            
            headers = ['instagram_id', 'name', 'phone', 'service', 'datetime', 
                      'status', 'revenue']
            ws.append(headers)
            
            ws.title = "Instruction"
            ws.append(["This is a template. Do not change headers."])
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return StreamingResponse(
                iter([buffer.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=bookings_template.xlsx"}
            )
        
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
            
    except Exception as e:
        log_error(f"Template download error: {e}", "export")
        return JSONResponse({"error": str(e)}, status_code=500)

# ===== ФУНКЦИИ ЭКСПОРТА ДАШБОРДА (Сводный отчет) =====

def export_dashboard_report_csv(stats, bot_analytics, bookings, lang='en'):
    """Экспорт сводного отчета дашборда в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 1. Заголовок
    writer.writerow([t(lang, 'adminPanel.dashboard.report', 'Отчет по салону'), datetime.now().strftime('%d.%m.%Y')])
    writer.writerow([])
    
    # 2. Основная статистика (KPI)
    writer.writerow([t(lang, 'adminPanel.dashboard.analytics', 'Аналитика')])
    
    # Клиенты
    writer.writerow([t(lang, 'adminPanel.dashboard.total_clients', 'Всего клиентов'), stats.get('clients', {}).get('total_active', 0)])
    writer.writerow([t(lang, 'adminPanel.dashboard.vip_clients', 'VIP-клиенты'), stats.get('clients', {}).get('vip', 0)])
    writer.writerow([t(lang, 'adminPanel.dashboard.new_clients', 'Новые клиенты'), stats.get('clients', {}).get('new', 0)])
    writer.writerow([t(lang, 'adminPanel.dashboard.active_clients', 'Активные клиенты'), stats.get('clients', {}).get('total_active', 0)])
    
    # Финансы
    writer.writerow([t(lang, 'adminPanel.dashboard.revenue', 'Доход'), stats.get('revenue', {}).get('total', 0)])
    writer.writerow([t(lang, 'adminPanel.dashboard.average_check', 'Средний чек'), stats.get('revenue', {}).get('average_check', 0)])
    
    # Эффективность
    writer.writerow([t(lang, 'adminPanel.dashboard.cancellations', 'Отмены'), f"{stats.get('bookings', {}).get('cancellation_rate', 0)}%"])
    writer.writerow([t(lang, 'adminPanel.dashboard.conversion', 'Конверсия'), f"{stats.get('bookings', {}).get('completion_rate', 0)}%"])
    
    writer.writerow([])
    
    # 3. Бот аналитика
    if bot_analytics:
        writer.writerow([t(lang, 'adminPanel.dashboard.bot_analytics', 'Бот Аналитика')])
        writer.writerow([t(lang, 'adminPanel.dashboard.active_sessions', 'Активных сессий'), bot_analytics.get('total_sessions', 0)])
        writer.writerow([t(lang, 'adminPanel.dashboard.conversion', 'Конверсия'), f"{bot_analytics.get('conversion_rate', 0)}%"])
        writer.writerow([])

    # 4. Последние записи
    writer.writerow([t(lang, 'adminPanel.dashboard.latest_bookings', 'Последние записи')])
    headers = [
        t(lang, 'adminPanel.dashboard.id', 'ID'),
        t(lang, 'adminPanel.dashboard.client', 'Клиент'),
        t(lang, 'common.service', 'Услуга'),
        t(lang, 'adminPanel.dashboard.datetime', 'Дата/Время'),
        t(lang, 'common.phone', 'Телефон'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.revenue', 'Доход'),
    ]
    writer.writerow(headers)
    
    for b in bookings[:50]: # Ограничим 50 последними для отчета
        try:
            date_obj = datetime.fromisoformat(b[3])
            date_str = date_obj.strftime('%d.%m.%Y %H:%M')
        except:
            date_str = str(b[3])
            
        writer.writerow([
            b[0], 
            b[5] or '', 
            b[2] or '', 
            date_str, 
            b[4] or '', 
            t(lang, f'adminPanel.dashboard.status_{b[6]}', b[6]),
            b[8] if len(b) > 8 else 0
        ])

    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_dashboard_report_excel(stats, bot_analytics, bookings, lang='en'):
    """Экспорт сводного отчета дашборда в Excel"""
    if not EXCEL_AVAILABLE:
        raise Exception("Excel экспорт недоступен")
    
    wb = Workbook()
    
    # --- Лист 1: Обзор (Overview) ---
    ws = wb.active
    ws.title = t(lang, 'adminPanel.dashboard.report', 'Обзор')
    
    # Стили
    header_font = Font(bold=True, size=14, color="EC4899")
    sub_header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    
    # Заголовок
    ws['A1'] = t(lang, 'adminPanel.dashboard.report', 'Отчет по салону')
    ws['A1'].font = header_font
    ws['B1'] = datetime.now().strftime('%d.%m.%Y')
    
    row = 3
    
    # KPI Stats Grid
    stats_to_show = [
        (t(lang, 'adminPanel.dashboard.total_clients', 'Всего клиентов'), stats.get('clients', {}).get('total_active', 0)),
        (t(lang, 'adminPanel.dashboard.vip_clients', 'VIP-клиенты'), stats.get('clients', {}).get('vip', 0)),
        (t(lang, 'adminPanel.dashboard.new_clients', 'Новые клиенты'), stats.get('clients', {}).get('new', 0)),
        (t(lang, 'adminPanel.dashboard.active_clients', 'Активные клиенты'), stats.get('clients', {}).get('total_active', 0)),
        (t(lang, 'adminPanel.dashboard.revenue', 'Доход'), f"{stats.get('revenue', {}).get('total', 0)} AED"),
        (t(lang, 'adminPanel.dashboard.average_check', 'Средний чек'), f"{stats.get('revenue', {}).get('average_check', 0)} AED"),
        (t(lang, 'adminPanel.dashboard.cancellations', 'Отмены'), f"{stats.get('bookings', {}).get('cancellation_rate', 0)}%"),
        (t(lang, 'adminPanel.dashboard.conversion', 'Конверсия'), f"{stats.get('bookings', {}).get('completion_rate', 0)}%")
    ]
    
    for label, value in stats_to_show:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value)
        row += 1
        
    row += 2
    
    # Бот аналитика
    if bot_analytics:
        ws.cell(row=row, column=1, value=t(lang, 'adminPanel.dashboard.bot_analytics', 'Бот Аналитика')).font = sub_header_font
        row += 1
        ws.cell(row=row, column=1, value=t(lang, 'adminPanel.dashboard.active_sessions', 'Активных сессий'))
        ws.cell(row=row, column=2, value=bot_analytics.get('total_sessions', 0))
        row += 1
        ws.cell(row=row, column=1, value=t(lang, 'adminPanel.dashboard.conversion', 'Конверсия'))
        ws.cell(row=row, column=2, value=f"{bot_analytics.get('conversion_rate', 0)}%")
        row += 2

    # --- Лист 2: Записи (Bookings) ---
    ws_bookings = wb.create_sheet(title=t(lang, 'adminPanel.dashboard.bookings', 'Записи'))
    
    headers = [
        t(lang, 'adminPanel.dashboard.id', 'ID'),
        t(lang, 'adminPanel.dashboard.client', 'Клиент'),
        t(lang, 'common.service', 'Услуга'),
        t(lang, 'adminPanel.dashboard.datetime', 'Дата/Время'),
        t(lang, 'common.phone', 'Телефон'),
        t(lang, 'common.status', 'Статус'),
        t(lang, 'adminPanel.dashboard.revenue', 'Доход'),
    ]
    ws_bookings.append(headers)
    
    # Стили заголовка таблицы
    fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws_bookings[1]:
        cell.fill = fill
        cell.font = font
    
    for b in bookings:
        try:
            date_obj = datetime.fromisoformat(b[3])
            date_str = date_obj.strftime('%d.%m.%Y %H:%M')
        except:
            date_str = str(b[3])
            
        ws_bookings.append([
            b[0], 
            b[5] or '', 
            b[2] or '', 
            date_str, 
            b[4] or '', 
            t(lang, f'adminPanel.dashboard.status_{b[6]}', b[6]),
            b[8] if len(b) > 8 else 0
        ])

    # Автоширина колонок
    for sheet in [ws, ws_bookings]:
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def export_dashboard_report_pdf(stats, bot_analytics, bookings, lang='en'):
    """Экспорт сводного отчета дашборда в PDF (Premium Style)"""
    if not PDF_AVAILABLE:
        raise Exception("PDF экспорт недоступен")
        
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                          rightMargin=40, leftMargin=40, 
                          topMargin=40, bottomMargin=40)
    elements = []
    
    # Регистрация шрифтов
    fontName = register_fonts()
    
    # Цветовая палитра
    PRIMARY_COLOR = colors.HexColor('#EC4899') # Pink-500
    SECONDARY_COLOR = colors.HexColor('#BE185D') # Pink-700
    TEXT_COLOR = colors.HexColor('#1F2937') # Gray-800
    LIGHT_GRAY = colors.HexColor('#F9FAFB') # Gray-50
    BORDER_COLOR = colors.HexColor('#E5E7EB') # Gray-200
    ACCENT_BG = colors.HexColor('#FDF2F8') # Pink-50
    
    # Стили
    styles = getSampleStyleSheet()
    
    # Стиль Заголовка
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontName=fontName,
        fontSize=24,
        textColor=PRIMARY_COLOR,
        alignment=TA_CENTER,
        spaceAfter=10,
        leading=30
    )
    
    # Стиль Подзаголовка (Дата)
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontName=fontName,
        fontSize=10,
        textColor=colors.HexColor('#6B7280'), # Gray-500
        alignment=TA_CENTER,
        spaceAfter=30
    )
    
    # Стиль Заголовков Разделов
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontName=fontName,
        fontSize=16,
        textColor=TEXT_COLOR,
        spaceBefore=20,
        spaceAfter=15,
        borderWidth=0,
        borderColor=PRIMARY_COLOR
    )
    
    # --- HEADER ---
    salon = get_salon_settings()
    salon_name = salon.get('name', 'Salon')
    elements.append(Paragraph(t(lang, 'adminPanel.dashboard.report', f'{salon_name} Report'), title_style))
    elements.append(Paragraph(f"{t(lang, 'adminPanel.dashboard.datetime', 'Generated')}: {datetime.now().strftime('%d.%m.%Y %H:%M')}", subtitle_style))
    
    # --- KPI ANALYTICS SECTION ---
    elements.append(Paragraph(t(lang, 'adminPanel.dashboard.analytics', 'Analytics Overview'), section_style))
    
    # Подготовка данных для таблицы KPI (2 колонки для компактности)
    # Формат: [Метка, Значение]
    kpi_data = [
        [t(lang, 'adminPanel.dashboard.total_clients', 'Всего клиентов'), str(stats.get('clients', {}).get('total_active', 0))],
        [t(lang, 'adminPanel.dashboard.vip_clients', 'VIP-клиенты'), str(stats.get('clients', {}).get('vip', 0))],
        [t(lang, 'adminPanel.dashboard.new_clients', 'Новые клиенты'), str(stats.get('clients', {}).get('new', 0))],
        [t(lang, 'adminPanel.dashboard.active_clients', 'Активные клиенты'), str(stats.get('clients', {}).get('total_active', 0))],
        [t(lang, 'adminPanel.dashboard.revenue', 'Доход'), f"{stats.get('revenue', {}).get('total', 0)} AED"],
        [t(lang, 'adminPanel.dashboard.average_check', 'Средний чек'), f"{stats.get('revenue', {}).get('average_check', 0)} AED"],
        [t(lang, 'adminPanel.dashboard.cancellations', 'Отмены'), f"{stats.get('bookings', {}).get('cancellation_rate', 0)}%"],
        [t(lang, 'adminPanel.dashboard.conversion', 'Конверсия'), f"{stats.get('bookings', {}).get('completion_rate', 0)}%"]
    ]

    # Стилизация таблицы KPI
    t_kpi = Table(kpi_data, colWidths=[300, 150])
    t_kpi.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), fontName),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4B5563')), # Muted labels
        ('TEXTCOLOR', (1, 0), (1, -1), TEXT_COLOR), # Darker values
        ('FONTNAME', (1, 0), (1, -1), fontName), # Bold values? No font variant loaded, rely on color
        
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        
        ('PADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        
        # Горизонтальные линии
        ('LINEBELOW', (0, 0), (-1, -2), 1, BORDER_COLOR),
        ('LINEBELOW', (0, -1), (-1, -1), 2, PRIMARY_COLOR), # Bottom accent line
        
        # Чередование фона
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, ACCENT_BG]),
    ]))
    elements.append(t_kpi)
    elements.append(Spacer(1, 20))

    # --- BOT ANALYTICS SECTION ---
    if bot_analytics:
        elements.append(Paragraph(t(lang, 'adminPanel.dashboard.bot_analytics', 'Bot Analytics'), section_style))
        
        bot_data = [
            [t(lang, 'adminPanel.dashboard.active_sessions', 'Active Sessions'), str(bot_analytics.get('total_sessions', 0))],
            [t(lang, 'adminPanel.dashboard.conversion', 'Conversion'), f"{bot_analytics.get('conversion_rate', 0)}%"]
        ]
        
        t_bot = Table(bot_data, colWidths=[300, 150])
        t_bot.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), fontName),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#4B5563')),
            ('TEXTCOLOR', (1, 0), (1, -1), TEXT_COLOR),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('PADDING', (0, 0), (-1, -1), 10),
            ('LINEBELOW', (0, 0), (-1, -2), 1, BORDER_COLOR),
            ('LINEBELOW', (0, -1), (-1, -1), 2, PRIMARY_COLOR),
             ('ROWBACKGROUNDS', (0, 0), (-1, -1), [colors.white, ACCENT_BG]),
        ]))
        elements.append(t_bot)
        elements.append(Spacer(1, 25))
        
    # --- RECENT BOOKINGS SECTION ---
    # Page Break if needed? No, flowable is better.
    elements.append(Paragraph(t(lang, 'adminPanel.dashboard.latest_bookings', 'Recent Bookings'), section_style))
    
    # Headers
    booking_headers = [
        t(lang, 'adminPanel.dashboard.client', 'Client'),
        t(lang, 'common.service', 'Service'),
        t(lang, 'adminPanel.dashboard.datetime', 'Date'),
        t(lang, 'common.status', 'Status'),
        t(lang, 'adminPanel.dashboard.revenue', 'Price')
    ]
    
    data_bookings = [booking_headers]
    
    for b in bookings[:50]: # Top 50
        try:
            date_obj = datetime.fromisoformat(b[3])
            date_str = date_obj.strftime('%d.%m %H:%M')
        except:
            date_str = str(b[3])[:11]
            
        status_text = t(lang, f'adminPanel.dashboard.status_{b[6]}', b[6])
        
        data_bookings.append([
            (b[5] or '')[:20],
            (b[2] or '')[:20],
            date_str,
            status_text,
            str(b[8])
        ])
        
    # Table Styling
    col_widths = [140, 140, 90, 90, 70]
    t_bookings = Table(data_bookings, colWidths=col_widths, repeatRows=1) # Repeat header on new page
    
    t_bookings.setStyle(TableStyle([
        # Header Row
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY_COLOR),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), fontName),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
        ('PADDING', (0, 0), (-1, 0), 12),
        
        # Content Rows
        ('FONTNAME', (0, 1), (-1, -1), fontName),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TEXTCOLOR', (0, 1), (-1, -1), TEXT_COLOR),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'), # Client, Service left
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'), # Date, Status, Revenue center/right?
        
        # Borders & Spacing
        ('GRID', (0, 1), (-1, -1), 0.5, BORDER_COLOR),
        ('BOX', (0, 0), (-1, -1), 1, PRIMARY_COLOR),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GRAY]),
        ('PADDING', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(t_bookings)
    
    # Footer info?
    elements.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontName=fontName,
        fontSize=8,
        textColor=colors.gray,
        alignment=TA_CENTER
    )
    elements.append(Paragraph(f"{salon_name} • {datetime.now().year}", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()