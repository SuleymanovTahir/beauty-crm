"""
API endpoints for admin panel features: Challenges, Referrals, Loyalty, Notifications, Gallery
"""
from fastapi import APIRouter, Request, Cookie, HTTPException, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional, List
from db.connection import get_db_connection
from utils.utils import require_auth
from utils.logger import log_info, log_error
from datetime import datetime
import os
import uuid
import json

router = APIRouter(tags=["Admin Features"])

# ============================================================================
# DASHBOARD STATS
# ============================================================================

@router.get("/admin/stats")
async def get_admin_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику для админ дашборда"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total clients (not users)
        c.execute("SELECT COUNT(*) FROM clients")
        total_users = c.fetchone()[0]

        # Active challenges
        c.execute("SELECT COUNT(*) FROM active_challenges WHERE is_active = TRUE")
        active_challenges = c.fetchone()[0]

        # Total loyalty points issued
        c.execute("SELECT COALESCE(SUM(points), 0) FROM loyalty_transactions WHERE points > 0")
        total_loyalty_points = c.fetchone()[0]

        # Total referrals
        c.execute("SELECT COUNT(*) FROM client_referrals")
        total_referrals = c.fetchone()[0]

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_users": total_users,
                "active_challenges": active_challenges,
                "total_loyalty_points": int(total_loyalty_points),
                "total_referrals": total_referrals
            }
        }
    except Exception as e:
        log_error(f"Error getting admin stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

# ============================================================================
# CHALLENGES API
# ============================================================================

@router.get("/admin/challenges")
async def get_admin_challenges(session_token: Optional[str] = Cookie(None)):
    """Получить все челленджи для админки"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Получаем все челленджи с информацией об участниках
        c.execute("""
            SELECT
                c.id,
                c.title,
                c.description,
                c.challenge_type as type,
                c.target_value,
                c.bonus_points as reward_points,
                c.start_date,
                c.end_date,
                c.is_active,
                COUNT(DISTINCT cp.client_id) as participants,
                COUNT(DISTINCT CASE WHEN cp.is_completed THEN cp.client_id END) as completions
            FROM active_challenges c
            LEFT JOIN challenge_progress cp ON c.id = cp.challenge_id
            GROUP BY c.id
            ORDER BY c.created_at DESC
        """)

        challenges = []
        for row in c.fetchall():
            # Определяем статус
            status = 'upcoming'
            if row[8]:  # is_active
                now = datetime.now().date()
                start_date = row[6]
                end_date = row[7]
                if start_date and end_date:
                    if start_date <= now <= end_date:
                        status = 'active'
                    elif now < start_date:
                        status = 'upcoming'
                    else:
                        status = 'completed'
                else:
                    status = 'active'
            else:
                status = 'completed'

            challenges.append({
                "id": str(row[0]),
                "title": row[1] or "Challenge",
                "description": row[2] or "",
                "type": row[3] or "visits",
                "target_value": row[4] or 0,
                "reward_points": row[5] or 0,
                "start_date": row[6].isoformat() if row[6] else datetime.now().isoformat(),
                "end_date": row[7].isoformat() if row[7] else datetime.now().isoformat(),
                "status": status,
                "participants": row[9] or 0,
                "completions": row[10] or 0
            })

        conn.close()
        return {"success": True, "challenges": challenges}
    except Exception as e:
        log_error(f"Error getting admin challenges: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/challenges")
async def create_admin_challenge(request: Request, session_token: Optional[str] = Cookie(None)):
    """Создать новый челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO active_challenges (
                title, description,
                challenge_type, target_value, bonus_points,
                start_date, end_date, is_active
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data.get("title", ""),
            data.get("description", ""),
            data.get("type", "visits"),
            data.get("target_value", 0),
            data.get("reward_points", 0),
            data.get("start_date"),
            data.get("end_date"),
            True
        ))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {"success": True, "id": new_id}
    except Exception as e:
        log_error(f"Error creating challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/challenges/{challenge_id}")
async def update_admin_challenge(challenge_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            UPDATE active_challenges SET
                title = %s,
                description = %s,
                challenge_type = %s,
                target_value = %s,
                bonus_points = %s,
                start_date = %s,
                end_date = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (
            data.get("title", ""),
            data.get("description", ""),
            data.get("type", "visits"),
            data.get("target_value", 0),
            data.get("reward_points", 0),
            data.get("start_date"),
            data.get("end_date"),
            challenge_id
        ))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/challenges/{challenge_id}")
async def delete_admin_challenge(challenge_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить челлендж"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Удаляем прогресс челленджа
        c.execute("DELETE FROM challenge_progress WHERE challenge_id = %s", (challenge_id,))
        # Удаляем сам челлендж
        c.execute("DELETE FROM active_challenges WHERE id = %s", (challenge_id,))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting challenge: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/challenges/{challenge_id}/check-progress")
async def check_challenge_progress(
    challenge_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Проверить и обновить прогресс челленджа для всех участников"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Получаем информацию о челлендже
        c.execute("""
            SELECT challenge_type, target_value, bonus_points, start_date, end_date
            FROM active_challenges
            WHERE id = %s AND is_active = TRUE
        """, (challenge_id,))

        challenge = c.fetchone()
        if not challenge:
            raise HTTPException(status_code=404, detail="Challenge not found or inactive")

        challenge_type, target_value, bonus_points, start_date, end_date = challenge

        # Проверяем прогресс в зависимости от типа челленджа
        if challenge_type == 'visits':
            # Считаем количество визитов (завершенных броней) для каждого клиента
            c.execute("""
                SELECT b.client_id, COUNT(*) as visit_count
                FROM bookings b
                WHERE b.status = 'completed'
                  AND (%s IS NULL OR b.booking_date >= %s)
                  AND (%s IS NULL OR b.booking_date <= %s)
                GROUP BY b.client_id
                HAVING COUNT(*) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        elif challenge_type == 'spending':
            # Считаем общую сумму трат для каждого клиента
            c.execute("""
                SELECT b.client_id, COALESCE(SUM(b.total_price), 0) as total_spent
                FROM bookings b
                WHERE b.status = 'completed'
                  AND (%s IS NULL OR b.booking_date >= %s)
                  AND (%s IS NULL OR b.booking_date <= %s)
                GROUP BY b.client_id
                HAVING COALESCE(SUM(b.total_price), 0) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        elif challenge_type == 'referrals':
            # Считаем количество успешных рефералов для каждого клиента
            c.execute("""
                SELECT r.referrer_id as client_id, COUNT(*) as referral_count
                FROM client_referrals r
                WHERE r.status = 'completed'
                  AND (%s IS NULL OR r.created_at >= %s)
                  AND (%s IS NULL OR r.created_at <= %s)
                GROUP BY r.referrer_id
                HAVING COUNT(*) >= %s
            """, (start_date, start_date, end_date, end_date, target_value))

        else:
            # Для типа 'services' нужна дополнительная логика
            conn.close()
            return {"success": False, "message": "Services type challenges not yet implemented"}

        completed_clients = c.fetchall()
        updated_count = 0

        for client_row in completed_clients:
            client_id = client_row[0]
            current_value = client_row[1]

            # Проверяем, есть ли уже запись прогресса
            c.execute("""
                SELECT id, is_completed
                FROM challenge_progress
                WHERE challenge_id = %s AND client_id = %s
            """, (challenge_id, client_id))

            progress = c.fetchone()

            if progress:
                progress_id, is_completed = progress
                if not is_completed:
                    # Обновляем прогресс и помечаем как завершенный
                    c.execute("""
                        UPDATE challenge_progress
                        SET current_value = %s,
                            is_completed = TRUE,
                            completed_at = CURRENT_TIMESTAMP,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s
                    """, (current_value, progress_id))

                    # Начисляем бонусные баллы клиенту
                    c.execute("""
                        UPDATE clients
                        SET loyalty_points = COALESCE(loyalty_points, 0) + %s
                        WHERE instagram_id = %s
                    """, (bonus_points, client_id))

                    # Создаем транзакцию лояльности
                    c.execute("""
                        INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
                        VALUES (%s, %s, 'earn', %s)
                    """, (client_id, bonus_points, f"Challenge completed: {challenge_type}"))

                    updated_count += 1
            else:
                # Создаем новую запись прогресса как завершенную
                c.execute("""
                    INSERT INTO challenge_progress (
                        challenge_id, client_id, current_value,
                        is_completed, completed_at
                    ) VALUES (%s, %s, %s, TRUE, CURRENT_TIMESTAMP)
                """, (challenge_id, client_id, current_value))

                # Начисляем бонусные баллы
                c.execute("""
                    UPDATE clients
                    SET loyalty_points = COALESCE(loyalty_points, 0) + %s
                    WHERE instagram_id = %s
                """, (bonus_points, client_id))

                # Создаем транзакцию
                c.execute("""
                    INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
                    VALUES (%s, %s, 'earn', %s)
                """, (client_id, bonus_points, f"Challenge completed: {challenge_type}"))

                updated_count += 1

        conn.commit()
        conn.close()

        return {
            "success": True,
            "updated_count": updated_count,
            "message": f"Updated {updated_count} clients who completed the challenge"
        }

    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Error checking challenge progress: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# REFERRAL PROGRAM API
# ============================================================================

@router.get("/admin/referrals")
async def get_admin_referrals(
    session_token: Optional[str] = Cookie(None),
    status: Optional[str] = None,
    sort_by: Optional[str] = "date",
    order: Optional[str] = "desc"
):
    """Получить все рефералы с фильтрацией и сортировкой"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Построение WHERE клаузы
        where_clause = ""
        params = []
        if status:
            where_clause = "WHERE r.status = %s"
            params.append(status)

        # Определение сортировки
        sort_column = {
            "date": "r.created_at",
            "points": "r.points_awarded",
            "status": "r.status"
        }.get(sort_by, "r.created_at")

        sort_order = "DESC" if order == "desc" else "ASC"

        query = f"""
            SELECT
                r.id,
                c1.name as referrer_name,
                c1.email as referrer_email,
                c2.name as referred_name,
                c2.email as referred_email,
                r.status,
                r.points_awarded,
                r.created_at
            FROM client_referrals r
            LEFT JOIN clients c1 ON r.referrer_id = c1.instagram_id
            LEFT JOIN clients c2 ON r.referred_id = c2.instagram_id
            {where_clause}
            ORDER BY {sort_column} {sort_order}
        """

        c.execute(query, params)

        referrals = []
        for row in c.fetchall():
            referrals.append({
                "id": str(row[0]),
                "referrer_name": row[1] or "Unknown",
                "referrer_email": row[2] or "",
                "referred_name": row[3] or "Unknown",
                "referred_email": row[4] or "",
                "status": row[5] or "pending",
                "points_awarded": row[6] or 0,
                "created_at": row[7].isoformat() if row[7] else datetime.now().isoformat()
            })

        conn.close()
        return {"success": True, "referrals": referrals}
    except Exception as e:
        log_error(f"Error getting referrals: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/referrals/settings")
async def get_referral_settings(session_token: Optional[str] = Cookie(None)):
    """Получить настройки реферальной программы"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT setting_value FROM settings WHERE setting_key = 'referral_settings'
        """)

        result = c.fetchone()
        if result:
            import json
            settings = json.loads(result[0])
        else:
            settings = {
                "referrer_bonus": 500,
                "referred_bonus": 200,
                "min_purchase_amount": 0
            }

        conn.close()
        return {"success": True, "settings": settings}
    except Exception as e:
        log_error(f"Error getting referral settings: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/referrals/settings")
async def update_referral_settings(request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить настройки реферальной программы"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        import json
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO settings (setting_key, setting_value)
            VALUES ('referral_settings', %s)
            ON CONFLICT (setting_key) DO UPDATE SET setting_value = EXCLUDED.setting_value
        """, (json.dumps(data),))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating referral settings: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/referrals/stats")
async def get_referral_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику рефералов"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total referrals
        c.execute("SELECT COUNT(*) FROM client_referrals")
        total_referrals = c.fetchone()[0] or 0

        # Completed referrals
        c.execute("SELECT COUNT(*) FROM client_referrals WHERE status = 'completed'")
        completed_referrals = c.fetchone()[0] or 0

        # Points distributed
        c.execute("SELECT COALESCE(SUM(points_awarded), 0) FROM client_referrals WHERE status = 'completed'")
        points_distributed = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_referrals": total_referrals,
                "completed_referrals": completed_referrals,
                "points_distributed": points_distributed
            }
        }
    except Exception as e:
        log_error(f"Error getting referral stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# LOYALTY MANAGEMENT API
# ============================================================================

@router.get("/admin/loyalty/tiers")
async def get_loyalty_tiers(session_token: Optional[str] = Cookie(None)):
    """Получить уровни лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT id, name, min_points, discount, color
            FROM loyalty_tiers
            ORDER BY min_points ASC
        """)

        tiers = []
        for row in c.fetchall():
            tiers.append({
                "id": str(row[0]),
                "name": row[1],
                "min_points": row[2] or 0,
                "discount": row[3] or 0,
                "color": row[4] or "#CD7F32"
            })

        # Если тиров нет, создаем дефолтные
        if not tiers:
            default_tiers = [
                ('Bronze', 0, 0, '#CD7F32'),
                ('Silver', 1000, 5, '#C0C0C0'),
                ('Gold', 5000, 10, '#FFD700'),
                ('Platinum', 10000, 15, '#E5E4E2')
            ]

            for tier in default_tiers:
                c.execute("""
                    INSERT INTO loyalty_tiers (name, min_points, discount, color)
                    VALUES (%s, %s, %s, %s)
                    RETURNING id, name, min_points, discount, color
                """, tier)
                row = c.fetchone()
                tiers.append({
                    "id": str(row[0]),
                    "name": row[1],
                    "min_points": row[2],
                    "discount": row[3],
                    "color": row[4]
                })

            conn.commit()

        conn.close()
        return {"success": True, "tiers": tiers}
    except Exception as e:
        log_error(f"Error getting loyalty tiers: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/loyalty/tiers/{tier_id}")
async def update_loyalty_tier(tier_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Обновить уровень лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            UPDATE loyalty_tiers SET
                name = %s,
                min_points = %s,
                discount = %s,
                color = %s
            WHERE id = %s
        """, (
            data.get("name"),
            data.get("min_points"),
            data.get("discount"),
            data.get("color"),
            tier_id
        ))

        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        log_error(f"Error updating loyalty tier: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/loyalty/transactions")
async def get_loyalty_transactions(session_token: Optional[str] = Cookie(None)):
    """Получить историю транзакций лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT
                lt.id,
                c.name as client_name,
                c.email as client_email,
                lt.points,
                lt.transaction_type as type,
                lt.reason,
                lt.created_at
            FROM loyalty_transactions lt
            LEFT JOIN clients c ON lt.client_id = c.instagram_id
            ORDER BY lt.created_at DESC
            LIMIT 100
        """)

        transactions = []
        for row in c.fetchall():
            transactions.append({
                "id": str(row[0]),
                "client_name": row[1] or "Unknown",
                "client_email": row[2] or "",
                "points": row[3] or 0,
                "type": row[4] or "adjust",
                "reason": row[5] or "",
                "created_at": row[6].isoformat() if row[6] else datetime.now().isoformat()
            })

        conn.close()
        return {"success": True, "transactions": transactions}
    except Exception as e:
        log_error(f"Error getting loyalty transactions: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/loyalty/adjust-points")
async def adjust_loyalty_points(request: Request, session_token: Optional[str] = Cookie(None)):
    """Скорректировать баллы лояльности клиента"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        # Найти клиента по email
        c.execute("SELECT instagram_id, loyalty_points FROM clients WHERE email = %s", (data.get("client_email"),))
        result = c.fetchone()

        if not result:
            raise HTTPException(status_code=404, detail="Client not found")

        client_id = result[0]
        current_points = result[1] or 0
        points = data.get("points", 0)
        new_points = current_points + points

        # Обновить баллы клиента
        c.execute("UPDATE clients SET loyalty_points = %s WHERE instagram_id = %s", (new_points, client_id))

        # Создать транзакцию
        c.execute("""
            INSERT INTO loyalty_transactions (client_id, points, transaction_type, reason)
            VALUES (%s, %s, %s, %s)
        """, (client_id, points, 'adjust', data.get("reason", "Manual adjustment")))

        conn.commit()
        conn.close()

        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Error adjusting loyalty points: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/loyalty/stats")
async def get_loyalty_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику программы лояльности"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total points issued
        c.execute("SELECT COALESCE(SUM(points), 0) FROM loyalty_transactions WHERE points > 0")
        total_points_issued = c.fetchone()[0] or 0

        # Points redeemed
        c.execute("SELECT COALESCE(SUM(ABS(points)), 0) FROM loyalty_transactions WHERE points < 0")
        points_redeemed = c.fetchone()[0] or 0

        # Active members (clients with points > 0)
        c.execute("SELECT COUNT(*) FROM clients WHERE loyalty_points > 0")
        active_members = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_points_issued": total_points_issued,
                "points_redeemed": points_redeemed,
                "active_members": active_members
            }
        }
    except Exception as e:
        log_error(f"Error getting loyalty stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# NOTIFICATIONS DASHBOARD API
# ============================================================================

@router.get("/admin/notifications")
async def get_admin_notifications(session_token: Optional[str] = Cookie(None)):
    """Получить историю уведомлений"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT
                id, title, message, notification_type,
                recipients_count, sent_count, failed_count,
                status, created_at, sent_at
            FROM notification_history
            ORDER BY created_at DESC
            LIMIT 100
        """)

        notifications = []
        for row in c.fetchall():
            notifications.append({
                "id": str(row[0]),
                "title": row[1] or "",
                "message": row[2] or "",
                "type": row[3] or "push",
                "recipients": row[4] or 0,
                "sent_count": row[5] or 0,
                "failed_count": row[6] or 0,
                "status": row[7] or "pending",
                "created_at": row[8].isoformat() if row[8] else datetime.now().isoformat(),
                "sent_at": row[9].isoformat() if row[9] else None
            })

        conn.close()
        return {"success": True, "notifications": notifications}
    except Exception as e:
        log_error(f"Error getting notifications: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/notifications/templates")
async def get_notification_templates(session_token: Optional[str] = Cookie(None)):
    """Получить шаблоны уведомлений"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT id, name, title, message, notification_type
            FROM notification_templates
            ORDER BY created_at DESC
        """)

        templates = []
        for row in c.fetchall():
            templates.append({
                "id": str(row[0]),
                "name": row[1] or "",
                "title": row[2] or "",
                "message": row[3] or "",
                "type": row[4] or "push"
            })

        conn.close()
        return {"success": True, "templates": templates}
    except Exception as e:
        log_error(f"Error getting notification templates: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/notifications/templates")
async def create_notification_template(request: Request, session_token: Optional[str] = Cookie(None)):
    """Создать шаблон уведомления"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO notification_templates (name, title, message, notification_type)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (
            data.get("name"),
            data.get("title"),
            data.get("message"),
            data.get("type", "push")
        ))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {"success": True, "id": new_id}
    except Exception as e:
        log_error(f"Error creating notification template: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/notifications/send")
async def send_notification(request: Request, session_token: Optional[str] = Cookie(None)):
    """Отправить уведомление"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        # Определить целевую аудиторию
        target_segment = data.get("target_segment", "all")
        tier_filter = data.get("tier_filter", "")

        # Новые параметры фильтрации
        appointment_filter = data.get("appointment_filter", "")
        appointment_date = data.get("appointment_date", "")
        appointment_start_date = data.get("appointment_start_date", "")
        appointment_end_date = data.get("appointment_end_date", "")
        service_filter = data.get("service_filter", "")

        # Получить список клиентов
        # Выбираем instagram_id и telegram_id для отправки
        columns = "instagram_id, telegram_id"
        
        if target_segment == "all":
            c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")
        elif target_segment == "active":
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND is_active = TRUE")
        elif target_segment == "inactive":
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND is_active = FALSE")
        elif target_segment == "tier" and tier_filter:
            c.execute(f"SELECT {columns} FROM clients WHERE (telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%') AND loyalty_tier = %s", (tier_filter,))
        elif target_segment == "appointment-based":
            # Фильтр по датам записей
            if appointment_filter == "tomorrow":
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime::date = CURRENT_DATE + INTERVAL '1 day'
                      AND b.status IN ('pending', 'confirmed')
                """)
            elif appointment_filter == "specific_date" and appointment_date:
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime::date = %s::date
                      AND b.status IN ('pending', 'confirmed')
                """, (appointment_date,))
            elif appointment_filter == "date_range" and appointment_start_date and appointment_end_date:
                c.execute(f"""
                    SELECT DISTINCT c.instagram_id, c.telegram_id
                    FROM clients c
                    JOIN bookings b ON c.instagram_id = b.client_id
                    WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                      AND b.datetime BETWEEN %s AND %s
                      AND b.status IN ('pending', 'confirmed')
                """, (appointment_start_date, appointment_end_date))
            else:
                c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")
        elif target_segment == "service-based" and service_filter:
            # Фильтр по процедурам/услугам
            c.execute(f"""
                SELECT DISTINCT c.instagram_id, c.telegram_id
                FROM clients c
                JOIN bookings b ON c.instagram_id = b.client_id
                WHERE (c.telegram_id IS NOT NULL OR c.instagram_id LIKE 'telegram_%')
                  AND (b.service_name ILIKE %s OR b.service_id::text = %s)
            """, (f"%{service_filter}%", service_filter))
        else:
            c.execute(f"SELECT {columns} FROM clients WHERE telegram_id IS NOT NULL OR instagram_id LIKE 'telegram_%'")

        recipients = c.fetchall()
        recipients_count = len(recipients)

        # Параметры планирования
        scheduled = data.get("scheduled", False)
        schedule_date = data.get("schedule_date", "")
        schedule_time = data.get("schedule_time", "")
        repeat_enabled = data.get("repeat_enabled", False)
        repeat_interval = data.get("repeat_interval", "")
        repeat_end_date = data.get("repeat_end_date", "")

        # Подготовить datetime для планирования
        schedule_datetime = None
        if scheduled and schedule_date and schedule_time:
            schedule_datetime = f"{schedule_date} {schedule_time}:00"

        # Сохранить параметры фильтрации в JSON
        filter_params = {
            "target_segment": target_segment,
            "tier_filter": tier_filter,
            "appointment_filter": appointment_filter,
            "appointment_date": appointment_date,
            "appointment_start_date": appointment_start_date,
            "appointment_end_date": appointment_end_date,
            "service_filter": service_filter
        }

        # Определить статус уведомления
        status = "pending" if scheduled else "sent"

        # Создать запись в истории
        c.execute("""
            INSERT INTO notification_history (
                title, message, notification_type,
                recipients_count, status,
                scheduled, schedule_datetime,
                repeat_enabled, repeat_interval, repeat_end_date,
                target_segment, filter_params
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data.get("title"),
            data.get("message"),
            data.get("type", "push"),
            recipients_count,
            status,
            scheduled,
            schedule_datetime,
            repeat_enabled,
            repeat_interval if repeat_enabled else None,
            repeat_end_date if repeat_enabled else None,
            target_segment,
            json.dumps(filter_params)
        ))

        notification_id = c.fetchone()[0]

        # Если не запланировано - отправить сразу
        # Если не запланировано - отправить сразу
        if not scheduled:
            # Реальная отправка уведомлений через Telegram
            from integrations.telegram_bot import telegram_bot
            
            sent_count = 0
            failed_count = 0
            message_text = data.get("message")
            
            for row in recipients:
                inst_id = row[0]
                tg_id = row[1] if len(row) > 1 else None
                
                chat_id = None
                if tg_id:
                     chat_id = tg_id
                elif inst_id and str(inst_id).startswith('telegram_'):
                     chat_id = str(inst_id).replace('telegram_', '')
                
                if chat_id:
                    try:
                        res = telegram_bot.send_message(int(chat_id), message_text)
                        if res.get("ok"):
                            sent_count += 1
                        else:
                            failed_count += 1
                            log_error(f"Telegram send failed for {chat_id}: {res}", "api")
                    except Exception as e:
                        failed_count += 1
                        log_error(f"Error sending to {chat_id}: {e}", "api")

            c.execute("""
                UPDATE notification_history
                SET sent_count = %s, failed_count = %s, sent_at = CURRENT_TIMESTAMP, status = 'sent'
                WHERE id = %s
            """, (sent_count, failed_count, notification_id))

        conn.commit()
        conn.close()

        return {"success": True, "id": notification_id, "scheduled": scheduled}
    except Exception as e:
        log_error(f"Error sending notification: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


# ============================================================================
# PHOTO GALLERY API
# ============================================================================

@router.get("/admin/gallery/photos")
async def get_gallery_photos(session_token: Optional[str] = Cookie(None)):
    """Получить все фото галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            SELECT
                id, url, title, description, category,
                uploaded_by, created_at, is_featured, views,
                before_photo_url, after_photo_url, client_id, is_visible
            FROM gallery_photos
            ORDER BY created_at DESC
        """)

        photos = []
        for row in c.fetchall():
            photos.append({
                "id": str(row[0]),
                "url": row[1] or "",
                "title": row[2] or "",
                "description": row[3] or "",
                "category": row[4] or "other",
                "uploaded_by": row[5] or "Admin",
                "created_at": row[6].isoformat() if row[6] else datetime.now().isoformat(),
                "is_featured": row[7] or False,
                "views": row[8] or 0,
                "before_photo_url": row[9] or "",
                "after_photo_url": row[10] or "",
                "client_id": row[11] or "",
                "is_visible": row[12] if len(row) > 12 else True
            })

        conn.close()
        return {"success": True, "photos": photos}
    except Exception as e:
        log_error(f"Error getting gallery photos: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/gallery/categories")
async def get_gallery_categories(session_token: Optional[str] = Cookie(None)):
    """Получить список категорий из базы данных"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Получаем уникальные категории из services таблицы
        c.execute("""
            SELECT DISTINCT category
            FROM services
            WHERE category IS NOT NULL AND category != ''
            ORDER BY category
        """)

        categories = []
        for row in c.fetchall():
            if row[0]:
                categories.append({
                    "value": row[0].lower().replace(' ', '_'),
                    "label": row[0]
                })

        # Добавляем дополнительные категории если нужно
        default_categories = [
            {"value": "portfolio", "label": "Портфолио"},
            {"value": "interior", "label": "Интерьер"},
            {"value": "other", "label": "Другое"}
        ]

        # Объединяем, избегая дубликатов
        existing_values = {cat["value"] for cat in categories}
        for cat in default_categories:
            if cat["value"] not in existing_values:
                categories.append(cat)

        conn.close()
        return {"success": True, "categories": categories}
    except Exception as e:
        log_error(f"Error getting gallery categories: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/admin/gallery/stats")
async def get_gallery_stats(session_token: Optional[str] = Cookie(None)):
    """Получить статистику галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Total photos
        c.execute("SELECT COUNT(*) FROM gallery_photos")
        total_photos = c.fetchone()[0] or 0

        # Total views
        c.execute("SELECT COALESCE(SUM(views), 0) FROM gallery_photos")
        total_views = c.fetchone()[0] or 0

        # Featured count
        c.execute("SELECT COUNT(*) FROM gallery_photos WHERE is_featured = TRUE")
        featured_count = c.fetchone()[0] or 0

        # Recent uploads (last 7 days)
        c.execute("SELECT COUNT(*) FROM gallery_photos WHERE created_at >= CURRENT_DATE - INTERVAL '7 days'")
        recent_uploads = c.fetchone()[0] or 0

        conn.close()

        return {
            "success": True,
            "stats": {
                "total_photos": total_photos,
                "total_views": total_views,
                "featured_count": featured_count,
                "recent_uploads": recent_uploads
            }
        }
    except Exception as e:
        log_error(f"Error getting gallery stats: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/admin/gallery/photos")
async def upload_gallery_photo(
    file: UploadFile = File(None),
    before_photo: UploadFile = File(None),
    after_photo: UploadFile = File(None),
    title: str = Form(...),
    description: str = Form(""),
    category: str = Form("other"),
    is_featured: str = Form("false"),
    client_id: Optional[str] = Form(None),
    session_token: Optional[str] = Cookie(None)
):
    """Загрузить фото в галерею (обычное или до/после)"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        upload_dir = "uploads/gallery"
        os.makedirs(upload_dir, exist_ok=True)

        photo_url = ""
        before_photo_url = ""
        after_photo_url = ""

        # Обработка обычного фото
        if file:
            file_extension = file.filename.split('.')[-1] if '.' in file.filename else 'jpg'
            unique_filename = f"{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await file.read()
                f.write(content)

            photo_url = f"/uploads/gallery/{unique_filename}"

        # Обработка фото "до"
        if before_photo:
            file_extension = before_photo.filename.split('.')[-1] if '.' in before_photo.filename else 'jpg'
            unique_filename = f"before_{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await before_photo.read()
                f.write(content)

            before_photo_url = f"/uploads/gallery/{unique_filename}"

        # Обработка фото "после"
        if after_photo:
            file_extension = after_photo.filename.split('.')[-1] if '.' in after_photo.filename else 'jpg'
            unique_filename = f"after_{uuid.uuid4()}.{file_extension}"
            file_path = os.path.join(upload_dir, unique_filename)

            with open(file_path, "wb") as f:
                content = await after_photo.read()
                f.write(content)

            after_photo_url = f"/uploads/gallery/{unique_filename}"

        # Если нет обычного фото, но есть фото "после", используем его как основное
        if not photo_url and after_photo_url:
            photo_url = after_photo_url

        # Сохранить в БД
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            INSERT INTO gallery_photos (
                url, title, description, category,
                uploaded_by, is_featured, client_id,
                before_photo_url, after_photo_url
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            photo_url,
            title,
            description,
            category,
            user.get("email", "admin"),
            is_featured.lower() == "true",
            client_id if client_id else None,
            before_photo_url if before_photo_url else None,
            after_photo_url if after_photo_url else None
        ))

        new_id = c.fetchone()[0]
        conn.commit()
        conn.close()

        return {
            "success": True,
            "id": new_id,
            "url": photo_url,
            "before_photo_url": before_photo_url,
            "after_photo_url": after_photo_url
        }
    except Exception as e:
        log_error(f"Error uploading gallery photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.delete("/admin/gallery/photos/{photo_id}")
async def delete_gallery_photo(photo_id: int, session_token: Optional[str] = Cookie(None)):
    """Удалить фото из галереи"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        conn = get_db_connection()
        c = conn.cursor()

        # Получить URL фото
        c.execute("SELECT url, before_photo_url, after_photo_url FROM gallery_photos WHERE id = %s", (photo_id,))
        result = c.fetchone()

        if result:
            photo_url = result[0]
            before_url = result[1]
            after_url = result[2]
            
            # Helper to delete file
            def delete_file(path_url):
                if path_url and path_url.startswith("/uploads/"):
                    file_path = path_url[1:]  # Remove leading /
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            log_info(f"Deleted file: {file_path}", "api")
                        except Exception as e:
                            log_error(f"Error deleting file {file_path}: {e}", "api")

            # Удаляем основное фото
            delete_file(photo_url)
            
            # Удаляем фото до/после если есть (и если отличаются от основного)
            if before_url and before_url != photo_url:
                delete_file(before_url)
            
            if after_url and after_url != photo_url:
                delete_file(after_url)

        # Удалить из БД
        c.execute("DELETE FROM gallery_photos WHERE id = %s", (photo_id,))
        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error deleting gallery photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/gallery/photos/{photo_id}/featured")
async def toggle_featured_photo(photo_id: int, request: Request, session_token: Optional[str] = Cookie(None)):
    """Переключить статус избранного для фото"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            UPDATE gallery_photos
            SET is_featured = %s
            WHERE id = %s
        """, (data.get("is_featured", False), photo_id))

        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error toggling featured photo: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/admin/gallery/photos/{photo_id}/visibility")
async def toggle_gallery_photo_visibility(
    photo_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Переключить видимость фото"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")

    try:
        data = await request.json()
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("""
            UPDATE gallery_photos
            SET is_visible = %s
            WHERE id = %s
        """, (data.get("is_visible", True), photo_id))

        conn.commit()
        conn.close()

        return {"success": True}
    except Exception as e:
        log_error(f"Error toggling photo visibility: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)


@router.post("/admin/export-report")
async def export_report(request: Request, session_token: Optional[str] = Cookie(None)):
    """Экспорт отчета в CSV/PDF/Excel"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "director", "manager"]:
        raise HTTPException(status_code=403, detail="Forbidden")
        
    try:
        data = await request.json()
        export_format = data.get('format', 'csv').lower()
        
        # Получаем данные
        from db import get_stats
        stats = get_stats()
        
        report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        period_str = f"{data.get('start_date', 'All')} - {data.get('end_date', 'Now')}"
        
        if export_format == 'csv':
            import io
            import csv
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            writer.writerow(["Report Generated", report_date])
            writer.writerow(["Period", period_str])
            writer.writerow([])
            writer.writerow(["Metric", "Value"])
            
            if stats:
                for k, v in stats.items():
                    if isinstance(v, (int, float, str)):
                        writer.writerow([k.replace('_', ' ').title(), v])
                    elif isinstance(v, dict) and 'value' in v:
                        writer.writerow([k.replace('_', ' ').title(), v['value']])

            output.seek(0)
            
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.csv"
                }
            )
            
        elif export_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment
                import io
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"
                
                # Header
                ws['A1'] = "Report Generated"
                ws['B1'] = report_date
                ws['A1'].font = Font(bold=True)
                
                ws['A2'] = "Period"
                ws['B2'] = period_str
                ws['A2'].font = Font(bold=True)
                
                # Data headers
                ws['A4'] = "Metric"
                ws['B4'] = "Value"
                ws['A4'].font = Font(bold=True)
                ws['B4'].font = Font(bold=True)
                
                row = 5
                if stats:
                    for k, v in stats.items():
                        if isinstance(v, (int, float, str)):
                            ws[f'A{row}'] = k.replace('_', ' ').title()
                            ws[f'B{row}'] = v
                            row += 1
                        elif isinstance(v, dict) and 'value' in v:
                            ws[f'A{row}'] = k.replace('_', ' ').title()
                            ws[f'B{row}'] = v['value']
                            row += 1
                
                # Auto-size columns (openpyxl: type stubs may not expose column_dimensions)
                ws.column_dimensions['A'].width = 30  # type: ignore[union-attr]
                ws.column_dimensions['B'].width = 20  # type: ignore[union-attr]
                
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                return StreamingResponse(
                    iter([output.getvalue()]),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={
                        "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.xlsx"
                    }
                )
            except ImportError:
                log_error("openpyxl not installed", "api")
                raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl.")
                
        elif export_format == 'pdf':
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.lib import colors
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                import io
                
                buffer = io.BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=A4)
                elements = []
                styles = getSampleStyleSheet()
                
                # Title
                title = Paragraph(f"<b>Report Generated: {report_date}</b>", styles['Heading1'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Period
                period = Paragraph(f"<b>Period:</b> {period_str}", styles['Normal'])
                elements.append(period)
                elements.append(Spacer(1, 20))
                
                # Data table
                data = [['Metric', 'Value']]
                if stats:
                    for k, v in stats.items():
                        if isinstance(v, (int, float, str)):
                            data.append([k.replace('_', ' ').title(), str(v)])
                        elif isinstance(v, dict) and 'value' in v:
                            data.append([k.replace('_', ' ').title(), str(v['value'])])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                doc.build(elements)
                
                buffer.seek(0)
                return StreamingResponse(
                    iter([buffer.getvalue()]),
                    media_type="application/pdf",
                    headers={
                        "Content-Disposition": f"attachment; filename=report_{datetime.now().strftime('%Y%m%d')}.pdf"
                    }
                )
            except ImportError:
                log_error("reportlab not installed", "api")
                raise HTTPException(status_code=501, detail="PDF export not available. Install reportlab.")
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported format: {export_format}")
        
    except HTTPException:
        raise
    except Exception as e:
        log_error(f"Export error: {e}", "api")
        raise HTTPException(status_code=500, detail=str(e))

