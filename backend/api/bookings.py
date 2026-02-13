"""
API Endpoints для работы с записями
"""
from fastapi import APIRouter, Request, Cookie, File, UploadFile, BackgroundTasks, Query
from fastapi.responses import JSONResponse
from typing import Optional, List
from pydantic import BaseModel
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import hashlib

from db import (
    save_booking,
    update_booking_status,
    get_or_create_client, update_client_info, log_activity,
    get_filtered_bookings,
    get_booking_stats
)
from db.connection import get_db_connection
from utils.utils import require_auth
from utils.logger import log_error, log_warning, log_info
from utils.cache import cache
from utils.language_utils import get_localized_name, validate_language
from services.smart_assistant import SmartAssistant
from notifications.master_notifications import notify_master_about_booking, get_master_info, save_notification_log
from utils.datetime_utils import get_current_time, get_salon_timezone

router = APIRouter(tags=["Bookings"])

class CreateBookingRequest(BaseModel):
    instagram_id: Optional[str] = None
    service: str
    date: str
    time: str
    phone: Optional[str] = ''
    name: Optional[str] = None
    master: Optional[str] = ''
    revenue: Optional[float] = 0
    source: Optional[str] = 'manual'
    promo_code: Optional[str] = None
    duration_minutes: Optional[int] = None

class UpdateStatusRequest(BaseModel):
    status: str


def _sanitize_phone_digits(phone: Optional[str]) -> str:
    return ''.join(ch for ch in str(phone or '') if ch.isdigit())


def _build_manual_client_id(name: Optional[str], phone: Optional[str]) -> str:
    phone_digits = _sanitize_phone_digits(phone)
    if phone_digits:
        return f"manual_phone_{phone_digits[-15:]}"

    normalized_name = " ".join(str(name or "").strip().lower().split())
    if normalized_name:
        digest = hashlib.sha1(normalized_name.encode("utf-8")).hexdigest()[:12]
        return f"manual_name_{digest}"

    return f"manual_{int(get_current_time().timestamp())}"


def _safe_duration(duration_minutes: Optional[int], fallback: int = 60) -> int:
    try:
        value = int(duration_minutes) if duration_minutes is not None else int(fallback)
        return max(1, value)
    except Exception:
        return max(1, int(fallback))


def _collect_nearest_available_slots(
    master_identifier: Optional[str],
    date_str: Optional[str],
    time_str: Optional[str],
    duration_minutes: Optional[int],
    limit: int = 5,
    days_to_scan: int = 14,
) -> List[dict]:
    master_value = str(master_identifier or "").strip()
    if not master_value:
        return []

    if not date_str:
        return []

    try:
        from services.master_schedule import MasterScheduleService

        timezone = ZoneInfo(get_salon_timezone())
        schedule_service = MasterScheduleService()
        duration = _safe_duration(duration_minutes, fallback=60)

        requested_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        raw_time = str(time_str or "").strip()[:5]
        try:
            requested_time = datetime.strptime(raw_time, "%H:%M").time()
        except Exception:
            requested_time = datetime.strptime("00:00", "%H:%M").time()

        requested_dt = datetime.combine(requested_date, requested_time).replace(tzinfo=timezone)
        now_dt = get_current_time().astimezone(timezone)
        threshold_dt = requested_dt if requested_dt > now_dt else now_dt

        nearest_slots: List[dict] = []
        for day_offset in range(0, max(1, days_to_scan)):
            current_date = requested_date + timedelta(days=day_offset)
            current_date_str = current_date.strftime("%Y-%m-%d")
            slots = schedule_service.get_available_slots(
                master_name=master_value,
                date=current_date_str,
                duration_minutes=duration,
                return_metadata=False,
            )
            for slot in slots:
                slot_time = str(slot).strip()[:5]
                try:
                    slot_dt = datetime.strptime(f"{current_date_str} {slot_time}", "%Y-%m-%d %H:%M").replace(tzinfo=timezone)
                except Exception:
                    continue

                if slot_dt < threshold_dt:
                    continue

                nearest_slots.append({
                    "date": current_date_str,
                    "time": slot_time,
                    "datetime": slot_dt.isoformat(),
                })
                if len(nearest_slots) >= limit:
                    return nearest_slots

        return nearest_slots
    except Exception as e:
        log_warning(f"Unable to collect nearest slots: {e}", "api")
        return []

def get_client_messengers_for_bookings(client_id: str):
    """Получить список мессенджеров клиента для bookings (DEPRECATED - use get_all_client_messengers)"""
    conn = get_db_connection()
    c = conn.cursor()

    messengers = []

    # Проверяем Instagram
    c.execute("SELECT COUNT(*) FROM chat_history WHERE instagram_id = %s", (client_id,))
    if c.fetchone()[0] > 0:
        messengers.append('instagram')

    # Проверяем другие мессенджеры
    c.execute("""
        SELECT DISTINCT messenger_type
        FROM messenger_messages
        WHERE client_id = %s
    """, (client_id,))

    for row in c.fetchall():
        if row[0] not in messengers:
            messengers.append(row[0])

    conn.close()
    return messengers

def get_all_client_messengers(client_ids: list):
    """Получить мессенджеры для всех клиентов за один запрос (оптимизировано)"""
    if not client_ids:
        return {}

    conn = get_db_connection()
    c = conn.cursor()

    # Результат: {client_id: [messenger1, messenger2, ...]}
    client_messengers = {client_id: [] for client_id in client_ids}

    # Получаем Instagram мессенджеры одним запросом
    c.execute("""
        SELECT DISTINCT instagram_id
        FROM chat_history
        WHERE instagram_id = ANY(%s)
    """, (client_ids,))

    for row in c.fetchall():
        client_id = row[0]
        if client_id in client_messengers and 'instagram' not in client_messengers[client_id]:
            client_messengers[client_id].append('instagram')

    # Получаем другие мессенджеры одним запросом
    c.execute("""
        SELECT client_id, messenger_type
        FROM messenger_messages
        WHERE client_id = ANY(%s)
        GROUP BY client_id, messenger_type
    """, (client_ids,))

    for row in c.fetchall():
        client_id, messenger_type = row
        if client_id in client_messengers and messenger_type not in client_messengers[client_id]:
            client_messengers[client_id].append(messenger_type)

    conn.close()
    return client_messengers

def _resolve_discount_source(source: Optional[str], promo_code: Optional[str]) -> Optional[str]:
    if str(promo_code or "").strip():
        return "promo_code"

    normalized_source = str(source or "").strip().lower()
    if not normalized_source:
        return None

    if "birthday" in normalized_source or "день_рождения" in normalized_source:
        return "birthday"
    if "referral" in normalized_source or "реферал" in normalized_source:
        return "referral"
    if "bonus" in normalized_source or "loyalty" in normalized_source or "бонус" in normalized_source:
        return "bonus"
    return None

def _build_master_display_maps(c, bookings_rows: list, language: str):
    normalized_language = validate_language(language)
    users_by_id = {}
    alias_to_user = {}

    master_ids = sorted({
        int(row[10])
        for row in bookings_rows
        if len(row) > 10 and row[10] is not None
    })

    if master_ids:
        c.execute(
            """
            SELECT id, full_name, username, nickname
            FROM users
            WHERE id = ANY(%s)
            """,
            (master_ids,)
        )
        for user_row in c.fetchall():
            user_id = int(user_row[0])
            full_name = str(user_row[1] or "").strip()
            username = str(user_row[2] or "").strip()
            nickname = str(user_row[3] or "").strip()
            users_by_id[user_id] = (full_name, username, nickname)
            for alias in (full_name, username, nickname):
                normalized_alias = alias.lower().strip()
                if normalized_alias:
                    alias_to_user[normalized_alias] = (user_id, full_name)

    unresolved_aliases = sorted({
        str(row[9] or "").strip().lower()
        for row in bookings_rows
        if len(row) > 9 and row[9] and (
            len(row) <= 10
            or row[10] is None
        )
    })
    unresolved_aliases = [alias for alias in unresolved_aliases if alias and alias not in alias_to_user]

    if unresolved_aliases:
        c.execute(
            """
            SELECT id, full_name, username, nickname
            FROM users
            WHERE is_service_provider = TRUE
              AND deleted_at IS NULL
              AND (
                LOWER(full_name) = ANY(%s)
                OR LOWER(username) = ANY(%s)
                OR LOWER(COALESCE(nickname, '')) = ANY(%s)
              )
            """,
            (unresolved_aliases, unresolved_aliases, unresolved_aliases)
        )
        for user_row in c.fetchall():
            user_id = int(user_row[0])
            full_name = str(user_row[1] or "").strip()
            username = str(user_row[2] or "").strip()
            nickname = str(user_row[3] or "").strip()
            users_by_id[user_id] = (full_name, username, nickname)
            for alias in (full_name, username, nickname):
                normalized_alias = alias.lower().strip()
                if normalized_alias:
                    alias_to_user[normalized_alias] = (user_id, full_name)

    localized_by_id = {}
    for user_id, values in users_by_id.items():
        full_name = values[0] or values[1]
        localized_by_id[user_id] = get_localized_name(user_id, full_name, normalized_language)

    return localized_by_id, alias_to_user, normalized_language

def _get_localized_master_name(
    raw_master: Optional[str],
    master_user_id: Optional[int],
    localized_by_id: dict,
    alias_to_user: dict,
    language: str
) -> Optional[str]:
    raw_value = str(raw_master or "").strip()
    if not raw_value:
        return None

    if master_user_id is not None:
        localized = localized_by_id.get(int(master_user_id))
        if localized:
            return localized

    alias_key = raw_value.lower()
    alias_match = alias_to_user.get(alias_key)
    if alias_match:
        localized = localized_by_id.get(int(alias_match[0]))
        if localized:
            return localized

    return get_localized_name(0, raw_value, language)

@router.get("/client/bookings")
def get_client_bookings(session_token: Optional[str] = Cookie(None)):
    """Получить записи текущего клиента (API)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    user_id = user.get("id")
    instagram_id = user.get("username")
    phone = user.get("phone")
    full_name = user.get("full_name")

    conn = get_db_connection()
    c = conn.cursor()

    try:
        # Primary: Search by user_id (most reliable)
        c.execute("""
            SELECT id, instagram_id, service_name, datetime, phone,
                   name, status, created_at, revenue, master
            FROM bookings
            WHERE user_id = %s
            ORDER BY datetime DESC
        """, (user_id,))
        bookings = c.fetchall()

        # Fallback: If no bookings found by user_id, try other methods
        if len(bookings) == 0:
            # Try by username/phone/name
            c.execute("""
                SELECT id, instagram_id, service_name, datetime, phone,
                       name, status, created_at, revenue, master
                FROM bookings
                WHERE instagram_id = %s
                   OR (phone IS NOT NULL AND phone = %s)
                   OR (name IS NOT NULL AND LOWER(name) = LOWER(%s))
                ORDER BY datetime DESC
            """, (instagram_id, phone or '', full_name or ''))
            bookings = c.fetchall()

    except Exception as e:
        from utils.logger import log_error
        log_error(f"Error fetching client bookings: {e}", "bookings")
        bookings = []
    finally:
        conn.close()
    
    # Format dates
    formatted_bookings = []
    for b in bookings:
        formatted_bookings.append({
            "id": b[0],
            "instagram_id": b[1],
            "service_name": b[2],
            "start_time": b[3], # Keep ISO format for JS
            "phone": b[4],
            "name": b[5],
            "status": b[6],
            "created_at": b[7],
            "revenue": b[8],
            "master_name": b[9] if len(b) > 9 else None
        })
        
    return {
        "bookings": formatted_bookings,
        "count": len(formatted_bookings)
    }

@router.get("/bookings")
def list_bookings(
    session_token: Optional[str] = Cookie(None),
    page: int = 1,
    limit: int = 50,
    search: Optional[str] = None,
    status: Optional[str] = None,
    master: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: Optional[str] = 'datetime',
    order: Optional[str] = 'desc',
    language: str = Query('ru')
):
    """Получить записи с пагинацией и фильтрацией"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    # RBAC: Clients cannot see all bookings
    # Also restrict roles that strictly don't have access (like marketer)
    allowed_roles = ["director", "admin", "manager", "sales", "employee"]
    if user["role"] not in allowed_roles:
        return JSONResponse({"error": "Forbidden"}, status_code=403)

    offset = (page - 1) * limit
    
    # RBAC: Employees see only their bookings (via user_id or master filter)
    filter_user_id = None
    if user["role"] == "employee":
        # We can pass user_id to filter by that
        filter_user_id = user["id"]
        # Or if master name filter is preferred, keep as is.
        # Ideally we use user_id to filter where user_id=... OR master=...
        # get_filtered_bookings handles user_id check.

    # Fetch data
    import time
    start_time = time.time()
    
    bookings, total_results = get_filtered_bookings(
        limit=limit,
        offset=offset,
        search=search,
        status=status,
        master=master,
        date_from=date_from,
        date_to=date_to,
        user_id=filter_user_id,
        sort_by=sort,
        order=order
    )
    
    db_duration = time.time() - start_time
    log_info(f"⏱️ get_filtered_bookings took {db_duration:.4f}s for {limit} items", "perf")

    # Fetch Stats (Global or Filtered)
    t1 = time.time()
    stats = get_booking_stats(
        search=search,
        master=master,
        date_from=date_from,
        date_to=date_to,
        user_id=filter_user_id
    )
    
    # Scrub financial stats for Sales role (they see individual bookings but not aggregate revenue)
    if user["role"] == "sales":
        if 'total_revenue' in stats: stats['total_revenue'] = 0
        if 'avg_check' in stats: stats['avg_check'] = 0
        
    log_info(f"⏱️ get_booking_stats took {time.time() - t1:.4f}s", "perf")

    # Добавляем информацию о мессенджерах для каждой записи
    bookings_with_messengers = []
    language_code = validate_language(language)

    # Оптимизация: получаем всех мессенджеров одним запросом
    t2 = time.time()
    client_ids = list(set([b[1] for b in bookings if b[1]]))  # Уникальные client_id
    all_messengers = get_all_client_messengers(client_ids) if client_ids else {}
    log_info(f"⏱️ get_all_client_messengers took {time.time() - t2:.4f}s for {len(client_ids)} clients", "perf")

    # Оптимизация: получаем все телефоны клиентов одним запросом
    t3 = time.time()
    client_phones = {}
    if client_ids:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT instagram_id, phone FROM clients WHERE instagram_id = ANY(%s)", (client_ids,))
        for row in c.fetchall():
            if row[0] and row[1]:
                client_phones[row[0]] = row[1]

        localized_by_id, alias_to_user, normalized_language = _build_master_display_maps(c, bookings, language_code)
        conn.close()
    else:
        localized_by_id = {}
        alias_to_user = {}
        normalized_language = language_code
    log_info(f"⏱️ fetching_client_phones took {time.time() - t3:.4f}s", "perf")

    t4 = time.time()
    for b in bookings:
        client_id = b[1]
        messengers = all_messengers.get(client_id, [])

        # Get phone from booking row
        phone = b[4] if len(b) > 4 else ''

        # Fallback to clients table if phone is missing in booking
        if not phone and client_id:
            phone = client_phones.get(client_id, '')

        master_user_id = b[10] if len(b) > 10 else None
        localized_master = _get_localized_master_name(
            raw_master=b[9] if len(b) > 9 else None,
            master_user_id=master_user_id,
            localized_by_id=localized_by_id,
            alias_to_user=alias_to_user,
            language=normalized_language
        )
        promo_code = b[13] if len(b) > 13 else None
        source_value = b[12] if len(b) > 12 else 'manual'
        discount_source = _resolve_discount_source(source_value, promo_code)

        bookings_with_messengers.append({
            "id": b[0],
            "client_id": client_id,
            "service": b[2] if len(b) > 2 else None,
            "service_name": b[2] if len(b) > 2 else None,  # ✅ Дублируем для совместимости
            "datetime": b[3] if len(b) > 3 else None,
            "phone": phone,
            "name": b[5] if len(b) > 5 else '',
            "status": b[6] if len(b) > 6 else 'pending',
            "created_at": b[7] if len(b) > 7 else None,
            "revenue": b[8] if len(b) > 8 else 0,
            "master": localized_master,
            "master_user_id": master_user_id,
            "user_id": b[11] if len(b) > 11 else None,
            "source": source_value,
            "promo_code": promo_code,
            "discount_source": discount_source,
            "messengers": messengers
        })
    log_info(f"⏱️ formatting_response took {time.time() - t4:.4f}s", "perf")
    log_info(f"⏱️ TOTAL list_bookings took {time.time() - start_time:.4f}s", "perf")

    return {
        "bookings": bookings_with_messengers,
        "count": len(bookings_with_messengers), # Legacy count of items in this page
        "total": total_results,                   # Total items in DB
        "page": page,
        "limit": limit,
        "stats": stats
    }

@router.get("/bookings/{booking_id}")
def get_booking_detail(
    booking_id: int,
    session_token: Optional[str] = Cookie(None),
    language: str = Query('ru')
):
    """Получить одну запись по ID"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    conn = get_db_connection()
    c = conn.cursor()

    c.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_name = 'bookings'
          AND column_name = 'master_user_id'
        LIMIT 1
        """
    )
    has_master_user_id = bool(c.fetchone())

    c.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_name = 'bookings'
          AND column_name = 'promo_code'
        LIMIT 1
        """
    )
    has_promo_code = bool(c.fetchone())

    c.execute(
        """
        SELECT 1
        FROM information_schema.columns
        WHERE table_name = 'bookings'
          AND column_name = 'source'
        LIMIT 1
        """
    )
    has_source = bool(c.fetchone())

    promo_code_select = "promo_code" if has_promo_code else "NULL AS promo_code"
    source_select = "source" if has_source else "'manual' AS source"
    master_user_id_select = "master_user_id" if has_master_user_id else "NULL AS master_user_id"

    c.execute(f"""
        SELECT id, instagram_id, service_name, datetime, phone,
               name, status, created_at, revenue, master, user_id, notes,
               {source_select}, {promo_code_select}, {master_user_id_select}
        FROM bookings
        WHERE id = %s
    """, (booking_id,))
    booking = c.fetchone()
    
    if not booking:
        conn.close()
        return JSONResponse({"error": "Booking not found"}, status_code=404)

    language_code = validate_language(language)
    localized_by_id, alias_to_user, normalized_language = _build_master_display_maps(c, [booking], language_code)
    raw_master_name = booking[9] if len(booking) > 9 else None
    master_user_id = booking[14] if len(booking) > 14 else None
    localized_master = _get_localized_master_name(
        raw_master=raw_master_name,
        master_user_id=master_user_id,
        localized_by_id=localized_by_id,
        alias_to_user=alias_to_user,
        language=normalized_language
    )
    source_value = booking[12] if len(booking) > 12 else "manual"
    promo_code = booking[13] if len(booking) > 13 else None
    discount_source = _resolve_discount_source(source_value, promo_code)
    conn.close()
        
    # RBAC: Clients can only see their own bookings
    if user["role"] == "client":
        booking_client_id = booking[1]
        booking_phone = booking[4]
        
        user_client_id = user.get("username")
        user_phone = user.get("phone")
        
        # Check fuzzy match
        is_owner = (user_client_id and str(user_client_id) == str(booking_client_id))
        if not is_owner and user_phone and booking_phone:
             # Basic phone normalization for comparison if needed, currently exact match
             is_owner = (str(user_phone) == str(booking_phone))
             
        if not is_owner:
            return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    return {
        "id": booking[0],
        "client_id": booking[1],
        "service": booking[2],
        "datetime": booking[3],
        "phone": booking[4],
        "name": booking[5],
        "status": booking[6],
        "created_at": booking[7],
        "revenue": booking[8] if booking[8] is not None else 0,
        "master": localized_master,
        "master_user_id": master_user_id,
        "user_id": booking[10],
        "notes": booking[11],
        "source": source_value,
        "promo_code": promo_code,
        "discount_source": discount_source
    }

async def process_booking_background_tasks(
    booking_id: int,
    data: dict,
    user_id: str
):
    """Background task handler for new bookings with timeouts"""
    import asyncio
    
    try:
        instagram_id = data.get('instagram_id')
        service = data.get('service')
        booking_date_value = str(data.get('date') or '').strip()
        booking_time_value = str(data.get('time') or '').strip()
        datetime_str = f"{booking_date_value} {booking_time_value}".strip()
        phone = data.get('phone', '')
        name = data.get('name')
        master = data.get('master', '')

        display_date = booking_date_value
        display_time = booking_time_value
        try:
            if booking_date_value:
                parsed_date = datetime.strptime(booking_date_value, "%Y-%m-%d")
                display_date = parsed_date.strftime("%d.%m.%Y")
            if booking_time_value:
                parsed_time = datetime.strptime(booking_time_value[:5], "%H:%M")
                display_time = parsed_time.strftime("%H:%M")
        except Exception:
            # Keep raw values if parsing fails
            pass

        # 1. 🧠 Smart Assistant Learning (with timeout)
        try:
            async def learn_task():
                assistant = SmartAssistant(instagram_id)
                assistant.learn_from_booking({
                    'service': service,
                    'master': master,
                    'datetime': datetime_str,
                    'phone': phone,
                    'name': name
                })
                log_info(f"🧠 SmartAssistant learned from booking for {instagram_id}", "bookings")
            
            await asyncio.wait_for(learn_task(), timeout=5.0)
        except asyncio.TimeoutError:
            log_warning(f"⏱️ SmartAssistant learning timed out for {instagram_id}", "bookings")
        except Exception as e:
            log_error(f"❌ SmartAssistant learning failed: {e}", "bookings")

        # 2. Notify Master (with timeout)
        if master and booking_id:
            try:
                async def notify_master_task():
                    notification_results = await notify_master_about_booking(
                        master_name=master,
                        client_name=name,
                        service=service,
                        datetime_str=datetime_str,
                        phone=phone,
                        booking_id=booking_id
                    )

                    # Log notifications
                    master_info = get_master_info(master)
                    if master_info:
                        for notif_type, success in notification_results.items():
                            if success:
                                save_notification_log(master_info["id"], booking_id, notif_type, "sent")
                            elif master_info.get(notif_type) or master_info.get(f"{notif_type}_username"):
                                save_notification_log(master_info["id"], booking_id, notif_type, "failed")
                
                await asyncio.wait_for(notify_master_task(), timeout=5.0)
            except asyncio.TimeoutError:
                log_warning(f"⏱️ Master notification timed out for booking {booking_id}", "bookings")
            except Exception as e:
                log_error(f"❌ Error sending master notification: {e}", "api")

        # 3. Notify Admin (with timeout)
        try:
            await asyncio.wait_for(notify_admin_about_booking(data), timeout=5.0)
        except asyncio.TimeoutError:
            log_warning(f"⏱️ Admin notification timed out for booking {booking_id}", "bookings")
        except Exception as e:
            log_error(f"❌ Error sending admin notification: {e}", "api")

        # 4. Notify Client (Confirmation)
        try:
            from services.universal_messenger import send_universal_message
            send_result = await send_universal_message(
                recipient_id=instagram_id,
                template_name="booking_confirmation",
                context={
                    "name": name or "Клиент",
                    "service": service,
                    "date": display_date,
                    "time": display_time,
                    "datetime": datetime_str,
                    "master": master or ""
                },
                booking_id=booking_id,
                platform='auto'
            )
            if send_result.get("success"):
                log_info(f"📩 Client confirmation sent for booking {booking_id}", "bookings")
            else:
                log_warning(
                    f"Client confirmation failed for booking {booking_id}: {send_result.get('error')}",
                    "bookings"
                )
        except Exception as e:
            log_error(f"❌ Error sending client confirmation: {e}", "bookings")
        
    except Exception as e:
        log_error(f"❌ Background task error: {e}", "background_tasks")


def process_booking_post_create_updates(
    instagram_id: str,
    phone: str,
    name: Optional[str],
    user_id: int,
    service: str,
):
    """Выполнить пост-обновления после создания записи вне критического пути ответа API."""
    try:
        if phone or name:
            update_client_info(instagram_id, phone=phone, name=name)
    except Exception as e:
        log_error(f"Error updating client profile after booking: {e}", "api")

    try:
        from db.users import update_user_info as db_update_user
        user_updates = {}
        if phone:
            user_updates['phone'] = phone
        if name:
            user_updates['full_name'] = name
        if user_updates and user_id:
            db_update_user(user_id, user_updates)
    except Exception as e:
        log_error(f"Error updating user profile after booking: {e}", "api")

    try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute(
            """
            SELECT id
            FROM workflow_stages
            WHERE entity_type = 'pipeline'
              AND (LOWER(name) LIKE '%book%' OR LOWER(name) LIKE '%запис%')
            LIMIT 1
            """
        )
        stage_row = c.fetchone()
        if stage_row:
            c.execute(
                "UPDATE clients SET pipeline_stage_id = %s WHERE instagram_id = %s",
                (stage_row[0], instagram_id)
            )
            conn.commit()
    except Exception as e:
        log_error(f"Error syncing client stage: {e}", "api")
    finally:
        if 'conn' in locals():
            conn.close()

    try:
        log_activity(user_id, "create_booking", "booking", instagram_id, f"Service: {service}")
    except Exception as e:
        log_error(f"Error writing booking activity log: {e}", "api")

    try:
        from db.clients import update_client_temperature
        update_client_temperature(instagram_id)
    except Exception as e:
        log_error(f"Error refreshing client temperature after booking: {e}", "api")

@router.post("/bookings")
def create_booking_api(
    data: CreateBookingRequest,
    background_tasks: BackgroundTasks,
    session_token: Optional[str] = Cookie(None)
):
    """Создать запись"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)

    try:
        raw_instagram_id = str(data.instagram_id or '').strip()
        service = data.service
        datetime_str = f"{data.date} {data.time}"
        phone = data.phone or ''
        name = data.name
        master = data.master or ''
        revenue = data.revenue or 0
        source = data.source or 'manual'
        promo_code = data.promo_code
        user_id = user["id"]

        if not raw_instagram_id:
            if not str(name or '').strip() and not _sanitize_phone_digits(phone):
                return JSONResponse({"error": "missing_client_identity"}, status_code=400)
            instagram_id = _build_manual_client_id(name=name, phone=phone)
        else:
            instagram_id = raw_instagram_id

        # Critical path: only synchronous booking save + strict slot validation.
        booking_id = save_booking(
            instagram_id,
            service,
            datetime_str,
            phone,
            name,
            master=master,
            user_id=user_id,
            revenue=revenue,
            source=source,
            promo_code=promo_code,
            duration_minutes=data.duration_minutes,
        )

        # Offload slow tasks
        background_tasks.add_task(
            process_booking_post_create_updates,
            instagram_id,
            phone,
            name,
            user_id,
            service,
        )
        if booking_id:
            payload = data.dict()
            payload["instagram_id"] = instagram_id
            background_tasks.add_task(process_booking_background_tasks, booking_id, payload, user_id)

        # Invalidate cache
        cache.clear_by_pattern("dashboard_*")
        cache.clear_by_pattern("funnel_*")

        return {"success": True, "message": "Booking created", "booking_id": booking_id}
    except ValueError as e:
        message = str(e)
        if message.startswith("slot_unavailable:"):
            reason = message.split(":", 1)[1] or "unavailable"
            nearest_slots = _collect_nearest_available_slots(
                master_identifier=data.master,
                date_str=data.date,
                time_str=data.time,
                duration_minutes=data.duration_minutes,
            )
            return JSONResponse(
                {
                    "error": "slot_unavailable",
                    "reason": reason,
                    "nearest_slots": nearest_slots,
                },
                status_code=409
            )
        log_error(f"Booking validation error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)
    except Exception as e:
        log_error(f"Booking creation error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

async def notify_admin_booking_status_change(booking_id: int, old_status: str, new_status: str):
    """Уведомить админа об изменении статуса записи"""
    from utils.email import send_email_sync
    from integrations.telegram_bot import send_telegram_alert
    import os
    import asyncio

    try:
        # Get booking details
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("""
            SELECT b.instagram_id, b.service_name, b.datetime, b.master, c.name, c.phone
            FROM bookings b
            LEFT JOIN clients c ON b.instagram_id = c.instagram_id
            WHERE b.id = %s
        """, (booking_id,))
        row = c.fetchone()
        conn.close()

        if not row:
            return

        instagram_id, service, datetime_str, master, client_name, phone = row
        client_display = client_name or instagram_id

        # Status translations
        status_emoji = {
            'pending': '⏳',
            'confirmed': '✅',
            'cancelled': '❌',
            'completed': '✔️',
            'no_show': '👻'
        }

        status_text = {
            'pending': 'status_pending',
            'confirmed': 'status_confirmed',
            'cancelled': 'status_cancelled',
            'completed': 'status_completed',
            'no_show': 'status_no_show'
        }

        admin_email = os.getenv('FROM_EMAIL') or os.getenv('SMTP_USERNAME')

        subject = f"{status_emoji.get(new_status, '📝')} Изменение статуса записи: {client_display}"
        message = (
            f"Клиент: {client_display}\n"
            f"Телефон: {phone or 'Не указан'}\n"
            f"Услуга: {service}\n"
            f"Мастер: {master or 'Любой'}\n"
            f"Дата/Время: {datetime_str}\n"
            f"Старый статус: {status_text.get(old_status, old_status)}\n"
            f"Новый статус: {status_text.get(new_status, new_status)}"
        )

        # Email
        if admin_email:
            try:
                send_email_sync([admin_email], subject, message)
            except Exception as e:
                print(f"Error sending email: {e}")

        # Telegram
        try:
            tg_msg = (
                f"{status_emoji.get(new_status, '📝')} <b>Booking Status Updated</b>\n\n"
                f"👤 <b>Client:</b> {client_display}\n"
                f"📞 <b>Phone:</b> <code>{phone or 'Not specified'}</code>\n"
                f"💅 <b>Service:</b> {service}\n"
                f"👨‍💼 <b>Professional:</b> {master or 'Any'}\n"
                f"🕒 <b>Time:</b> {datetime_str}\n"
                f"📊 <b>Status:</b> {status_text.get(old_status, old_status)} → {status_text.get(new_status, new_status)}"
            )
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(send_telegram_alert(tg_msg))
            loop.close()
        except Exception as e:
            print(f"Error sending telegram: {e}")
    except Exception as e:
        print(f"Error notifying admin about status change: {e}")


def notify_admin_booking_status_change_background(booking_id: int, old_status: str, new_status: str):
    """Background runner to avoid blocking the status update response."""
    import asyncio

    try:
        asyncio.run(notify_admin_booking_status_change(booking_id, old_status, new_status))
    except Exception as e:
        log_error(f"Error in background status notification for booking {booking_id}: {e}", "notifications")

async def notify_admin_about_booking(data: dict):
    """Notify admin about new booking with timeouts"""
    from utils.email import send_email_sync
    from integrations.telegram_bot import send_telegram_alert
    import os
    import asyncio
    
    name = data.get('name')
    phone = data.get('phone')
    service = data.get('service')
    datetime_str = f"{data.get('date')} {data.get('time')}"
    master = data.get('master', 'any_master')
    source = data.get('source', 'manual')
    
    # Beautiful source names
    source_display = "🔧 Manual Entry"
    source_emoji = "🔧"
    if source == 'public_landing':
        source_display = "🏠 Public Landing"
        source_emoji = "🏠"
    elif source == 'client_cabinet':
        source_display = "👤 Client Cabinet"
        source_emoji = "👤"
    elif source == 'instagram':
        source_display = "📸 Instagram"
        source_emoji = "📸"
    elif source == 'whatsapp':
        source_display = "💬 WhatsApp"
        source_emoji = "💬"
    elif source == 'website':
        source_display = "🌐 Website"
        source_emoji = "🌐"
    
    # 1. Email Admin (with timeout)
    admin_email = os.getenv('FROM_EMAIL') or os.getenv('SMTP_USERNAME')
    if admin_email:
        subject = f"📅 Новая запись: {name}"
        message_text = (
            f"Имя: {name}\n"
            f"Телефон: {phone}\n"
            f"Услуга: {service}\n"
            f"Мастер: {master}\n"
            f"Время: {datetime_str}\n"
            f"Источник: {source_display}"
        )
        try:
            loop = asyncio.get_event_loop()
            await asyncio.wait_for(
                loop.run_in_executor(None, lambda: send_email_sync([admin_email], subject, message_text)),
                timeout=3.0
            )
        except asyncio.TimeoutError:
            log_warning(f"⏱️ Admin email timed out for booking: {name}", "notifications")
        except Exception as e:
            log_error(f"❌ Error sending admin email: {e}", "notifications")

    # 2. Telegram Admin (with timeout)
    try:
        telegram_message = (
            f"📅 <b>New Booking!</b>\n\n"
            f"👤 <b>Name:</b> {name}\n"
            f"📞 <b>Phone:</b> <code>{phone}</code>\n"
            f"💇‍♀️ <b>Service:</b> {service}\n"
            f"👤 <b>Professional:</b> {master}\n"
            f"🕒 <b>Time:</b> {datetime_str}\n"
            f"━━━━━━━━━━━━━━━━━━\n"
            f"{source_emoji} <b>Source:</b> {source_display}"
        )
        await asyncio.wait_for(send_telegram_alert(telegram_message), timeout=3.0)
    except asyncio.TimeoutError:
        log_warning(f"⏱️ Admin telegram timed out for booking: {name}", "notifications")
    except Exception as e:
        log_error(f"❌ Error sending admin telegram: {e}", "notifications")

@router.post("/bookings/{booking_id}/status")
def update_booking_status_api(
    booking_id: int,
    data: UpdateStatusRequest,
    background_tasks: BackgroundTasks,
    session_token: Optional[str] = Cookie(None)
):
    """Изменить статус записи"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    status = data.status

    # RBAC: Clients can only cancel their own bookings
    if user["role"] == "client":
        if status != "cancelled":
            return JSONResponse({"error": "Forbidden: Clients can only cancel bookings"}, status_code=403)
            
        # Verify ownership
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT instagram_id, phone FROM bookings WHERE id = %s", (booking_id,))
        row = c.fetchone()
        conn.close()
        
        if not row:
             return JSONResponse({"error": "Booking not found"}, status_code=404)
        
        is_owner = (user.get("username") and str(user.get("username")) == str(row[0]))
        if not is_owner and user.get("phone") and row[1]:
             is_owner = (str(user.get("phone")) == str(row[1]))
             
        if not is_owner:
            return JSONResponse({"error": "Forbidden"}, status_code=403)

    if not status:
        return JSONResponse({"error": "Status required"}, status_code=400)

    # Get old status
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT status FROM bookings WHERE id = %s", (booking_id,))
    old_status_row = c.fetchone()
    conn.close()

    old_status = old_status_row[0] if old_status_row else None

    success = update_booking_status(booking_id, status)
    if success:
        log_activity(user["id"], "update_booking_status", "booking",
                    str(booking_id), f"Status: {status}")

        if old_status and old_status != status:
            background_tasks.add_task(
                notify_admin_booking_status_change_background,
                booking_id,
                old_status,
                status
            )
        
        # loyalty
        if status == 'completed':
            try:
                conn = get_db_connection()
                c = conn.cursor()
                c.execute("""
                    SELECT b.instagram_id, b.revenue, b.service_name, s.category 
                    FROM bookings b
                    LEFT JOIN services s ON b.service_name = s.name
                    WHERE b.id = %s
                """, (booking_id,))
                b_row = c.fetchone()
                conn.close()

                if b_row:
                    client_id, revenue, service_name, category = b_row
                    revenue = revenue or 0
                    from services.loyalty import LoyaltyService
                    loyalty = LoyaltyService()
                    if not loyalty.has_earned_for_booking(booking_id):
                        points = loyalty.points_for_booking(revenue, service_category=category)
                        if points > 0:
                            loyalty.earn_points(client_id=client_id, points=points, reason=f"Посещение: {service_name}", booking_id=booking_id)
            except Exception as e:
                log_error(f"Error earning loyalty points: {e}", "api")

        cache.clear_by_pattern("dashboard_*")
        cache.clear_by_pattern("funnel_*")

        return {"success": True, "message": "Booking status updated"}
    
    return JSONResponse({"error": "Update failed"}, status_code=400)

@router.post("/bookings/import")
async def import_bookings(
    file: UploadFile = File(...),
    session_token: Optional[str] = Cookie(None)
):
    """Импортировать записи из CSV или Excel файла"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    try:
        if not file.filename:
            return JSONResponse({"error": "No file provided"}, status_code=400)
        
        file_ext = file.filename.split('.')[-1].lower()
        
        if file_ext not in ['csv', 'xlsx', 'xls']:
            return JSONResponse({"error": "Invalid file format. Use CSV or Excel"}, 
                              status_code=400)
        
        content = await file.read()
        
        imported_count = 0
        skipped_count = 0
        errors = []
        
        if file_ext == 'csv':
            import csv
            import io
            import time
            
            csv_content = content.decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            
            for row in csv_reader:
                try:
                    instagram_id = row.get('instagram_id') or row.get('ID') or \
                                  f"import_{int(time.time())}_{imported_count}"
                    name = row.get('name') or row.get('Имя') or 'Импортированный клиент'
                    phone = row.get('phone') or row.get('Телефон') or ''
                    service = row.get('service') or row.get('Услуга') or 'Неизвестно'
                    datetime_str = row.get('datetime') or row.get('Дата/Время') or \
                                  datetime.now().isoformat()
                    status = row.get('status') or row.get('Статус') or 'pending'
                    revenue = float(row.get('revenue') or row.get('Доход') or 0)
                    
                    get_or_create_client(instagram_id, username=name)
                    
                    if phone or name:
                        update_client_info(instagram_id, name=name, phone=phone)
                    
                    conn = get_db_connection()
                    c = conn.cursor()
                    
                    c.execute("""INSERT INTO bookings 
                                 (instagram_id, service_name, datetime, phone, name, 
                                  status, created_at, revenue)
                                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                              (instagram_id, service, datetime_str, phone, name, status, 
                               datetime.now().isoformat(), revenue))
                    
                    conn.commit()
                    conn.close()
                    
                    imported_count += 1
                    
                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Строка {imported_count + skipped_count}: {str(e)}")
        
        elif file_ext in ['xlsx', 'xls']:
            try:
                from openpyxl import load_workbook
            except ImportError:
                return JSONResponse({"error": "Excel support not available"}, 
                                  status_code=500)
            
            import io
            import time
            
            wb = load_workbook(io.BytesIO(content))
            ws = wb.active
            
            if ws is None:
                return JSONResponse({"error": "No active worksheet found"}, status_code=400)
            
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    row_dict = dict(zip(headers, row))
                    
                    # Извлекаем данные из row_dict с fallback значениями
                    instagram_id = str(row_dict.get('instagram_id') or row_dict.get('ID') or f"import_{int(time.time())}_{imported_count}")
                    name = str(row_dict.get('name') or row_dict.get('Имя') or 'Импортированный клиент')
                    phone = str(row_dict.get('phone') or row_dict.get('Телефон') or '')
                    service = str(row_dict.get('service') or row_dict.get('Услуга') or 'Неизвестно')
                    datetime_str = str(
                        row_dict.get('datetime') or
                        row_dict.get('Дата/Время') or
                        datetime.now().isoformat()
                    )
                    status = str(row_dict.get('status') or row_dict.get('Статус') or 'pending')
                    revenue = row_dict.get('revenue') or row_dict.get('Доход') or 0

                    
                    get_or_create_client(instagram_id, username=name)
                    
                    if phone or name:
                        update_client_info(instagram_id, name=name, phone=phone)
                    
                    conn = get_db_connection()
                    c = conn.cursor()
                    
                    c.execute("""INSERT INTO bookings 
                                 (instagram_id, service_name, datetime, phone, name, 
                                  status, created_at, revenue)
                                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
                              (instagram_id, service, datetime_str, phone, name, status, 
                               datetime.now().isoformat(), revenue))
                    
                    conn.commit()
                    conn.close()
                    
                    imported_count += 1
                    
                except Exception as e:
                    skipped_count += 1
                    errors.append(f"Строка {imported_count + skipped_count + 1}: {str(e)}")
        
        log_activity(user["id"], "import_bookings", "bookings", str(imported_count), 
                     f"Imported {imported_count} bookings")
        
        return {
            "success": True,
            "imported": imported_count,
            "skipped": skipped_count,
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        log_error(f"Import error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/bookings/{booking_id}/waitlist")
async def add_to_booking_waitlist(
    booking_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Добавить клиента в лист ожидания (#17)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        data = await request.json()
        instagram_id = data.get('instagram_id')
        service = data.get('service')
        preferred_date = data.get('date')
        preferred_time = data.get('time')
        
        if not all([instagram_id, service, preferred_date, preferred_time]):
            return JSONResponse({"error": "Missing required fields"}, status_code=400)
        
        from db.bookings import add_to_waitlist
        add_to_waitlist(instagram_id, service, preferred_date, preferred_time)
        
        log_activity(user["id"], "add_to_waitlist", "booking", 
                    str(booking_id), f"Added to waitlist")
        
        return {"success": True, "message": "Added to waitlist"}
        
    except Exception as e:
        log_error(f"Waitlist error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/api/bookings/patterns/{client_id}")
def get_client_booking_pattern(client_id: str, session_token: Optional[str] = Cookie(None)):
    """Получить обычный паттерн записи клиента (#7)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        from db.bookings import get_client_usual_booking_pattern
        pattern = get_client_usual_booking_pattern(client_id)
        
        return {
            "has_pattern": pattern is not None,
            "pattern": pattern
        }
        
    except Exception as e:
        log_error(f"Pattern error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/bookings/incomplete/{client_id}")
async def get_incomplete_client_booking(
    client_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Получить незавершённую запись (#4)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        from db.bookings import get_incomplete_booking
        incomplete = get_incomplete_booking(client_id)
        
        return {
            "has_incomplete": incomplete is not None,
            "booking": incomplete
        }
        
    except Exception as e:
        log_error(f"Incomplete booking error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/bookings/course-progress/{client_id}")
async def get_course_progress(
    client_id: str,
    service: str,
    session_token: Optional[str] = Cookie(None)
):
    """Получить прогресс по курсу (#11)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        from db.bookings import get_client_course_progress
        progress = get_client_course_progress(client_id, service)
        
        return {
            "has_course": progress is not None,
            "progress": progress
        }
        
    except Exception as e:
        log_error(f"Course progress error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/bookings/no-show-risk/{client_id}")
async def get_no_show_risk(
    client_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Получить риск no-show клиента (#19)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        from db.clients import calculate_no_show_risk
        risk = calculate_no_show_risk(client_id)
        
        risk_level = "low"
        if risk > 0.5:
            risk_level = "high"
        elif risk > 0.3:
            risk_level = "medium"

        return {
            "risk_score": risk,
            "risk_level": risk_level,
            "requires_deposit": risk > 0.4
        }

    except Exception as e:
        log_error(f"No-show risk error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.put("/bookings/{booking_id}")
async def update_booking_api(
    booking_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить запись (полное обновление)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
    # RBAC: Clients cannot direct update bookings (must use cancellation)
    if user["role"] == "client":
        return JSONResponse({"error": "Forbidden: Clients cannot edit bookings directly"}, status_code=403)

    data = await request.json()

    try:
        # Получаем старую запись для сравнения
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("SELECT instagram_id, service_name, datetime, master, name, phone, revenue, source FROM bookings WHERE id = %s",
                  (booking_id,))
        old_booking = c.fetchone()

        if not old_booking:
            conn.close()
            return JSONResponse({"error": "Booking not found"}, status_code=404)

        current_instagram_id, old_service, old_datetime, old_master, old_name, old_phone, old_revenue, old_source = old_booking

        # Обновляем запись
        new_service = data.get('service', old_service)
        new_datetime = f"{data.get('date')} {data.get('time')}" if data.get('date') and data.get('time') else old_datetime
        raw_master = data.get('master', old_master)
        from db.bookings import normalize_master_value
        normalized_master = normalize_master_value(raw_master)
        if normalized_master is None and raw_master and str(raw_master).strip().lower() not in {"any", "any_master", "global", "любой", "не указан", "не указано"}:
            normalized_master = str(raw_master).strip()
        new_master = normalized_master
        new_name = data.get('name', old_name)
        new_phone = data.get('phone', old_phone)
        new_revenue = data.get('revenue', old_revenue)
        new_source = data.get('source', old_source)

        has_master_user_id = False
        master_user_id = None
        c.execute(
            """
            SELECT 1
            FROM information_schema.columns
            WHERE table_name = 'bookings' AND column_name = 'master_user_id'
            LIMIT 1
            """
        )
        has_master_user_id = bool(c.fetchone())
        if has_master_user_id and new_master:
            c.execute(
                """
                SELECT id
                FROM users
                WHERE is_service_provider = TRUE
                  AND deleted_at IS NULL
                  AND (
                    LOWER(full_name) = LOWER(%s)
                    OR LOWER(username) = LOWER(%s)
                    OR LOWER(COALESCE(nickname, '')) = LOWER(%s)
                  )
                ORDER BY
                    CASE
                        WHEN LOWER(username) = LOWER(%s) THEN 0
                        WHEN LOWER(full_name) = LOWER(%s) THEN 1
                        ELSE 2
                    END,
                    id ASC
                LIMIT 1
                """,
                (new_master, new_master, new_master, new_master, new_master),
            )
            master_row = c.fetchone()
            if master_row:
                master_user_id = int(master_row[0])

        # Sync client info if phone/name changed
        if new_phone != old_phone or new_name != old_name:
            from db.clients import update_client_info
            update_client_info(current_instagram_id, phone=new_phone, name=new_name)

        if has_master_user_id:
            c.execute(
                """
                UPDATE bookings
                SET service_name = %s, datetime = %s, master = %s, master_user_id = %s, name = %s, phone = %s, revenue = %s, source = %s
                WHERE id = %s
                """,
                (new_service, new_datetime, new_master, master_user_id, new_name, new_phone, new_revenue, new_source, booking_id),
            )
        else:
            c.execute(
                """
                UPDATE bookings
                SET service_name = %s, datetime = %s, master = %s, name = %s, phone = %s, revenue = %s, source = %s
                WHERE id = %s
                """,
                (new_service, new_datetime, new_master, new_name, new_phone, new_revenue, new_source, booking_id),
            )

        conn.commit()
        conn.close()

        log_activity(user["id"], "update_booking", "booking", str(booking_id),
                    f"Updated booking: {new_service}")

        # Отправляем уведомление мастеру об изменении
        if new_master:
            try:
                notification_results = await notify_master_about_booking(
                    master_name=new_master,
                    client_name=new_name,
                    service=new_service,
                    datetime_str=new_datetime,
                    phone=new_phone,
                    booking_id=booking_id,
                    notification_type="booking_change"
                )

                # Сохраняем логи
                master_info = get_master_info(new_master)
                if master_info:
                    for notif_type, success in notification_results.items():
                        if success or master_info.get(notif_type) or master_info.get(f"{notif_type}_username"):
                            save_notification_log(
                                master_id=master_info["id"],
                                booking_id=booking_id,
                                notification_type=notif_type,
                                status="sent" if success else "failed"
                            )

            except Exception as e:
                log_error(f"Error sending booking change notification: {e}", "api")

        return {"success": True, "message": "Booking updated", "booking_id": booking_id}

    except Exception as e:
        log_error(f"Booking update error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.delete("/bookings/{booking_id}")
async def delete_booking_api(
    booking_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить запись (с уведомлением мастеру)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
        
    if user["role"] != "director":
        return JSONResponse({"error": "Forbidden: Only Director can delete bookings"}, status_code=403)

    try:
        # Получаем информацию о записи для уведомления
        conn = get_db_connection()
        c = conn.cursor()

        c.execute("SELECT service_name, datetime, master, name, phone FROM bookings WHERE id = %s",
                  (booking_id,))
        booking = c.fetchone()

        if not booking:
            conn.close()
            return JSONResponse({"error": "Booking not found"}, status_code=404)

        service, datetime_str, master, name, phone = booking

        # Отправляем уведомление мастеру об отмене ПЕРЕД удалением
        # (чтобы можно было сохранить лог уведомления)
        if master:
            try:
                notification_results = await notify_master_about_booking(
                    master_name=master,
                    client_name=name,
                    service=service,
                    datetime_str=datetime_str,
                    phone=phone,
                    booking_id=booking_id,
                    notification_type="booking_cancel"
                )

                # Сохраняем логи
                master_info = get_master_info(master)
                if master_info:
                    for notif_type, success in notification_results.items():
                        if success or master_info.get(notif_type) or master_info.get(f"{notif_type}_username"):
                            save_notification_log(
                                master_id=master_info["id"],
                                booking_id=booking_id,
                                notification_type=notif_type,
                                status="sent" if success else "failed"
                            )

            except Exception as e:
                log_error(f"Error sending booking cancel notification: {e}", "api")

        # ✅ Используем Soft Delete вместо физического удаления
        from utils.soft_delete import soft_delete_booking
        from utils.audit import log_audit
        
        success = soft_delete_booking(booking_id, user)
        
        if success:
            # Логируем в аудит
            log_audit(
                user=user,
                action='delete',
                entity_type='booking',
                entity_id=str(booking_id),
                old_value={
                    "service": service,
                    "datetime": datetime_str,
                    "master": master,
                    "client": name
                }
            )
            return {"success": True, "message": "Booking moved to trash"}
        else:
            return JSONResponse({"error": "Booking not found or already deleted"}, status_code=404)


    except Exception as e:
        log_error(f"Booking deletion error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)
