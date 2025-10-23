from fastapi import APIRouter, Request, Query, Cookie, HTTPException, Form
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional
from datetime import datetime
import csv
import io
import time
import sqlite3

# ===== ИМПОРТЫ =====
from logger import logger, log_error, log_warning
from database import (
    get_all_clients, get_client_by_id, get_all_bookings, 
    update_client_info, update_client_status, pin_client,
    get_chat_history, mark_messages_as_read, save_message,
     get_stats, get_analytics_data, get_funnel_data,
    get_all_services, update_booking_status, log_activity,
    get_unread_messages_count, get_or_create_client, save_booking,
    get_all_users, delete_user, 
    get_custom_statuses, create_service, update_service, delete_service,
    delete_client, get_user_by_session, DATABASE_NAME,
    get_all_special_packages, create_special_package, 
    update_special_package, delete_special_package,
    get_all_roles, create_custom_role, delete_custom_role,
    get_role_permissions, update_role_permissions, AVAILABLE_PERMISSIONS,
    get_salon_settings, update_salon_settings,  # ✅ ДОБАВЬТЕ
    get_bot_settings, update_bot_settings,
)
from config import CLIENT_STATUSES
from integrations import send_message


# ===== ИМПОРТЫ ДЛЯ PDF И EXCEL =====
try:
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    log_warning("ReportLab не установлен. Экспорт в PDF недоступен.", "api")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    log_warning("openpyxl не установлен. Экспорт в Excel недоступен.", "api")

router = APIRouter(prefix="/api", tags=["API"])

# ===== MIDDLEWARE АВТОРИЗАЦИИ =====
def require_auth(session_token: Optional[str] = Cookie(None)):
    """Проверить авторизацию"""
    if not session_token:
        return None
    user = get_user_by_session(session_token)
    return user if user else None

def get_total_unread():
    """Получить общее кол-во непрочитанных"""
    clients = get_all_clients()
    return sum(get_unread_messages_count(c[0]) for c in clients)

def get_all_statuses():
    """Получить все статусы"""
    statuses = CLIENT_STATUSES.copy()
    for status in get_custom_statuses():
        statuses[status[1]] = {
            "label": status[2],
            "color": status[3],
            "icon": status[4]
        }
    return statuses

def get_client_display_name(client):
    """Форматировать имя клиента"""
    if client[3]:  # name
        return client[3]
    elif client[1]:  # username
        return f"@{client[1]}"
    else:
        return client[0][:15] + "..."

# ===== ФУНКЦИИ ЭКСПОРТА =====
def export_clients_csv(clients):
    """Экспорт клиентов в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['ID', 'Имя', 'Username', 'Телефон', 'Статус', 'Сообщений', 'LTV', 'Первый контакт', 'Последний контакт'])
    
    for c in clients:
        writer.writerow([
            c[0],  # id
            c[3] or '',  # name
            c[1] or '',  # username
            c[2] or '',  # phone
            c[8] if len(c) > 8 else 'new',  # status
            c[6],  # total_messages
            c[9] if len(c) > 9 else 0,  # lifetime_value
            c[4],  # first_contact
            c[5]   # last_contact
        ])
    
    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_clients_pdf(clients):
    """Экспорт клиентов в PDF"""
    if not PDF_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF экспорт недоступен")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#ec4899'),
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"База клиентов - {get_salon_settings['name']}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    date_text = Paragraph(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 20))
    
    data = [['ID', 'Имя', 'Телефон', 'Статус', 'Сообщений', 'LTV']]
    
    for c in clients:
        data.append([
            str(c[0])[:15],
            (c[3] or c[1] or 'N/A')[:25],
            (c[2] or '-')[:15],
            (c[8] if len(c) > 8 else 'new')[:10],
            str(c[6]),
            f"{c[9] if len(c) > 9 else 0} AED"
        ])
    
    table = Table(data, colWidths=[60, 120, 80, 60, 60, 80])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def export_clients_excel(clients):
    """Экспорт клиентов в Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel экспорт недоступен")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Клиенты"
    
    headers = ['ID', 'Имя', 'Username', 'Телефон', 'Статус', 'Сообщений', 'LTV (AED)', 'Первый контакт', 'Последний контакт']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for c in clients:
        ws.append([
            c[0],
            c[3] or '',
            c[1] or '',
            c[2] or '',
            c[8] if len(c) > 8 else 'new',
            c[6],
            c[9] if len(c) > 9 else 0,
            c[4],
            c[5]
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def export_bookings_csv(bookings):
    """Экспорт записей в CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    writer.writerow(['ID', 'Клиент', 'Услуга', 'Дата/Время', 'Телефон', 'Статус', 'Доход (AED)', 'Создана'])
    
    for b in bookings:
        writer.writerow([
            b[0],  # id
            b[5] or '',  # name
            b[2] or '',  # service_name
            b[3],  # datetime
            b[4] or '',  # phone
            b[6],  # status
            b[8] if len(b) > 8 else 0,  # revenue
            b[7]  # created_at
        ])
    
    output.seek(0)
    return output.getvalue().encode('utf-8')

def export_bookings_pdf(bookings):
    """Экспорт записей в PDF"""
    if not PDF_AVAILABLE:
        raise HTTPException(status_code=500, detail="PDF экспорт недоступен")
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#ec4899'),
        spaceAfter=30,
        alignment=1
    )
    
    title = Paragraph(f"Записи - {get_salon_settings['name']}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    date_text = Paragraph(f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal'])
    elements.append(date_text)
    elements.append(Spacer(1, 20))
    
    data = [['ID', 'Клиент', 'Услуга', 'Дата', 'Статус', 'Доход']]
    
    for b in bookings:
        try:
            date_obj = datetime.fromisoformat(b[3])
            date_str = date_obj.strftime('%d.%m %H:%M')
        except:
            date_str = str(b[3])[:16]
            
        data.append([
            str(b[0]),
            (b[5] or 'N/A')[:20],
            (b[2] or 'N/A')[:25],
            date_str,
            (b[6] or '')[:10],
            f"{b[8] if len(b) > 8 else 0} AED"
        ])
    
    table = Table(data, colWidths=[30, 80, 100, 70, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ec4899')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def export_bookings_excel(bookings):
    """Экспорт записей в Excel"""
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel экспорт недоступен")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Записи"
    
    headers = ['ID', 'Клиент', 'Услуга', 'Дата/Время', 'Телефон', 'Статус', 'Доход (AED)', 'Создана']
    ws.append(headers)
    
    header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for b in bookings:
        ws.append([
            b[0],
            b[5] or '',
            b[2] or '',
            b[3],
            b[4] or '',
            b[6],
            b[8] if len(b) > 8 else 0,
            b[7]
        ])
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ===== ОСНОВНЫЕ API ENDPOINTS =====
@router.get("/dashboard")
async def get_dashboard(session_token: Optional[str] = Cookie(None)):
    """Получить данные дашборда"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    stats = get_stats()
    analytics = get_analytics_data()
    funnel = get_funnel_data()
    
    return {
        "stats": stats,
        "analytics": analytics,
        "funnel": funnel,
        "unread_count": get_total_unread()
    }

# ===== КЛИЕНТЫ =====
@router.post("/clients")
async def create_client_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Создать нового клиента вручную"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    try:
        instagram_id = data.get('instagram_id') or f"manual_{int(time.time())}"
        get_or_create_client(instagram_id, username=data.get('name'))
        
        if data.get('phone') or data.get('notes') or data.get('name'):
            update_client_info(
                instagram_id,
                name=data.get('name'),
                phone=data.get('phone'),
                notes=data.get('notes')
            )
        
        log_activity(user["id"], "create_client", "client", instagram_id, f"Client: {data.get('name')}")
        return {"success": True, "message": "Client created", "id": instagram_id}
    except Exception as e:
        log_error(f"Error creating client: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.get("/clients")
async def list_clients(session_token: Optional[str] = Cookie(None)):
    """Получить всех клиентов"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    clients = get_all_clients()
    return {
        "clients": [
            {
                "id": c[0],
                "instagram_id": c[0],
                "username": c[1],
                "phone": c[2],
                "name": c[3],
                "display_name": get_client_display_name(c),
                "first_contact": c[4],
                "last_contact": c[5],
                "total_messages": c[6],
                "status": c[8] if len(c) > 8 else "new",
                "lifetime_value": c[9] if len(c) > 9 else 0,
                "profile_pic": c[10] if len(c) > 10 else None,
                "notes": c[11] if len(c) > 11 else "",
                "is_pinned": c[12] if len(c) > 12 else 0
            }
            for c in clients
        ],
        "count": len(clients),
        "unread_count": get_total_unread()
    }

@router.get("/clients/{client_id}")
async def get_client_detail(client_id: str, session_token: Optional[str] = Cookie(None)):
    """Получить деталь клиента"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    client = get_client_by_id(client_id)
    if not client:
        return JSONResponse({"error": "Client not found"}, status_code=404)
    
    history = get_chat_history(client_id, limit=50)
    bookings = [b for b in get_all_bookings() if b[1] == client_id]
    
    return {
        "client": {
            "id": client[0],
            "username": client[1],
            "phone": client[2],
            "name": client[3],
            "first_contact": client[4],
            "last_contact": client[5],
            "total_messages": client[6],
            "status": client[8] if len(client) > 8 else "new",
            "lifetime_value": client[9] if len(client) > 9 else 0,
            "notes": client[11] if len(client) > 11 else ""
        },
        "chat_history": [
            {
                "message": msg[0],
                "sender": msg[1],
                "timestamp": msg[2],
                "type": msg[3] if len(msg) > 3 else "text"
            }
            for msg in history
        ],
        "bookings": [
            {
                "id": b[0],
                "service": b[2],
                "datetime": b[3],
                "phone": b[4],
                "status": b[6],
                "revenue": b[8] if len(b) > 8 else 0
            }
            for b in bookings
        ]
    }

@router.post("/clients/{client_id}/delete")
async def delete_client_api(
    client_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить клиента"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    try:
        success = delete_client(client_id)
        if success:
            log_activity(user["id"], "delete_client", "client", 
                        client_id, "Client deleted")
            return {"success": True, "message": "Client deleted"}
        else:
            return JSONResponse({"error": "Client not found"}, 
                              status_code=404)
    except Exception as e:
        log_error(f"Error deleting client: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.post("/clients/{client_id}/status")
async def update_client_status_api(
    client_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Изменить статус клиента"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    data = await request.json()
    status = data.get('status')
    
    if not status:
        return JSONResponse({"error": "Status required"}, status_code=400)
    
    update_client_status(client_id, status)
    log_activity(user["id"], "update_client_status", "client", client_id, f"Status: {status}")
    
    return {"success": True, "message": "Client status updated"}

@router.post("/clients/{client_id}/update")
async def update_client_api(
    client_id: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить информацию клиента"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    data = await request.json()
    success = update_client_info(
        client_id,
        name=data.get('name'),
        phone=data.get('phone'),
        notes=data.get('notes')
    )
    
    if success:
        log_activity(user["id"], "update_client_info", "client", client_id, "Info updated")
        return {"success": True, "message": "Client updated"}
    
    return JSONResponse({"error": "Update failed"}, status_code=400)

@router.post("/clients/{client_id}/pin")
async def pin_client_api(
    client_id: str,
    session_token: Optional[str] = Cookie(None)
):
    """Закрепить/открепить клиента"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    client = get_client_by_id(client_id)
    if not client:
        return JSONResponse({"error": "Client not found"}, status_code=404)
    
    is_pinned = client[12] if len(client) > 12 else 0
    pin_client(client_id, not is_pinned)
    
    log_activity(user["id"], "pin_client", "client", client_id, 
                 f"{'Pinned' if not is_pinned else 'Unpinned'}")
    
    return {
        "success": True,
        "pinned": not is_pinned,
        "message": "Pinned" if not is_pinned else "Unpinned"
    }

# ===== ЗАПИСИ =====
@router.get("/bookings")
async def list_bookings(session_token: Optional[str] = Cookie(None)):
    """Получить все записи"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    bookings = get_all_bookings()
    return {
        "bookings": [
            {
                "id": b[0],
                "client_id": b[1],
                "service": b[2],
                "datetime": b[3],
                "phone": b[4],
                "name": b[5],
                "status": b[6],
                "created_at": b[7],
                "revenue": b[8] if len(b) > 8 else 0
            }
            for b in bookings
        ],
        "count": len(bookings)
    }

@router.post("/bookings")
async def create_booking_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Создать запись"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    data = await request.json()
    
    try:
        instagram_id = data.get('instagram_id')
        service = data.get('service')
        datetime_str = f"{data.get('date')} {data.get('time')}"
        phone = data.get('phone', '')
        name = data.get('name')
        
        get_or_create_client(instagram_id, username=name)
        save_booking(instagram_id, service, datetime_str, phone, name)
        
        log_activity(user["id"], "create_booking", "booking", instagram_id, f"Service: {service}")
        
        return {"success": True, "message": "Booking created"}
    except Exception as e:
        log_error(f"Booking creation error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.post("/bookings/{booking_id}/status")
async def update_booking_status_api(
    booking_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Изменить статус записи"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    data = await request.json()
    status = data.get('status')
    
    if not status:
        return JSONResponse({"error": "Status required"}, status_code=400)
    
    success = update_booking_status(booking_id, status)
    if success:
        log_activity(user["id"], "update_booking_status", "booking", str(booking_id), f"Status: {status}")
        return {"success": True, "message": "Booking status updated"}
    
    return JSONResponse({"error": "Update failed"}, status_code=400)

@router.get("/bookings/{booking_id}")
async def get_booking_detail(
    booking_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Получить одну запись по ID"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    bookings = get_all_bookings()
    booking = next((b for b in bookings if b[0] == booking_id), None)
    
    if not booking:
        return JSONResponse({"error": "Booking not found"}, status_code=404)
    
    return {
        "id": booking[0],
        "client_id": booking[1],
        "service": booking[2],
        "datetime": booking[3],
        "phone": booking[4],
        "name": booking[5],
        "status": booking[6],
        "created_at": booking[7],
        "revenue": booking[8] if len(booking) > 8 else 0
    }

# ===== ЧАТ =====
@router.get("/chat/messages")
async def get_chat_messages(
    client_id: str = Query(...),
    limit: int = Query(50),
    session_token: Optional[str] = Cookie(None)
):
    """Получить сообщения чата"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    messages_raw = get_chat_history(client_id, limit=limit)
    mark_messages_as_read(client_id, user["id"])
    
    return {
        "messages": [
            {
                "id": msg[4] if len(msg) > 4 else None,
                "message": msg[0],
                "sender": msg[1],
                "timestamp": msg[2],
                "type": msg[3] if len(msg) > 3 else "text"
            }
            for msg in messages_raw
        ]
    }

@router.post("/chat/send")
async def send_chat_message(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Отправить сообщение клиенту"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    data = await request.json()
    instagram_id = data.get('instagram_id')
    message = data.get('message')
    
    if not instagram_id or not message:
        return JSONResponse({"error": "Missing data"}, status_code=400)
    
    result = await send_message(instagram_id, message)
    
    if "error" not in result:
        save_message(instagram_id, message, "bot")
        log_activity(user["id"], "send_message", "client", instagram_id, f"Message sent")
        return {"success": True, "message": "Message sent"}
    
    return JSONResponse({"error": "Send failed"}, status_code=500)

# ===== АНАЛИТИКА =====
@router.get("/analytics")
async def get_analytics_api(
    period: int = Query(30),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """Получить аналитику"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    if date_from and date_to:
        return get_analytics_data(date_from=date_from, date_to=date_to)
    else:
        return get_analytics_data(days=period)

@router.get("/funnel")
async def get_funnel_api(session_token: Optional[str] = Cookie(None)):
    """Получить данные воронки"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    return get_funnel_data()

@router.get("/stats")
async def get_stats_api(session_token: Optional[str] = Cookie(None)):
    """Получить статистику"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    return get_stats()

# ===== УСЛУГИ =====
@router.get("/services")
async def list_services(
    active_only: bool = Query(True),
    session_token: Optional[str] = Cookie(None)
):
    """Получить услуги"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    services = get_all_services(active_only=active_only)
    
    return {
        "services": [
            {
                "id": s[0],
                "key": s[1],
                "name": s[2],
                "name_ru": s[3] if len(s) > 3 else s[2],
                "price": s[5] if len(s) > 5 else 0,
                "currency": s[6] if len(s) > 6 else "AED",
                "category": s[7] if len(s) > 7 else "other",
                "description": s[8] if len(s) > 8 else "",
                "is_active": s[12] if len(s) > 12 else True
            }
            for s in services
        ],
        "count": len(services)
    }

@router.post("/services")
async def create_service_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Создать новую услугу"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    try:
        success = create_service(
            service_key=data.get('key'),
            name=data.get('name'),
            name_ru=data.get('name_ru'),
            price=float(data.get('price', 0)),
            currency=data.get('currency', 'AED'),
            category=data.get('category'),
            description=data.get('description'),
            description_ru=data.get('description_ru'),
            benefits=data.get('benefits', [])
        )
        
        if success:
            log_activity(user["id"], "create_service", "service", data.get('key'), "Service created")
            return {"success": True, "message": "Service created"}
        else:
            return JSONResponse({"error": "Service key already exists"}, status_code=400)
    except Exception as e:
        log_error(f"Error creating service: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.post("/services/{service_id}/update")
async def update_service_api(
    service_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить услугу"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    try:
        update_service(service_id, **data)
        log_activity(user["id"], "update_service", "service", str(service_id), "Service updated")
        return {"success": True, "message": "Service updated"}
    except Exception as e:
        log_error(f"Error updating service: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.post("/services/{service_id}/delete")
async def delete_service_api(
    service_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить услугу"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    try:
        delete_service(service_id)
        log_activity(user["id"], "delete_service", "service", str(service_id), "Service deleted")
        return {"success": True, "message": "Service deleted"}
    except Exception as e:
        log_error(f"Error deleting service: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

# ===== СПЕЦИАЛЬНЫЕ ПАКЕТЫ =====
@router.get("/special-packages")
async def list_special_packages(
    active_only: bool = Query(True),
    session_token: Optional[str] = Cookie(None)
):
    """Получить специальные пакеты"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    packages = get_all_special_packages(active_only=active_only)
    
    return {
        "packages": [
            {
                "id": p[0],
                "name": p[1],
                "name_ru": p[2],
                "description": p[3],
                "description_ru": p[4],
                "original_price": p[5],
                "special_price": p[6],
                "currency": p[7],
                "discount_percent": p[8],
                "services_included": p[9].split(',') if p[9] else [],
                "promo_code": p[10],
                "keywords": p[11].split(',') if p[11] else [],
                "valid_from": p[12],
                "valid_until": p[13],
                "is_active": p[14],
                "usage_count": p[15],
                "max_usage": p[16]
            }
            for p in packages
        ],
        "count": len(packages)
    }

@router.post("/special-packages")
async def create_special_package_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Создать специальный пакет"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    try:
        package_id = create_special_package(
            name=data.get('name'),
            name_ru=data.get('name_ru'),
            original_price=float(data.get('original_price')),
            special_price=float(data.get('special_price')),
            currency=data.get('currency', 'AED'),
            keywords=data.get('keywords', []),
            valid_from=data.get('valid_from'),
            valid_until=data.get('valid_until'),
            description=data.get('description'),
            description_ru=data.get('description_ru'),
            services_included=data.get('services_included', []),
            promo_code=data.get('promo_code'),
            max_usage=data.get('max_usage')
        )
        
        if package_id:
            log_activity(user["id"], "create_special_package", "package", str(package_id), "Package created")
            return {"success": True, "message": "Package created", "id": package_id}
        else:
            return JSONResponse({"error": "Failed to create package"}, status_code=400)
    except Exception as e:
        log_error(f"Error creating package: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.post("/special-packages/{package_id}")
async def update_special_package_api(
    package_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить специальный пакет"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    try:
        update_special_package(package_id, **data)
        log_activity(user["id"], "update_special_package", "package", str(package_id), "Package updated")
        return {"success": True, "message": "Package updated"}
    except Exception as e:
        log_error(f"Error updating package: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

@router.delete("/special-packages/{package_id}")
async def delete_special_package_api(
    package_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить специальный пакет"""
    user = require_auth(session_token)
    if not user or user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    try:
        delete_special_package(package_id)
        log_activity(user["id"], "delete_special_package", "package", str(package_id), "Package deleted")
        return {"success": True, "message": "Package deleted"}
    except Exception as e:
        log_error(f"Error deleting package: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=400)

# ===== ПОЛЬЗОВАТЕЛИ =====
@router.get("/users")
async def list_users(session_token: Optional[str] = Cookie(None)):
    """Получить пользователей (только для admin)"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    users = get_all_users()
    
    return {
        "users": [
            {
                "id": u[0],
                "username": u[1],
                "full_name": u[3],
                "email": u[4],
                "role": u[5],
                "created_at": u[6]
            }
            for u in users
        ],
        "count": len(users)
    }

@router.post("/users/{user_id}/delete")
async def delete_user_api(
    user_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить пользователя"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    if user["id"] == user_id:
        return JSONResponse({"error": "Cannot delete yourself"}, status_code=400)
    
    success = delete_user(user_id)
    
    if success:
        log_activity(user["id"], "delete_user", "user", str(user_id), "User deleted")
        return {"success": True, "message": "User deleted"}
    
    return JSONResponse({"error": "Delete failed"}, status_code=400)

# ===== ЭКСПОРТ (ВСЕ ФОРМАТЫ) =====
@router.get("/export/clients")
async def export_clients(
    format: str = Query("csv"),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать клиентов в CSV, PDF или Excel"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    clients = get_all_clients()
    
    try:
        if format == "csv":
            content = export_clients_csv(clients)
            media_type = "text/csv"
            filename = f"clients_{datetime.now().strftime('%Y%m%d')}.csv"
        elif format == "pdf":
            content = export_clients_pdf(clients)
            media_type = "application/pdf"
            filename = f"clients_{datetime.now().strftime('%Y%m%d')}.pdf"
        elif format == "excel":
            content = export_clients_excel(clients)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"clients_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/bookings")
async def export_bookings(
    format: str = Query("csv"),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать записи в CSV, PDF или Excel с фильтрацией по дате"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    all_bookings = get_all_bookings()
    
    if date_from and date_to:
        filtered_bookings = []
        for b in all_bookings:
            try:
                booking_date = datetime.fromisoformat(b[3])
                start_date = datetime.fromisoformat(date_from)
                end_date = datetime.fromisoformat(date_to)
                
                if start_date <= booking_date <= end_date:
                    filtered_bookings.append(b)
            except:
                continue
        bookings = filtered_bookings
    else:
        bookings = all_bookings
    
    try:
        if format == "csv":
            content = export_bookings_csv(bookings)
            media_type = "text/csv"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.csv"
        elif format == "pdf":
            content = export_bookings_pdf(bookings)
            media_type = "application/pdf"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.pdf"
        elif format == "excel":
            content = export_bookings_excel(bookings)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"bookings_{datetime.now().strftime('%Y%m%d')}.xlsx"
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/export/analytics")
async def export_analytics(
    format: str = Query("csv"),
    period: int = Query(30),
    date_from: str = Query(None),
    date_to: str = Query(None),
    session_token: Optional[str] = Cookie(None)
):
    """Экспортировать аналитику в CSV, PDF или Excel"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    if date_from and date_to:
        analytics = get_analytics_data(date_from=date_from, date_to=date_to)
    else:
        analytics = get_analytics_data(days=period)
    
    try:
        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            
            writer.writerow(['Дата', 'Записей'])
            for date, count in analytics.get('bookings_by_day', []):
                writer.writerow([date, count])
            
            writer.writerow([])
            writer.writerow(['Услуга', 'Количество', 'Доход'])
            for name, count, revenue in analytics.get('services_stats', []):
                writer.writerow([name, count, revenue])
            
            output.seek(0)
            content = output.getvalue().encode('utf-8')
            media_type = "text/csv"
            filename = f"analytics_{datetime.now().strftime('%Y%m%d')}.csv"
        
        elif format == "pdf":
            if not PDF_AVAILABLE:
                return JSONResponse({"error": "PDF export not available"}, status_code=500)
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            
            styles = getSampleStyleSheet()
            title = Paragraph("Аналитика", styles['Heading1'])
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            data = [['Дата', 'Записей']]
            for date, count in analytics.get('bookings_by_day', []):
                data.append([date, str(count)])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            doc.build(elements)
            buffer.seek(0)
            content = buffer.getvalue()
            media_type = "application/pdf"
            filename = f"analytics_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        elif format == "excel":
            if not EXCEL_AVAILABLE:
                return JSONResponse({"error": "Excel export not available"}, status_code=500)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Аналитика"
            
            ws.append(['Дата', 'Записей'])
            for date, count in analytics.get('bookings_by_day', []):
                ws.append([date, count])
            
            ws.append([])
            ws.append(['Услуга', 'Количество', 'Доход'])
            for name, count, revenue in analytics.get('services_stats', []):
                ws.append([name, count, revenue])
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            content = buffer.getvalue()
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"analytics_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        else:
            return JSONResponse({"error": "Invalid format"}, status_code=400)
        
        return StreamingResponse(
            iter([content]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        log_error(f"Export analytics error: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.get("/unread-count")
async def get_unread_count(session_token: Optional[str] = Cookie(None)):
    """Получить количество непрочитанных"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    return {"count": get_total_unread()}

# ===== НАСТРОЙКИ БОТА =====
@router.get("/bot-settings")
async def get_bot_settings_api(session_token: Optional[str] = Cookie(None)):
    """Получить настройки бота (из БД)"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    # ✅ ИСПРАВЛЕНО: admin И manager могут просматривать настройки
    if user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    try:
        settings = get_bot_settings()
        return settings
    except Exception as e:
        log_error(f"Ошибка получения настроек бота: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)

@router.post("/bot-settings")
async def update_bot_settings_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить настройки бота"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    # ✅ ИСПРАВЛЕНО: admin И manager могут редактировать
    if user["role"] not in ["admin", "manager"]:
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    success = update_bot_settings(data)
    
    if success:
        log_activity(user["id"], "update_bot_settings", "bot", "general", "Bot settings updated")
        return {"success": True, "message": "Bot settings updated"}
    else:
        return JSONResponse({"error": "Update failed"}, status_code=500)

# ===== НАСТРОЙКИ САЛОНА =====
@router.get("/salon-settings")
async def get_salon_settings_api(session_token: Optional[str] = Cookie(None)):
    """Получить настройки салона"""
    user = require_auth(session_token)
    if not user:
        return JSONResponse({"error": "Unauthorized"}, status_code=401)
    
    try:
        settings = get_salon_settings()
        return settings
    except Exception as e:
        log_error(f"Ошибка получения настроек салона: {e}", "api")
        return JSONResponse(
            {"error": str(e), "message": "Запустите init_settings.py"}, 
            status_code=500
        )


@router.post("/salon-settings")
async def update_salon_settings_api(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить настройки салона (только admin)"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    success = update_salon_settings(data)
    
    if success:
        log_activity(user["id"], "update_salon_settings", "settings", "salon", "Salon settings updated")
        return {"success": True, "message": "Salon settings updated"}
    else:
        return JSONResponse({"error": "Update failed"}, status_code=400)


# ===== РОЛИ И ПРАВА ДОСТУПА =====
@router.get("/roles")
async def list_roles(session_token: Optional[str] = Cookie(None)):
    """Получить все роли"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    roles = get_all_roles()
    
    return {
        "roles": roles,
        "count": len(roles)
    }

@router.post("/roles")
async def create_role(
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Создать кастомную роль"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    
    if not data.get('role_key') or not data.get('role_name'):
        return JSONResponse({"error": "Missing required fields"}, status_code=400)
    
    success = create_custom_role(
        data['role_key'],
        data['role_name'],
        data.get('role_description', ''),
        user["id"]
    )
    
    if success:
        log_activity(user["id"], "create_role", "role", data['role_key'], "Role created")
        return {"success": True, "message": "Role created"}
    else:
        return JSONResponse({"error": "Role already exists"}, status_code=400)

@router.delete("/roles/{role_key}")
async def delete_role(
    role_key: str,
    session_token: Optional[str] = Cookie(None)
):
    """Удалить кастомную роль"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    success = delete_custom_role(role_key)
    
    if success:
        log_activity(user["id"], "delete_role", "role", role_key, "Role deleted")
        return {"success": True, "message": "Role deleted"}
    else:
        return JSONResponse({"error": "Cannot delete built-in roles"}, status_code=400)

@router.get("/roles/{role_key}/permissions")
async def get_role_permissions_api(
    role_key: str,
    session_token: Optional[str] = Cookie(None)
):
    """Получить права роли"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    permissions = get_role_permissions(role_key)
    
    return {
        "role_key": role_key,
        "permissions": permissions,
        "available_permissions": AVAILABLE_PERMISSIONS
    }

@router.post("/roles/{role_key}/permissions")
async def update_role_permissions_api(
    role_key: str,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Обновить права роли"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    data = await request.json()
    permissions = data.get('permissions', {})
    
    success = update_role_permissions(role_key, permissions)
    
    if success:
        log_activity(user["id"], "update_permissions", "role", role_key, "Permissions updated")
        return {"success": True, "message": "Permissions updated"}
    else:
        return JSONResponse({"error": "Update failed"}, status_code=400)

@router.get("/permissions/available")
async def list_available_permissions(session_token: Optional[str] = Cookie(None)):
    """Получить список всех доступных прав"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    return {
        "permissions": [
            {"key": key, "name": name}
            for key, name in AVAILABLE_PERMISSIONS.items()
        ]
    }

@router.get("/users/{user_id}/permissions")
async def get_user_permissions(
    user_id: int,
    session_token: Optional[str] = Cookie(None)
):
    """Получить права конкретного пользователя"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()
    
    c.execute("SELECT role FROM users WHERE id = ?", (user_id,))
    result = c.fetchone()
    
    if not result:
        conn.close()
        return JSONResponse({"error": "User not found"}, status_code=404)
    
    role = result[0]
    conn.close()
    
    permissions = get_role_permissions(role)
    
    return {
        "user_id": user_id,
        "role": role,
        "permissions": permissions
    }

@router.post("/users/{user_id}/role")
async def update_user_role(
    user_id: int,
    request: Request,
    session_token: Optional[str] = Cookie(None)
):
    """Изменить роль пользователя"""
    user = require_auth(session_token)
    if not user or user["role"] != "admin":
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    
    if user["id"] == user_id:
        return JSONResponse({"error": "Cannot change your own role"}, status_code=400)
    
    data = await request.json()
    new_role = data.get('role')
    
    if not new_role:
        return JSONResponse({"error": "Role required"}, status_code=400)
    
    conn = sqlite3.connect(DATABASE_NAME)
    c = conn.cursor()
    
    try:
        c.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, user_id))
        conn.commit()
        
        if c.rowcount > 0:
            log_activity(user["id"], "update_user_role", "user", str(user_id), f"Role: {new_role}")
            conn.close()
            return {"success": True, "message": "Role updated"}
        else:
            conn.close()
            return JSONResponse({"error": "User not found"}, status_code=404)
    except Exception as e:
        conn.rollback()
        conn.close()
        log_error(f"Error updating user role: {e}", "api")
        return JSONResponse({"error": str(e)}, status_code=500)